"""
政策汇编生成模块 v3
- 国家级政策一个Sheet，福建省政策（含地市）一个Sheet
- 每个Sheet内按14个政策类型分组
- 核心要点用AI总结
- 政策唯一ID = 发文编号（无发文编号则用日期+标题前20字）
"""

import json, re, logging
from datetime import datetime
from pathlib import Path
from typing import Optional
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent
OUTPUT_DIR = PROJECT_ROOT / "output"
DATA_DIR = OUTPUT_DIR / "data"

# 14个政策类型及其关键词匹配规则
POLICY_CATEGORIES = OrderedDict([
    ("新能源", ["新能源", "风电", "光伏", "可再生能源", "分布式能源", "生物质能", "地热能"]),
    ("光伏", ["光伏", "太阳能", "光电"]),
    ("储能", ["储能", "新型储能", "抽水蓄能", "电化学储能", "压缩空气储能"]),
    ("充电桩/站", ["充电桩", "充电站", "充电基础设施", "充换电"]),
    ("虚拟电厂", ["虚拟电厂", "需求侧响应", "负荷聚合"]),
    ("零碳园区", ["零碳园区", "零碳", "低碳园区", "近零碳"]),
    ("微电网", ["微电网", "微网", "源网荷储"]),
    ("综合能源服务", ["综合能源服务", "综合能源", "多能互补", "能源互联网"]),
    ("节能降碳", ["节能降碳", "碳达峰", "碳中和", "能耗双控", "碳排放"]),
    ("节能改造", ["节能改造", "节能诊断", "节能审查", "能效提升", "设备更新"]),
    ("能源托管", ["能源托管", "合同能源管理"]),
    ("电力运维服务", ["电力运维", "配电运维", "运维服务"]),
    ("售电", ["售电", "售电公司", "电力零售"]),
    ("电改", ["电改", "电力体制改革", "电力市场", "输配电价", "电力交易", "现货市场", "绿电", "绿证", "上网电价"]),
])


def _classify_policy(title, summary):
    """根据标题和摘要分类到14个政策类型"""
    text = (title + ' ' + summary).lower()
    matched = []
    for cat, keywords in POLICY_CATEGORIES.items():
        if any(kw.lower() in text for kw in keywords):
            matched.append(cat)
    return matched if matched else ["综合"]


def _generate_policy_id(policy):
    """生成政策唯一ID：优先用发文编号，否则用日期+标题前20字"""
    wh = policy.get('发文编号', '').strip()
    if wh and len(wh) >= 6:
        return wh
    date = policy.get('发布日期', '').strip()
    title = policy.get('文件标题', '').strip()
    if date:
        return f"{date}_{title[:30]}"
    return title[:40]


def _summarize_core_points(policy):
    """对核心要点进行智能总结"""
    raw = policy.get('核心要点', '')
    title = policy.get('文件标题', '')

    # 如果原文太短或没有内容，用标题
    if not raw or len(raw) < 30:
        # 尝试从标题提取要点
        title_clean = re.sub(r'[（(].*?[）)]', '', title)  # 去掉括号内容
        title_clean = re.sub(r'(国家发展改革委|国家能源局|国务院|福建省|等部门|关于印发|的通知|的通知$|意见$|办法$)', '', title_clean)
        return title_clean.strip()[:200] if title_clean.strip() else title[:200]

    # 清理
    cleaned = re.sub(r'(目录项的基本信息|索引号|主办单位|制发日期|索\s*引\s*号)[\s：:0-9A-Za-z/\-]+', '', raw)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    # 提取关键句子
    sentences = re.split(r'[。；;]', cleaned)
    key_points = []
    key_signals = ['为', '要', '应', '将', '推进', '推动', '加快', '促进', '加强', '完善', '建立',
                   '发展', '目标', '任务', '措施', '要求', '重点', '落实', '实施', '支持',
                   '鼓励', '规范', '管理', '机制', '体系', '标准', '市场', '改革', '创新']

    for s in sentences:
        s = s.strip()
        if not s or len(s) < 10:
            continue
        if any(kw in s for kw in key_signals):
            key_points.append(s)
        elif len(key_points) < 3:
            key_points.append(s)

    if not key_points:
        return cleaned[:200]

    summary = '；'.join(key_points[:5])
    if len(summary) > 250:
        summary = summary[:250] + '…'
    return summary


def generate_assembly(results=None):
    """生成分类汇编Excel"""
    if results is None:
        latest = DATA_DIR / "policies_latest.json"
        if not latest.exists():
            logger.error("无采集结果")
            return ""
        with open(latest, 'r', encoding='utf-8') as f:
            results = json.load(f)

    if not results:
        return ""

    # 按层级分类：国家级 vs 福建省（含地市）
    national = []
    fujian = []
    for r in results:
        level = r.get('内容层级', '')
        if '福建' in level or '地市' in level:
            fujian.append(r)
        else:
            national.append(r)

    # 对每个政策重新分类（14类）
    for r in national + fujian:
        cats = _classify_policy(r.get('文件标题', ''), r.get('核心要点', ''))
        r['_categories'] = cats
        r['_policy_id'] = _generate_policy_id(r)
        r['_summary'] = _summarize_core_points(r)

    wb = openpyxl.Workbook()
    
    # 样式
    hdr_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cat_font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
    cat_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    cell_font = Font(name='微软雅黑', size=10)
    cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    headers = ['政策唯一ID(发文编号)', '文件标题', '发布单位', '发布日期', '政策类型', 
               '核心要点总结', '原文链接', '附件PDF']

    def write_sheet(ws, policies, sheet_title):
        ws.title = sheet_title
        col_widths = [26, 50, 24, 13, 14, 55, 45, 35]
        row = 1

        # 按14个分类分组
        grouped = OrderedDict()
        for cat_name in POLICY_CATEGORIES:
            grouped[cat_name] = []
        grouped["综合"] = []

        for p in policies:
            for cat in p.get('_categories', ['综合']):
                if cat in grouped:
                    grouped[cat].append(p)
                else:
                    grouped["综合"].append(p)

        # 写入每个分组
        for cat_name, cat_policies in grouped.items():
            if not cat_policies:
                continue

            # 分类标题行
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            cell = ws.cell(row=row, column=1, value=f"▌ {cat_name}（{len(cat_policies)}条）")
            cell.font = cat_font
            cell.fill = cat_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin
            for c in range(2, len(headers)+1):
                ws.cell(row=row, column=c).fill = cat_fill
                ws.cell(row=row, column=c).border = thin
            ws.row_dimensions[row].height = 28
            row += 1

            # 表头行
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align
                cell.border = thin
            ws.row_dimensions[row].height = 30
            row += 1

            # 数据行
            for pi, p in enumerate(cat_policies):
                pdf_urls = p.get('附件PDF', [])
                pdf_str = '; '.join([x.get('url', '') for x in pdf_urls]) if pdf_urls else ''
                
                row_data = [
                    p.get('_policy_id', ''),
                    p.get('文件标题', ''),
                    p.get('发布单位', ''),
                    p.get('发布日期', ''),
                    '、'.join(p.get('_categories', ['综合'])),
                    p.get('_summary', ''),
                    p.get('原文链接', ''),
                    pdf_str,
                ]
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font = cell_font
                    cell.border = thin
                    cell.alignment = center_align if ci in (1, 3, 4, 5) else cell_align
                    if pi % 2 == 0:
                        cell.fill = alt_fill
                ws.row_dimensions[row].height = 50 if len(p.get('_summary', '')) > 100 else 28
                row += 1

            row += 1  # 分组间空行

        # 设置列宽
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # 冻结
        ws.freeze_panes = 'A2'

    # Sheet 1: 国家级政策
    ws1 = wb.active
    write_sheet(ws1, national, "国家级政策")

    # Sheet 2: 福建省政策
    ws2 = wb.create_sheet()
    write_sheet(ws2, fujian, "福建省政策")

    # Sheet 3: 统计
    ws3 = wb.create_sheet("分类统计")
    _write_stats(ws3, national, fujian)

    # 保存
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    fp = OUTPUT_DIR / f"政策汇编_{ts}.xlsx"
    wb.save(fp)
    latest = OUTPUT_DIR / "政策汇编_最新.xlsx"
    wb.save(latest)
    logger.info(f"汇编已生成: {fp}")
    return str(fp)


def _write_stats(ws, national, fujian):
    """写入统计Sheet"""
    hdr_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell_font = Font(name='微软雅黑', size=10)
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    hdr_align = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    ws.cell(row=1, column=1, value="政策分类统计").font = Font(name='微软雅黑', size=14, bold=True)
    ws.merge_cells('A1:D1')

    for ci, h in enumerate(['政策类型', '国家级', '福建省', '合计'], 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin

    row = 4
    for cat in POLICY_CATEGORIES:
        n_cnt = sum(1 for p in national if cat in p.get('_categories', []))
        f_cnt = sum(1 for p in fujian if cat in p.get('_categories', []))
        if n_cnt == 0 and f_cnt == 0:
            continue
        for ci, v in enumerate([cat, n_cnt, f_cnt, n_cnt + f_cnt], 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = cell_font
            cell.border = thin
            cell.alignment = hdr_align if ci > 1 else Alignment(horizontal='left', vertical='center')
        row += 1

    total_n = len(national)
    total_f = len(fujian)
    for ci, v in enumerate(['合计', total_n, total_f, total_n + total_f], 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.font = Font(name='微软雅黑', size=10, bold=True)
        cell.border = thin
        cell.alignment = hdr_align if ci > 1 else Alignment(horizontal='left', vertical='center')


def generate_assembly_markdown(results=None):
    """生成Markdown汇编"""
    if results is None:
        latest = DATA_DIR / "policies_latest.json"
        if not latest.exists():
            return ""
        with open(latest, 'r', encoding='utf-8') as f:
            results = json.load(f)

    national = [r for r in results if '福建' not in r.get('内容层级','') and '地市' not in r.get('内容层级','')]
    fujian = [r for r in results if '福建' in r.get('内容层级','') or '地市' in r.get('内容层级','')]

    lines = [
        "# 能源政策汇编",
        f"> 更新: {datetime.now():%Y-%m-%d %H:%M} | 国家级: {len(national)}条 | 福建省: {len(fujian)}条",
        "", "---", "",
    ]

    for section_name, policies in [("国家级政策", national), ("福建省政策", fujian)]:
        lines.append(f"## {section_name}（{len(policies)}条）")
        lines.append("")
        
        grouped = OrderedDict()
        for cat in POLICY_CATEGORIES:
            grouped[cat] = []
        grouped["综合"] = []
        for p in policies:
            for cat in _classify_policy(p.get('文件标题',''), p.get('核心要点','')):
                if cat in grouped:
                    grouped[cat].append(p)
                else:
                    grouped["综合"].append(p)

        for cat, cat_policies in grouped.items():
            if not cat_policies:
                continue
            lines.append(f"### {cat}（{len(cat_policies)}条）")
            lines.append("")
            lines.append("| 发文编号 | 标题 | 发布单位 | 日期 | 核心要点 |")
            lines.append("|----------|------|----------|------|----------|")
            for p in cat_policies:
                pid = _generate_policy_id(p)
                summary = _summarize_core_points(p)[:80]
                lines.append(
                    f"| {pid} | {p.get('文件标题','')[:50]} | {p.get('发布单位','')} "
                    f"| {p.get('发布日期','')} | {summary} |"
                )
            lines.append("")

    md_path = OUTPUT_DIR / "政策汇编_最新.md"
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    return str(md_path)


if __name__ == '__main__':
    p = generate_assembly()
    print(f"Excel: {p}")
