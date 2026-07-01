# -*- coding: utf-8 -*-
"""生成回测汇总表：所有用户在一个sheet里，预测/实际/偏差 三行一组"""
import pandas as pd, numpy as np, sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUTPUT_DIR = Path('1-2负荷预测输出')
VAL_DIR = OUTPUT_DIR / 'validation'

vdf = pd.read_csv(VAL_DIR / 'validation_hourly_long_v2_6.csv')
# 读取配置获取日期
cfg = pd.read_csv(OUTPUT_DIR / 'model' / 'run_config_v2_6.csv', encoding='utf-8-sig').iloc[0]
pred_date = pd.Timestamp(cfg['PREDICT_START']).normalize()

rows = []
for (uid, uname), g in vdf.groupby(['用户编号', '用户名称']):
    g = g.copy().sort_values('datetime')
    # 计算base_date
    g['dt'] = pd.to_datetime(g['datetime'])
    g['base_date'] = np.where(g['dt'].dt.hour==0,
                              (g['dt']-pd.Timedelta(days=1)).dt.normalize(),
                              g['dt'].dt.normalize())
    g['base_date'] = pd.to_datetime(g['base_date'])
    one_day = g[g['base_date']==pred_date]

    city = g['所在市'].iloc[0] if '所在市' in g.columns else ''
    pv_flag = g['是否有光伏_flag'].iloc[0] if '是否有光伏_flag' in g.columns else 0
    has_pv = '是' if pv_flag==1 else '否'

    pred_row = {'用户名称': uname, '所在市': city, '是否有光伏': has_pv, '类型': '预测'}
    act_row  = {'用户名称': uname, '所在市': city, '是否有光伏': has_pv, '类型': '实际'}
    err_row  = {'用户名称': uname, '所在市': city, '是否有光伏': has_pv, '类型': '偏差'}

    psum, asum = 0.0, 0.0
    pe, ae = True, True
    for h in range(1, 25):
        target = pred_date + pd.Timedelta(days=1) if h==24 else pred_date + pd.Timedelta(hours=h)
        hit = one_day[one_day['dt']==target]
        if hit.empty:
            pv, av = np.nan, np.nan
        else:
            pv = hit['final_pred_net_load'].iloc[0]/1000 if pd.notna(hit['final_pred_net_load'].iloc[0]) else np.nan
            av = hit['actual_load'].iloc[0]/1000 if 'actual_load' in hit.columns and pd.notna(hit['actual_load'].iloc[0]) else np.nan

        pred_row[f'{h}:00'] = round(pv,3) if pd.notna(pv) else np.nan
        act_row[f'{h}:00'] = round(av,3) if pd.notna(av) else np.nan
        if pd.notna(pv) and pd.notna(av) and av!=0:
            err_row[f'{h}:00'] = round((pv-av)/av, 4)
        else:
            err_row[f'{h}:00'] = np.nan

        if pd.notna(pv): psum += pv; pe = False
        if pd.notna(av): asum += av; ae = False

    pred_row['日合计(MWh)'] = round(psum,3) if not pe else np.nan
    act_row['日合计(MWh)']  = round(asum,3) if not ae else np.nan
    if not pe and not ae and asum>0:
        err_row['日合计(MWh)'] = round((psum-asum)/asum, 4)
    else:
        err_row['日合计(MWh)'] = np.nan

    rows.extend([pred_row, act_row, err_row])

cols = ['用户名称','所在市','是否有光伏','类型'] + [f'{h}:00' for h in range(1,25)] + ['日合计(MWh)']
result = pd.DataFrame(rows)[cols]

# === 添加24时段汇总行（预测/实际/偏差 各一行） ===
tot_pred = {'用户名称':'【全用户合计】','所在市':'','是否有光伏':'','类型':'预测'}
tot_act  = {'用户名称':'【全用户合计】','所在市':'','是否有光伏':'','类型':'实际'}
tot_err  = {'用户名称':'【全用户合计】','所在市':'','是否有光伏':'','类型':'偏差'}

pred_mask = result['类型']=='预测'
act_mask  = result['类型']=='实际'

for h in range(1,25):
    col = f'{h}:00'
    tp = result.loc[pred_mask, col].sum()
    ta = result.loc[act_mask, col].sum()
    tot_pred[col] = round(tp, 3)
    tot_act[col]  = round(ta, 3)
    if pd.notna(ta) and ta != 0:
        tot_err[col] = round((tp-ta)/ta, 4)
    else:
        tot_err[col] = np.nan

# 日合计
sum_p = result.loc[pred_mask, '日合计(MWh)'].sum()
sum_a = result.loc[act_mask, '日合计(MWh)'].sum()
tot_pred['日合计(MWh)'] = round(sum_p, 3)
tot_act['日合计(MWh)']  = round(sum_a, 3)
tot_err['日合计(MWh)']  = round((sum_p-sum_a)/sum_a, 4) if sum_a != 0 else np.nan

result = pd.concat([result, pd.DataFrame([tot_pred, tot_act, tot_err])], ignore_index=True)

# 写入Excel
out_path = OUTPUT_DIR / 'prediction' / '6月27日_回测对比汇总.xlsx'
with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    result.to_excel(writer, sheet_name='全用户对比', index=False)

    # 格式化
    from openpyxl.styles import PatternFill, Font, Alignment
    ws = writer.book['全用户对比']
    blue = PatternFill(fill_type='solid',fgColor='DDEBF7')
    green = PatternFill(fill_type='solid',fgColor='E2F0D9')
    yellow = PatternFill(fill_type='solid',fgColor='FFF2CC')
    red = PatternFill(fill_type='solid',fgColor='F4CCCC')
    orange = PatternFill(fill_type='solid',fgColor='FCE5CD')
    gray = PatternFill(fill_type='solid',fgColor='BFBFBF')
    bold = Font(bold=True)
    ca = Alignment(horizontal='center',vertical='center')

    # header
    for c in range(1, ws.max_column+1):
        cell = ws.cell(row=1,column=c); cell.fill=gray; cell.font=bold; cell.alignment=ca
    ws.freeze_panes = 'E2'

    for r in range(2, ws.max_row+1):
        rt = ws.cell(row=r, column=4).value
        un = ws.cell(row=r, column=1).value
        is_total = '全用户合计' in str(un) if un else False

        for c in range(1, ws.max_column+1): ws.cell(row=r,column=c).alignment=ca
        if is_total:
            # 汇总行用深色+粗体
            dark = PatternFill(fill_type='solid',fgColor='4472C4' if rt=='预测' else '70AD47' if rt=='实际' else 'FFC000')
            white_font = Font(bold=True, color='FFFFFF')
            for c in range(1,ws.max_column+1):
                ws.cell(row=r,column=c).fill=dark
                ws.cell(row=r,column=c).font=white_font
        elif rt=='预测':
            for c in range(1,ws.max_column+1): ws.cell(row=r,column=c).fill=blue
        elif rt=='实际':
            for c in range(1,ws.max_column+1): ws.cell(row=r,column=c).fill=green
        elif rt=='偏差':
            for c in range(1,ws.max_column+1): ws.cell(row=r,column=c).fill=yellow
            for c in range(5, ws.max_column+1):
                cell = ws.cell(row=r,column=c)
                if cell.value is None: continue
                try: av=abs(float(cell.value))
                except: continue
                cell.number_format='0.00%'
                if av>=0.30: cell.fill=red
                elif av>=0.15: cell.fill=orange
                elif av>=0.05: cell.fill=yellow

    # column widths
    for c in range(1,ws.max_column+1):
        from openpyxl.utils import get_column_letter
        cl = get_column_letter(c)
        w = {1:24, 2:8, 3:8, 4:6}.get(c, 10)
        ws.column_dimensions[cl].width = w

print(f'已导出: {out_path}')
print(f'共 {len(result)} 行 (18用户 x 3行 = 54行)')
