#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""提取真实交易数据，用于 dashboard.html"""

import csv
import json
import os
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent

def read_csv_raw(p, encodings=('utf-8-sig', 'utf-8', 'gbk')):
    for enc in encodings:
        try:
            with open(p, 'r', encoding=enc) as f:
                return [r for r in csv.reader(f)]
        except:
            continue
    return []

# ========== 1. 现货出清价格 (6月1-22日, 24时段) ==========
price_file = BASE / '2-1价格预测输入/6月/现货出清价格/6月现货出清价格_小时级汇总.xlsx'
import openpyxl
wb = openpyxl.load_workbook(price_file, data_only=True)
ws = wb['6月现货小时级汇总']
price_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
price_daily = []  # [[date_str, h1, h2, ..., h24, daily_avg]]
for row in price_rows:
    if row[0] is None:
        continue
    date = str(row[0])[:10]
    prices = []
    for v in row[1:25]:
        try: prices.append(float(v))
        except: prices.append(0.0)
    daily_avg = float(row[25]) if row[25] else sum(prices)/len(prices)
    price_daily.append({'date': date, 'prices': prices, 'avg': round(daily_avg, 2)})

# 提取典型日负荷数据
print(f"[1] 现货价格: {len(price_daily)} 天")
for p in price_daily[:3]:
    print(f"  {p['date']}: avg={p['avg']}, range={min(p['prices']):.1f}-{max(p['prices']):.1f}")

# ========== 2. 日清分结果 ==========
settle_file = BASE / '日清分/福建亿云能源科技有限公司_2026-06-01到2026-06-22_日清分结果.xlsx'
wb2 = openpyxl.load_workbook(settle_file, data_only=True)
ws2 = wb2['日清分']
settle_rows = list(ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True))
settle_daily = []
for row in settle_rows:
    if row[0] is None: continue
    date = str(row[1])[:10]
    settle_daily.append({
        'date': date,
        'volume': float(row[2] or 0),  # 分时汇总日电量
        'total_volume': float(row[3] or 0),  # 结算总电量
        'avg_price': float(row[4] or 0),  # 结算均价
        'total_fee': float(row[5] or 0),  # 结算总电费
    })

print(f"\n[2] 日清分: {len(settle_daily)} 天")
for s in settle_daily[:3]:
    print(f"  {s['date']}: vol={s['volume']:.1f}MWh, price={s['avg_price']:.2f}元/MWh, fee={s['total_fee']:.2f}元")

# ========== 3. 6月日交易 (提取所有天) ==========
daily_trade_dir = BASE / '3-1交易记录/6月交易'
daily_trades = []
import re
for f in sorted(daily_trade_dir.glob('6-*.xlsx')):
    match = re.match(r'6-(\d+)', f.name)
    if not match: continue
    day = match.group(1)
    wb3 = openpyxl.load_workbook(f, data_only=True)
    ws3 = wb3[wb3.sheetnames[0]]
    rows = list(ws3.iter_rows(min_row=2, max_row=ws3.max_row, values_only=True))
    # 统计购方成交数据
    buy_volume = 0.0
    total_price_weighted = 0.0
    buy_count = 0
    for r in rows:
        if r[2] == '购方' and r[10] is not None:  # 成交总量
            try:
                vol = float(r[10])
                price = float(r[11]) if r[11] else 0
                buy_volume += vol
                total_price_weighted += vol * price
                buy_count += 1
            except: pass
    if buy_volume > 0:
        daily_trades.append({
            'date': f'2026-06-{day}',
            'volume': round(buy_volume, 3),
            'avg_price': round(total_price_weighted/buy_volume, 2)
        })

print(f"\n[3] 日交易: {len(daily_trades)} 天")
for t in daily_trades[:5]:
    print(f"  {t['date']}: vol={t['volume']}, avg_price={t['avg_price']}")

# ========== 4. 用户负荷 CSV ==========
user_csv = BASE / 'AI用/福建省莆田市新兴达饲料有限公司.csv'
user_load = []
if user_csv.exists():
    rows = read_csv_raw(user_csv)
    if rows:
        header = rows[0]
        # get latest day's 24h load
        last_row = rows[-1]
        try:
            for i in range(2, 26):  # columns from 1:00 to 24:00
                user_load.append(float(last_row[i]))
        except: pass
print(f"\n[4] 用户负荷(莆田新兴达饲料): {len(user_load)}h, sum={sum(user_load):.1f}MWh")

# ========== 5. 预测CSV ==========
pred_csv = BASE / '1-2负荷预测输出/prediction/predict_long_v2_6.csv'
pred_data = {}
if pred_csv.exists():
    rows = read_csv_raw(pred_csv)
    if rows:
        header = rows[0]
        for r in rows[1:]:
            if len(r) < 3: continue
            user = r[1] if len(r) > 1 else ''
            try:
                hour = int(r[2]) if len(r) > 2 else 0
                pred_val = float(r[5]) if len(r) > 5 and r[5] else 0
            except: continue
            if user not in pred_data:
                pred_data[user] = [0]*24
            if 0 <= hour < 24:
                pred_data[user][hour] = round(pred_val, 3)
# 汇总所有用户的总负荷
total_load = [0.0]*24
for u, vals in pred_data.items():
    for h in range(24):
        total_load[h] += vals[h]
print(f"\n[5] 预测: {len(pred_data)} 用户, 总负荷曲线={[round(v,2) for v in total_load[:12]]}...")

# ========== 6. 省调负荷 (6月1日) ==========
gd_file = BASE / '2-1价格预测输入/6月/省调负荷/0601.xlsx'
gd_hourly = []
if gd_file.exists():
    wb6 = openpyxl.load_workbook(gd_file, data_only=True)
    ws = wb6['sheet1']
    vals = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        try:
            v = float(row[9]) if row[9] else None
            if v: vals.append(v)
        except: pass
    # 15-min to hourly
    for i in range(0, len(vals), 4):
        chunk = vals[i:i+4]
        gd_hourly.append(round(sum(chunk)/len(chunk), 0))
print(f"\n[6] 省调负荷(0601): {len(gd_hourly)}h, 峰值={max(gd_hourly):.0f}MW")

# ========== 7. 新能源总出力 (6月1日) ==========
ne_file = BASE / '2-1价格预测输入/6月/新能源总出力/0601.xlsx'
ne_hourly = []
if ne_file.exists():
    wb7 = openpyxl.load_workbook(ne_file, data_only=True)
    ws = wb7[wb7.sheetnames[0]]
    vals = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        for col_idx in [9, 10, 11]:
            try:
                v = float(row[col_idx]) if row[col_idx] else None
                if v and v < 50000:  # reasonable range
                    vals.append(v)
                    break
            except: pass
    if vals:
        for i in range(0, min(len(vals), 96), 4):
            chunk = vals[i:i+4]
            ne_hourly.append(round(sum(chunk)/len(chunk), 0))
print(f"\n[7] 新能源出力(0601): {len(ne_hourly)}h, 峰值={max(ne_hourly) if ne_hourly else 'N/A'}MW")

# ========== 8. 水电出力 (6月1日) ==========
hydro_file = BASE / '2-1价格预测输入/6月/水电（含抽蓄）总出力/0601.xlsx'
hydro_hourly = []
if hydro_file.exists():
    wb8 = openpyxl.load_workbook(hydro_file, data_only=True)
    ws = wb8[wb8.sheetnames[0]]
    vals = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        try:
            v = float(row[9]) if row[9] else None
            if v: vals.append(v)
        except: pass
    if vals:
        for i in range(0, min(len(vals), 96), 4):
            chunk = vals[i:i+4]
            hydro_hourly.append(round(sum(chunk)/len(chunk), 0))
print(f"\n[8] 水电出力(0601): {len(hydro_hourly)}h")

# ========== 组装输出 ==========
output = {
    'price': price_daily,
    'settle': settle_daily,
    'daily_trades': daily_trades,
    'user_load': [round(v, 4) for v in user_load],
    'total_load': [round(v, 4) for v in total_load],
    'gd_load': gd_hourly,
    'ne_gen': ne_hourly,
    'hydro_gen': hydro_hourly,
    'users': list(pred_data.keys()),
}

# 计算一些汇总指标
if price_daily:
    output['price_summary'] = {
        'max': max(p['avg'] for p in price_daily),
        'min': min(p['avg'] for p in price_daily),
        'avg': round(sum(p['avg'] for p in price_daily) / len(price_daily), 2),
        'latest': price_daily[-1]['avg'],
        'latest_date': price_daily[-1]['date'],
    }

if settle_daily:
    total_vol = sum(s['total_volume'] for s in settle_daily)
    total_fee = sum(s['total_fee'] for s in settle_daily)
    output['settle_summary'] = {
        'total_vol': round(total_vol, 2),
        'total_fee': round(total_fee, 2),
        'avg_price': round(total_fee / total_vol, 2) if total_vol > 0 else 0,
        'days': len(settle_daily),
    }

if daily_trades:
    output['trade_summary'] = {
        'total_vol': round(sum(t['volume'] for t in daily_trades), 2),
        'days': len(daily_trades),
    }

out_path = BASE / 'dashboard_data.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)
print(f"\n✓ 数据已导出到 {out_path}")
print(f"  - 价格数据: {len(price_daily)}天")
print(f"  - 清分数据: {len(settle_daily)}天")
print(f"  - 交易数据: {len(daily_trades)}天")
print(f"  - 用户: {len(pred_data)}个")
