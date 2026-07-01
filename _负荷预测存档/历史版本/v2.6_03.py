# -*- coding: utf-8 -*-
"""
【V2.6 升级版】负荷预测验证脚本
核心升级:
  P0: 全局关口交差指标（基于 final_pred_net_load）
  P1: 专项诊断1 - 剥离光伏后工厂真实用电需求的拟合能力
  P2: 白天 8:00-19:00 专项分析
  P3: 4家光伏用户关口专项
  P4: 低负荷分类器混淆矩阵 + 漏判(FN)监控
"""

import re
import glob
import warnings
import numpy as np
import pandas as pd

from pathlib import Path
from sklearn.metrics import mean_absolute_error, mean_squared_error
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# =========================================================
# 1. 路径配置
# =========================================================
BASE_DIR = Path(__file__).resolve().parent

USER_MASTER_PATH = BASE_DIR / "input" / "user_master" / "01_用户主档案表.csv"
WEATHER_DIR = BASE_DIR / "input" / "weather"
LOAD_DIR = BASE_DIR / "input" / "load"

OUTPUT_MODEL = BASE_DIR / "output" / "model"
OUTPUT_PREDICTION = BASE_DIR / "output" / "prediction"
OUTPUT_VALIDATION = BASE_DIR / "output" / "validation"
OUTPUT_LOGS = BASE_DIR / "output" / "logs"

for p in [OUTPUT_MODEL, OUTPUT_PREDICTION, OUTPUT_VALIDATION, OUTPUT_LOGS]:
    p.mkdir(parents=True, exist_ok=True)


# =========================================================
# 2. 日志
# =========================================================
LOG_FILE = OUTPUT_LOGS / "03_validate_v2_6_log.txt"


def log(msg):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(str(msg) + "\n")


with open(LOG_FILE, "w", encoding="utf-8") as f:
    f.write("=== V2.6 验证日志 ===\n")


# =========================================================
# 3. 读取运行配置
# =========================================================
CONFIG_PATH = OUTPUT_MODEL / "run_config_v2_6.csv"
if not CONFIG_PATH.exists():
    raise FileNotFoundError("未找到 run_config_v2_6.csv，请先运行 01.py")

RUN_CONFIG = pd.read_csv(CONFIG_PATH, encoding="utf-8-sig").iloc[0]
PREDICT_START_TS = pd.Timestamp(RUN_CONFIG["PREDICT_START"])
PREDICT_END_TS = pd.Timestamp(RUN_CONFIG["PREDICT_END"])
LOW_LOAD_THRESHOLD = float(RUN_CONFIG["LOW_LOAD_THRESHOLD"])
LOW_LOAD_PROBA_THRESHOLD = float(RUN_CONFIG["LOW_LOAD_PROBA_THRESHOLD"])


# =========================================================
# 4. 节假日和气象相关（为 reverse-pv 计算）
# =========================================================
HOLIDAY_MAP = {
    "2024-01-01": "元旦",
    "2024-02-10": "春节", "2024-02-11": "春节", "2024-02-12": "春节", "2024-02-13": "春节",
    "2024-02-14": "春节", "2024-02-15": "春节", "2024-02-16": "春节", "2024-02-17": "春节",
    "2024-04-04": "清明节", "2024-04-05": "清明节", "2024-04-06": "清明节",
    "2024-05-01": "劳动节", "2024-05-02": "劳动节", "2024-05-03": "劳动节", "2024-05-04": "劳动节", "2024-05-05": "劳动节",
    "2024-06-08": "端午节", "2024-06-09": "端午节", "2024-06-10": "端午节",
    "2024-09-15": "中秋节", "2024-09-16": "中秋节", "2024-09-17": "中秋节",
    "2024-10-01": "国庆节", "2024-10-02": "国庆节", "2024-10-03": "国庆节", "2024-10-04": "国庆节",
    "2024-10-05": "国庆节", "2024-10-06": "国庆节", "2024-10-07": "国庆节",
    "2025-01-01": "元旦",
    "2025-01-28": "春节", "2025-01-29": "春节", "2025-01-30": "春节", "2025-01-31": "春节",
    "2025-02-01": "春节", "2025-02-02": "春节", "2025-02-03": "春节", "2025-02-04": "春节",
    "2025-04-04": "清明节", "2025-04-05": "清明节", "2025-04-06": "清明节",
    "2025-05-01": "劳动节", "2025-05-02": "劳动节", "2025-05-03": "劳动节", "2025-05-04": "劳动节", "2025-05-05": "劳动节",
    "2025-05-31": "端午节", "2025-06-01": "端午节", "2025-06-02": "端午节",
    "2025-10-01": "国庆节", "2025-10-02": "国庆节", "2025-10-03": "国庆节", "2025-10-04": "国庆节",
    "2025-10-05": "国庆节", "2025-10-06": "国庆节", "2025-10-07": "国庆节", "2025-10-08": "中秋节",
    "2026-01-01": "元旦",
    "2026-02-17": "春节", "2026-02-18": "春节", "2026-02-19": "春节", "2026-02-20": "春节",
    "2026-02-21": "春节", "2026-02-22": "春节", "2026-02-23": "春节",
    "2026-04-04": "清明节", "2026-04-05": "清明节", "2026-04-06": "清明节",
    "2026-05-01": "劳动节", "2026-05-02": "劳动节", "2026-05-03": "劳动节",
    "2026-06-19": "端午节", "2026-06-20": "端午节", "2026-06-21": "端午节",
    "2026-09-25": "中秋节", "2026-09-26": "中秋节", "2026-09-27": "中秋节",
    "2026-10-01": "国庆节", "2026-10-02": "国庆节", "2026-10-03": "国庆节", "2026-10-04": "国庆节",
    "2026-10-05": "国庆节", "2026-10-06": "国庆节", "2026-10-07": "国庆节",
}

ADJUST_WORKDAYS = set([
    "2024-02-04", "2024-02-18", "2024-04-07", "2024-04-28", "2024-05-11", "2024-09-14", "2024-09-29", "2024-10-12",
    "2025-01-26", "2025-02-08", "2025-04-27", "2025-09-28", "2025-10-11",
    "2026-02-15", "2026-02-28", "2026-04-26", "2026-05-09", "2026-09-27", "2026-10-10",
])


# =========================================================
# 5. 通用函数
# =========================================================
def normalize_text(x):
    if pd.isna(x):
        return None
    x = str(x).strip()
    x = x.replace("\u3000", "").replace(" ", "")
    return x


def safe_read_table(path):
    path = str(path)
    if path.lower().endswith(".csv"):
        try:
            return pd.read_csv(path, encoding="utf-8-sig")
        except Exception:
            try:
                return pd.read_csv(path, encoding="gbk")
            except Exception:
                return pd.read_csv(path)
    else:
        return pd.read_excel(path)


def clean_col_name(c):
    c = str(c).strip()
    c = c.replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")
    c = c.replace("（", "(").replace("）", ")")
    c = c.replace("：", ":")
    return c


def extract_city_district(region_text):
    if pd.isna(region_text):
        return None, None
    txt = str(region_text).strip()
    txt = txt.replace("―", "-").replace("－", "-").replace("C", "-")
    parts = re.split(r"[\/\\\-\_\s]+", txt)
    parts = [p for p in parts if p]
    city, district = None, None
    if len(parts) >= 3:
        city = parts[-2]
        district = parts[-1]
    elif len(parts) == 2:
        city = parts[-1]
    elif len(parts) == 1:
        city = parts[0]
    return city, district


def calc_pct_error(pred, actual):
    if pd.isna(pred) or pd.isna(actual) or actual == 0:
        return np.nan
    return (pred - actual) / actual


def calc_metrics(df, actual_col="actual_load", pred_col="final_pred_net_load"):
    tmp = df.dropna(subset=[actual_col, pred_col]).copy()
    if tmp.empty:
        return {"mae": np.nan, "rmse": np.nan, "mape": np.nan, "sample_count": 0}

    mae = mean_absolute_error(tmp[actual_col], tmp[pred_col])
    rmse = np.sqrt(mean_squared_error(tmp[actual_col], tmp[pred_col]))

    ape = np.where(
        (tmp[actual_col].notna()) & (tmp[actual_col] != 0),
        np.abs(tmp[pred_col] - tmp[actual_col]) / tmp[actual_col],
        np.nan
    )
    mape = np.nanmean(ape)

    return {
        "mae": mae,
        "rmse": rmse,
        "mape": mape,
        "sample_count": len(tmp)
    }


# =========================================================
# 6. 光伏物理公式估算（用于反推 actual_total_load）
# =========================================================
def estimate_pv_generation(radiation, temp, capacity):
    if pd.isna(radiation) or pd.isna(temp) or pd.isna(capacity):
        return 0.0
    if radiation <= 0:
        return 0.0
    temp_coeff = 1 + (-0.004) * (temp - 25)
    pv_estimated = capacity * 1000 * 0.75 * (radiation / 1000) * temp_coeff
    return max(0.0, pv_estimated)


# =========================================================
# 7. 读取主档案
# =========================================================
def load_user_master():
    df = safe_read_table(USER_MASTER_PATH)
    df.columns = [str(c).strip() for c in df.columns]
    df["用户名称_norm"] = df["用户名称"].apply(normalize_text)
    df["是否有光伏_flag"] = df["是否有光伏"].astype(str).str.strip().isin(
        ["是", "有", "1", "true", "True", "Y", "y"]
    ).astype(int)

    if "光伏容量(MW)" in df.columns:
        df["光伏容量(MW)"] = pd.to_numeric(df["光伏容量(MW)"], errors="coerce").fillna(0)
    else:
        df["光伏容量(MW)"] = 0.0

    return df


# =========================================================
# 8. 读取气象（用于反推光伏）
# =========================================================
def hourly_weather_column_mapper(cols):
    mapping = {}
    for c in cols:
        c1 = clean_col_name(c)
        if c1 == "地区":
            mapping[c] = "region"
        elif c1 == "时间":
            mapping[c] = "datetime"
        elif c1 == "天气":
            mapping[c] = "weather"
        elif "温度" in c1 and "露点" not in c1 and "体感" not in c1:
            mapping[c] = "temperature"
        elif "降水量" in c1:
            mapping[c] = "rainfall"
        elif c1 == "风向":
            mapping[c] = "wind_direction"
        elif "短波辐射" in c1 and "总量" not in c1:
            mapping[c] = "shortwave_radiation"
    return mapping


def looks_like_hourly_weather_columns(cols):
    cols_clean = [clean_col_name(c) for c in cols]
    joined = "".join(cols_clean)
    keys = ["地区", "时间", "天气", "温度", "降水量", "湿度"]
    return sum([1 for k in keys if k in joined]) >= 4


def smart_read_hourly_weather_file(path):
    path = str(path)
    if path.lower().endswith(".csv"):
        for enc in ["utf-8-sig", "gbk", "utf-8"]:
            try:
                df = pd.read_csv(path, encoding=enc)
                if looks_like_hourly_weather_columns(df.columns):
                    return df
            except Exception:
                pass
        return pd.read_csv(path)

    xls = pd.ExcelFile(path)
    for sheet_name in xls.sheet_names:
        for header_row in [0, 1, 2, 3, 4, 5]:
            try:
                df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
                if looks_like_hourly_weather_columns(df.columns):
                    return df
            except Exception:
                continue
    return pd.read_excel(path, sheet_name=0)


def load_hourly_weather():
    log("读取小时气象文件（用于反推光伏）...")
    weather_files = (
        glob.glob(str(WEATHER_DIR / "*.xlsx")) +
        glob.glob(str(WEATHER_DIR / "*.xls")) +
        glob.glob(str(WEATHER_DIR / "*.csv"))
    )

    if not weather_files:
        log("[警告] 未找到小时气象文件，跳过硬气象专项诊断")
        return None

    weather_list = []
    for fp in weather_files:
        try:
            df = smart_read_hourly_weather_file(fp)
            df = df.rename(columns=hourly_weather_column_mapper(df.columns))

            required = ["region", "datetime", "weather", "temperature"]
            if any([c not in df.columns for c in required]):
                continue

            df["datetime"] = pd.to_datetime(df["datetime"], errors="coerce")
            df = df.dropna(subset=["datetime"]).copy()

            num_cols = ["temperature", "shortwave_radiation"]
            for c in num_cols:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors="coerce")

            city_list, district_list = [], []
            for x in df["region"]:
                city, district = extract_city_district(x)
                city_list.append(normalize_text(city))
                district_list.append(normalize_text(district))

            df["所在市_norm"] = city_list
            df["所在区_norm"] = district_list
            weather_list.append(df)
        except Exception as e:
            log(f"[错误] 气象读取失败：{Path(fp).name} -> {e}")

    if not weather_list:
        return None

    weather_df = pd.concat(weather_list, ignore_index=True)
    weather_df["datetime"] = pd.to_datetime(weather_df["datetime"]).dt.floor("h")
    weather_df = (
        weather_df.sort_values("datetime")
        .drop_duplicates(subset=["所在市_norm", "所在区_norm", "datetime"], keep="last")
        .reset_index(drop=True)
    )
    log(f"气象数据共 {len(weather_df)} 条")
    return weather_df


# =========================================================
# 9. 读取实际值
# =========================================================
def normalize_load_sheet(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    header_row_idx = None
    for i in range(min(len(df), 10)):
        row_values = [str(x).strip() for x in df.iloc[i].tolist()]
        joined = "".join(row_values)
        if ("电量年月日" in joined) and ("户号" in joined):
            header_row_idx = i
            break

    if header_row_idx is not None:
        new_header = [str(x).strip() for x in df.iloc[header_row_idx].tolist()]
        df = df.iloc[header_row_idx + 1:].copy()
        df.columns = new_header
    else:
        if "电量年月日" not in df.columns or "户号" not in df.columns:
            raise ValueError("未识别到负荷表头")

    df.columns = [str(c).strip() for c in df.columns]
    return df


def parse_actual_sheet(file_path, user_info):
    target_sheet = f"{str(PREDICT_START_TS.year)[2:]}.{PREDICT_START_TS.month}"

    raw = pd.read_excel(file_path, sheet_name=target_sheet, header=0)
    df = normalize_load_sheet(raw)

    required = ["电量年月日", "户号"]
    for c in required:
        if c not in df.columns:
            raise ValueError(f"{Path(file_path).name}-{target_sheet} 缺少字段: {c}")

    hour_cols = [f"{h}:00" for h in range(1, 25)]
    for c in hour_cols:
        if c not in df.columns:
            raise ValueError(f"{Path(file_path).name}-{target_sheet} 缺少小时列: {c}")

    df["电量年月日"] = df["电量年月日"].astype(str).str.strip()
    df = df[df["电量年月日"].str.fullmatch(r"\d{8}", na=False)].copy()
    if df.empty:
        return None

    long_df = df.melt(
        id_vars=["电量年月日", "户号"],
        value_vars=hour_cols,
        var_name="hour_str",
        value_name="actual_load"
    )

    long_df["date"] = pd.to_datetime(long_df["电量年月日"], format="%Y%m%d", errors="coerce")
    long_df["actual_load"] = pd.to_numeric(long_df["actual_load"], errors="coerce")
    long_df["户号"] = long_df["户号"].astype(str).str.strip()

    def build_datetime(row):
        d = row["date"]
        h = int(str(row["hour_str"]).split(":")[0])
        if pd.isna(d):
            return pd.NaT
        if h == 24:
            return d + pd.Timedelta(days=1)
        return d + pd.Timedelta(hours=h)

    long_df["datetime"] = long_df.apply(build_datetime, axis=1)
    long_df["用户编号"] = user_info["用户编号"]
    long_df["用户名称"] = user_info["用户名称"]
    long_df["是否有光伏_flag"] = user_info["是否有光伏_flag"]
    long_df["光伏容量(MW)"] = user_info.get("光伏容量(MW)", 0)

    long_df = long_df[
        (long_df["datetime"] >= PREDICT_START_TS) &
        (long_df["datetime"] < PREDICT_END_TS)
    ].copy()

    return long_df[["用户编号", "用户名称", "户号", "datetime", "actual_load",
                    "是否有光伏_flag", "光伏容量(MW)"]].dropna(subset=["datetime"])


def load_actual(user_master_df):
    log("读取实际预测区间数据...")
    files = glob.glob(str(LOAD_DIR / "*.xlsx")) + glob.glob(str(LOAD_DIR / "*.xls"))

    all_list = []
    target_sheet = f"{str(PREDICT_START_TS.year)[2:]}.{PREDICT_START_TS.month}"

    for fp in files:
        file_name = Path(fp).stem
        user_name_norm = normalize_text(file_name)

        matched = user_master_df[user_master_df["用户名称_norm"] == user_name_norm]
        if matched.empty:
            continue

        user_info = matched.iloc[0]

        try:
            xls = pd.ExcelFile(fp)
        except Exception as e:
            log(f"[错误] 打开失败：{Path(fp).name} -> {e}")
            continue

        if target_sheet not in xls.sheet_names:
            log(f"[警告] 缺少 {target_sheet} sheet：{Path(fp).name}")
            continue

        try:
            one = parse_actual_sheet(fp, user_info)
            if one is not None and not one.empty:
                all_list.append(one)
                log(f"已读取实际：{Path(fp).name} - {target_sheet}")
        except Exception as e:
            log(f"[错误] 读取实际失败：{Path(fp).name} -> {e}")

    if not all_list:
        raise ValueError("未读取到任何预测区间实际数据")

    df = pd.concat(all_list, ignore_index=True)
    df = df.sort_values(["用户编号", "datetime"]).reset_index(drop=True)
    return df


# =========================================================
# 10. 读取预测结果
# =========================================================
def load_prediction_long():
    pred_path = OUTPUT_PREDICTION / "predict_long_v2_6.csv"
    if not pred_path.exists():
        raise FileNotFoundError("未找到 predict_long_v2_6.csv，请先运行 02.py")

    pred_df = pd.read_csv(pred_path, encoding="utf-8-sig")
    pred_df["datetime"] = pd.to_datetime(pred_df["datetime"], errors="coerce")

    keep_cols = [
        "用户编号", "用户名称", "户号", "datetime",
        "final_pred_net_load", "pred_total_load", "pred_pv",
        "is_low_load", "proba_low", "predict_status",
        "是否有光伏_flag", "weather_match_level"
    ]
    keep_cols = [c for c in keep_cols if c in pred_df.columns]
    pred_df = pred_df[keep_cols].copy()

    if "predict_status" not in pred_df.columns:
        pred_df["predict_status"] = "已预测"

    return pred_df


# =========================================================
# 11. 合并验证（含光伏反推）
# =========================================================
def build_validation_long(actual_df, pred_df, weather_df):
    log("合并实际与预测...")

    merge_cols = ["用户编号", "用户名称", "户号", "datetime"]
    df = actual_df.merge(pred_df, on=merge_cols, how="left", suffixes=("", "_pred"))

    # 合并是否有光伏信息
    if "是否有光伏_flag_pred" in df.columns:
        df["是否有光伏_flag"] = df["是否有光伏_flag"].fillna(df["是否有光伏_flag_pred"])
        df = df.drop(columns=["是否有光伏_flag_pred"])

    # ==== V2 专项诊断: 反推 actual_total_load ====
    # 如果预测结果中没有 pred_pv, 尝试从气象数据计算
    df["pv_est"] = 0.0

    if weather_df is not None and "shortwave_radiation" in df.columns:
        # 预测结果中已经有气象数据
        mask_pv = df["是否有光伏_flag"] == 1
        if mask_pv.any():
            df.loc[mask_pv, "pv_est"] = df[mask_pv].apply(
                lambda r: estimate_pv_generation(
                    r.get("shortwave_radiation", 0),
                    r.get("temperature", 25),
                    r.get("光伏容量(MW)", 0)
                ), axis=1
            )
    elif weather_df is not None:
        # 需要从气象表合并
        actual_with_location = actual_df.merge(
            pd.DataFrame({
                "用户编号": actual_df["用户编号"].unique()
            }), on="用户编号"
        )
        # 简化：用已训练的天气表做匹配
        try:
            weather_cleaned_path = OUTPUT_PROCESSED / "hourly_weather_cleaned_v2_6.csv"
            if weather_cleaned_path.exists():
                w_clean = pd.read_csv(weather_cleaned_path, encoding="utf-8-sig")
                w_clean["datetime"] = pd.to_datetime(w_clean["datetime"]).dt.floor("h")
                # 这里不做复杂匹配，仅记录
                log("[注] 气象数据可用但未做实时反推，光伏诊断基于预测中的 pred_pv")
        except Exception:
            pass

    # actual_total_load = 实际负荷 + 光伏估算
    df["actual_total_load"] = df["actual_load"]
    mask_pv = df["是否有光伏_flag"] == 1
    if mask_pv.any():
        df.loc[mask_pv, "actual_total_load"] = (
            df.loc[mask_pv, "actual_load"] + df.loc[mask_pv, "pv_est"]
        )

    # 关口误差
    df["net_error"] = df["final_pred_net_load"] - df["actual_load"]
    df["net_abs_error"] = df["net_error"].abs()
    df["net_pct_error"] = df.apply(
        lambda x: calc_pct_error(x["final_pred_net_load"], x["actual_load"]), axis=1
    )

    # total_load 误差（工厂真实需求拟合能力）
    df["total_error"] = df["pred_total_load"] - df["actual_total_load"]
    df["total_abs_error"] = df["total_error"].abs()

    df["ape_net"] = np.where(
        (df["actual_load"].notna()) & (df["actual_load"] != 0) & (df["final_pred_net_load"].notna()),
        df["net_abs_error"] / df["actual_load"],
        np.nan
    )

    df["ape_total"] = np.where(
        (df["actual_total_load"].notna()) & (df["actual_total_load"] != 0) & (df["pred_total_load"].notna()),
        df["total_abs_error"] / df["actual_total_load"],
        np.nan
    )

    df["status"] = np.where(
        df["final_pred_net_load"].isna(),
        df["predict_status"].fillna("未预测"),
        "已预测"
    )

    df["hour"] = pd.to_datetime(df["datetime"]).dt.hour
    df["is_daytime_8_19"] = df["hour"].isin(list(range(8, 20))).astype(int)
    df["is_low_load_actual"] = (df["actual_load"] < LOW_LOAD_THRESHOLD).astype(int)

    return df.sort_values(["用户编号", "datetime"]).reset_index(drop=True)


# =========================================================
# 12. 全时段指标
# =========================================================
def export_metrics(validation_df):
    total_metrics = calc_metrics(validation_df, "actual_load", "final_pred_net_load")

    pd.DataFrame([{
        "scope": "total",
        "mae": total_metrics["mae"],
        "rmse": total_metrics["rmse"],
        "mape": total_metrics["mape"],
        "sample_count": total_metrics["sample_count"]
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_metrics_total_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    by_user_rows = []
    for (uid, uname), g in validation_df.groupby(["用户编号", "用户名称"]):
        m = calc_metrics(g)
        by_user_rows.append({
            "用户编号": uid,
            "用户名称": uname,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    pd.DataFrame(by_user_rows).to_csv(
        OUTPUT_VALIDATION / "validation_metrics_by_user_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    log("="*50)
    log(f"【全局最终关口电量】")
    log(f"  MAE:  {total_metrics['mae']:.2f} kW")
    log(f"  RMSE: {total_metrics['rmse']:.2f} kW")
    log(f"  MAPE: {total_metrics['mape']:.4%}")
    log(f"  N:    {total_metrics['sample_count']}")
    log("="*50)


# =========================================================
# 13. V2 专项诊断1: 工厂真实用电需求拟合能力
# =========================================================
def export_total_load_diagnostics(validation_df):
    log("\n[专项诊断1] 剥离光伏后的工厂原始用电需求 (Total Load) 拟合能力")

    valid = validation_df.dropna(subset=["pred_total_load", "actual_total_load"]).copy()
    if valid.empty:
        log("  [跳过] 无 pred_total_load 数据")
        return

    total_mae = mean_absolute_error(valid["actual_total_load"], valid["pred_total_load"])
    total_rmse = np.sqrt(mean_squared_error(valid["actual_total_load"], valid["pred_total_load"]))

    # 同时也计算全局关口 MAE 做对比
    net_valid = validation_df.dropna(subset=["final_pred_net_load", "actual_load"]).copy()
    net_mae = mean_absolute_error(net_valid["actual_load"], net_valid["final_pred_net_load"]) if not net_valid.empty else np.nan

    log(f"  --> 模型真实拟合能力 MAE (Total Load): {total_mae:.2f} kW")
    log(f"  --> 全局关口 MAE (Net Load):           {net_mae:.2f} kW")
    if not pd.isna(total_mae) and not pd.isna(net_mae):
        gap = net_mae - total_mae
        log(f"  --> 光伏气象预测误差贡献:              {gap:.2f} kW")
        if gap > 0:
            log("  --> 结论: 光伏气象预测误差是最大短板，建议提升气象预报精度")
        else:
            log("  --> 结论: 模型自身拟合能力是瓶颈，重点优化模型")

    pd.DataFrame([{
        "total_load_mae": total_mae,
        "total_load_rmse": total_rmse,
        "net_load_mae": net_mae,
        "pv_error_contribution": net_mae - total_mae,
        "sample_count": len(valid)
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_total_load_diagnostics_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )


# =========================================================
# 14. V2 专项诊断2: 白天 8:00-19:00
# =========================================================
def export_daytime_metrics(validation_df):
    log("\n[专项诊断2] 白天 8:00-19:00 专项分析")

    day_df = validation_df[validation_df["is_daytime_8_19"] == 1].copy()

    total_metrics = calc_metrics(day_df)
    pd.DataFrame([{
        "scope": "daytime_8_19_total",
        "mae": total_metrics["mae"],
        "rmse": total_metrics["rmse"],
        "mape": total_metrics["mape"],
        "sample_count": total_metrics["sample_count"]
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_daytime_metrics_total_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    log(f"  --> 白天 MAE:  {total_metrics['mae']:.2f} kW")
    log(f"  --> 白天 RMSE: {total_metrics['rmse']:.2f} kW")
    log(f"  --> 白天 MAPE: {total_metrics['mape']:.4%}")

    # 按用户
    by_user_rows = []
    for (uid, uname), g in day_df.groupby(["用户编号", "用户名称"]):
        m = calc_metrics(g)
        by_user_rows.append({
            "用户编号": uid,
            "用户名称": uname,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })
    pd.DataFrame(by_user_rows).to_csv(
        OUTPUT_VALIDATION / "validation_daytime_metrics_by_user_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    # 按是否有光伏
    if "是否有光伏_flag" in day_df.columns:
        by_pv_rows = []
        for pv_flag, g in day_df.groupby("是否有光伏_flag"):
            m = calc_metrics(g)
            label = "光伏用户" if pv_flag == 1 else "非光伏用户"
            by_pv_rows.append({
                "是否有光伏_flag": pv_flag,
                "label": label,
                "mae": m["mae"],
                "rmse": m["rmse"],
                "mape": m["mape"],
                "sample_count": m["sample_count"]
            })
        pd.DataFrame(by_pv_rows).to_csv(
            OUTPUT_VALIDATION / "validation_daytime_metrics_by_pv_v2_6.csv",
            index=False, encoding="utf-8-sig"
        )

        for row in by_pv_rows:
            log(f"  --> {row['label']}: MAE={row['mae']:.2f}kW, RMSE={row['rmse']:.2f}kW")


# =========================================================
# 15. V2 专项诊断3: 光伏用户关口专项
# =========================================================
def export_pv_user_diagnostics(validation_df):
    log("\n[专项诊断3] 光伏用户关口专项分析")

    mask_pv = validation_df["是否有光伏_flag"] == 1
    if not mask_pv.any():
        log("  [跳过] 无光伏用户")
        return

    df_pv = validation_df[mask_pv]
    m = calc_metrics(df_pv)
    log(f"  --> 光伏用户全局关口 MAE:  {m['mae']:.2f} kW")
    log(f"  --> 光伏用户全局关口 RMSE: {m['rmse']:.2f} kW")
    log(f"  --> 光伏用户全局关口 MAPE: {m['mape']:.4%}")

    # 白天时段
    df_pv_day = df_pv[df_pv["is_daytime_8_19"] == 1]
    m_day = calc_metrics(df_pv_day)
    log(f"  --> 光伏用户白天 MAE:       {m_day['mae']:.2f} kW")
    log(f"  --> 光伏用户白天 RMSE:      {m_day['rmse']:.2f} kW")
    log(f"  --> 光伏用户白天 MAPE:      {m_day['mape']:.4%}")

    pd.DataFrame([{
        "scope": "pv_users_total",
        "mae": m["mae"],
        "rmse": m["rmse"],
        "mape": m["mape"],
        "sample_count": m["sample_count"]
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_pv_user_total_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    pd.DataFrame([{
        "scope": "pv_users_daytime",
        "mae": m_day["mae"],
        "rmse": m_day["rmse"],
        "mape": m_day["mape"],
        "sample_count": m_day["sample_count"]
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_pv_user_daytime_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )


# =========================================================
# 16. V2 专项诊断4: 低负荷分类器混淆矩阵
# =========================================================
def export_low_load_classifier_report(validation_df):
    log("\n[专项诊断4] 低负荷分类器混淆矩阵与漏判监控")

    if "is_low_load" not in validation_df.columns:
        log("  [跳过] 无 is_low_load 预测标签")
        return

    eval_df = validation_df.dropna(subset=["is_low_load", "is_low_load_actual"]).copy()
    if eval_df.empty:
        log("  [跳过] 无可评估样本")
        return

    eval_df["is_low_load"] = pd.to_numeric(eval_df["is_low_load"], errors="coerce").fillna(0).astype(int)

    tp = ((eval_df["is_low_load"] == 1) & (eval_df["is_low_load_actual"] == 1)).sum()
    fp = ((eval_df["is_low_load"] == 1) & (eval_df["is_low_load_actual"] == 0)).sum()
    tn = ((eval_df["is_low_load"] == 0) & (eval_df["is_low_load_actual"] == 0)).sum()
    fn = ((eval_df["is_low_load"] == 0) & (eval_df["is_low_load_actual"] == 1)).sum()

    total = len(eval_df)
    pred_low = tp + fp
    actual_low = tp + fn

    log(f"  --> 总样本数:  {total}")
    log(f"  --> 预测低负荷: {pred_low} ({pred_low/total*100:.1f}%)")
    log(f"  --> 实际低负荷: {actual_low} ({actual_low/total*100:.1f}%)")
    log(f"  --> TP(正确判定低): {tp}")
    log(f"  --> FP(误判为低):   {fp}")
    log(f"  --> TN(正确判定正常): {tn}")
    log(f"  --> FN(漏判低负荷):  {fn}")
    if fn > 0:
        log(f"  ⚠ 漏判数={fn} 条！若此数值居高不下，建议降低 LOW_LOAD_PROBA_THRESHOLD (当前={LOW_LOAD_PROBA_THRESHOLD})")

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    log(f"  --> Precision: {precision:.4f}")
    log(f"  --> Recall:    {recall:.4f} (越接近1说明漏判越少)")
    log(f"  --> F1:        {f1:.4f}")

    pd.DataFrame([{
        "total_count": total,
        "pred_low_count": pred_low,
        "actual_low_count": actual_low,
        "tp": tp,
        "fp": fp,
        "tn": tn,
        "fn": fn,
        "precision": precision,
        "recall": recall,
        "f1": f1,
        "low_load_threshold": LOW_LOAD_THRESHOLD,
        "low_load_proba_threshold": LOW_LOAD_PROBA_THRESHOLD
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_low_load_classifier_report_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    # 白天专项
    day_df = eval_df[eval_df["is_daytime_8_19"] == 1].copy()
    if not day_df.empty:
        tp_d = ((day_df["is_low_load"] == 1) & (day_df["is_low_load_actual"] == 1)).sum()
        fn_d = ((day_df["is_low_load"] == 0) & (day_df["is_low_load_actual"] == 1)).sum()
        log(f"  --> 白天漏判数(FN): {fn_d} 条")


# =========================================================
# 17. 按真实值分层分析
# =========================================================
def export_actual_value_bucket_metrics(validation_df):
    log("\n导出按真实值分层误差分析...")

    df = validation_df.copy()

    def get_bucket(x):
        if pd.isna(x):
            return "missing"
        elif x < 20:
            return "<20"
        elif x < 50:
            return "20-50"
        elif x < 100:
            return "50-100"
        elif x < 200:
            return "100-200"
        else:
            return ">=200"

    df["actual_bucket"] = df["actual_load"].apply(get_bucket)
    df_eval = df[df["actual_bucket"] != "missing"].copy()
    bucket_order = ["<20", "20-50", "50-100", "100-200", ">=200"]

    rows_all = []
    for bucket, g in df_eval.groupby("actual_bucket"):
        m = calc_metrics(g)
        rows_all.append({
            "actual_bucket": bucket,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    out_all = pd.DataFrame(rows_all)
    if not out_all.empty:
        out_all["actual_bucket"] = pd.Categorical(out_all["actual_bucket"], categories=bucket_order, ordered=True)
        out_all = out_all.sort_values("actual_bucket")

    out_all.to_csv(
        OUTPUT_VALIDATION / "validation_metrics_by_actual_bucket_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    # 白天分层
    day_df = df_eval[df_eval["is_daytime_8_19"] == 1].copy()
    rows_day = []
    for bucket, g in day_df.groupby("actual_bucket"):
        m = calc_metrics(g)
        rows_day.append({
            "actual_bucket": bucket,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    out_day = pd.DataFrame(rows_day)
    if not out_day.empty:
        out_day["actual_bucket"] = pd.Categorical(out_day["actual_bucket"], categories=bucket_order, ordered=True)
        out_day = out_day.sort_values("actual_bucket")

    out_day.to_csv(
        OUTPUT_VALIDATION / "validation_daytime_metrics_by_actual_bucket_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    log("按真实值分层误差分析导出完成")


# =========================================================
# 18. 天气匹配层级分析
# =========================================================
def export_weather_match_metrics(validation_df):
    log("\n导出天气匹配层级分析...")

    if "weather_match_level" not in validation_df.columns:
        log("[跳过] 预测结果中不存在 weather_match_level")
        return

    df = validation_df.copy()

    count_df = (
        df.groupby("weather_match_level", dropna=False)
        .size()
        .reset_index(name="sample_count")
    )
    count_df.to_csv(
        OUTPUT_VALIDATION / "validation_weather_match_level_count_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    metric_rows = []
    for level, g in df.groupby("weather_match_level", dropna=False):
        m = calc_metrics(g)
        metric_rows.append({
            "weather_match_level": level,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    pd.DataFrame(metric_rows).to_csv(
        OUTPUT_VALIDATION / "validation_weather_match_level_metrics_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    day_df = df[df["is_daytime_8_19"] == 1].copy()
    day_rows = []
    for level, g in day_df.groupby("weather_match_level", dropna=False):
        m = calc_metrics(g)
        day_rows.append({
            "weather_match_level": level,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })
    pd.DataFrame(day_rows).to_csv(
        OUTPUT_VALIDATION / "validation_daytime_weather_match_level_metrics_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )

    log("天气匹配层级分析导出完成")


# =========================================================
# 19. 按小时分析
# =========================================================
def export_hourly_metrics(validation_df):
    log("\n导出按小时误差分析...")

    df = validation_df.copy()

    rows = []
    for h in range(24):
        hour_df = df[df["hour"] == h].copy()
        m = calc_metrics(hour_df)
        rows.append({
            "hour": h,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    pd.DataFrame(rows).to_csv(
        OUTPUT_VALIDATION / "validation_metrics_by_hour_v2_6.csv",
        index=False, encoding="utf-8-sig"
    )
    log("按小时误差分析导出完成")


# =========================================================
# 20. 每个用户宽表：每天3行（预测/实际/偏差）
# =========================================================
def build_validation_wide_for_one_user(g):
    g = g.copy().sort_values("datetime")

    g["base_date"] = np.where(
        g["datetime"].dt.hour == 0,
        (g["datetime"] - pd.Timedelta(days=1)).dt.normalize(),
        g["datetime"].dt.normalize()
    )
    g["base_date"] = pd.to_datetime(g["base_date"])

    day_list = sorted(g["base_date"].dropna().unique())
    rows = []

    for d in day_list:
        day = pd.to_datetime(d)
        one_day = g[g["base_date"] == day].copy()

        account_value = one_day["户号"].dropna().astype(str).iloc[0] if one_day["户号"].notna().any() else ""

        pred_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "预测"}
        actual_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "实际"}
        error_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "偏差"}

        pred_sum = 0.0
        actual_sum = 0.0
        pred_empty = True
        actual_empty = True

        for h in range(1, 25):
            target_dt = day + pd.Timedelta(days=1) if h == 24 else day + pd.Timedelta(hours=h)
            hit = one_day[one_day["datetime"] == target_dt]

            if hit.empty:
                pred_val = np.nan
                actual_val = np.nan
            else:
                pred_val = hit["final_pred_net_load"].iloc[0] if "final_pred_net_load" in hit.columns else np.nan
                actual_val = hit["actual_load"].iloc[0] if "actual_load" in hit.columns else np.nan

            pct_val = calc_pct_error(pred_val, actual_val)

            pred_row[f"{h}:00"] = pred_val
            actual_row[f"{h}:00"] = actual_val
            error_row[f"{h}:00"] = pct_val

            if pd.notna(pred_val):
                pred_sum += pred_val
                pred_empty = False
            if pd.notna(actual_val):
                actual_sum += actual_val
                actual_empty = False

        pred_row["合计"] = np.nan if pred_empty else pred_sum
        actual_row["合计"] = np.nan if actual_empty else actual_sum
        error_row["合计"] = calc_pct_error(pred_sum, actual_sum) if (not pred_empty and not actual_empty) else np.nan

        rows.extend([pred_row, actual_row, error_row])

    final_cols = ["电量年月日", "户号", "类型"] + [f"{h}:00" for h in range(1, 25)] + ["合计"]
    return pd.DataFrame(rows)[final_cols]


# =========================================================
# 21. 全体汇总宽表
# =========================================================
def build_validation_total_wide(validation_df):
    df = validation_df.copy().sort_values("datetime")

    df["base_date"] = np.where(
        df["datetime"].dt.hour == 0,
        (df["datetime"] - pd.Timedelta(days=1)).dt.normalize(),
        df["datetime"].dt.normalize()
    )
    df["base_date"] = pd.to_datetime(df["base_date"])

    day_list = sorted(df["base_date"].dropna().unique())
    rows = []

    for d in day_list:
        day = pd.to_datetime(d)
        one_day = df[df["base_date"] == day].copy()

        pred_row = {"电量年月日": day.strftime("%Y%m%d"), "类型": "预测"}
        actual_row = {"电量年月日": day.strftime("%Y%m%d"), "类型": "实际"}
        error_row = {"电量年月日": day.strftime("%Y%m%d"), "类型": "偏差"}

        pred_sum_day = 0.0
        actual_sum_day = 0.0
        pred_empty = True
        actual_empty = True

        for h in range(1, 25):
            target_dt = day + pd.Timedelta(days=1) if h == 24 else day + pd.Timedelta(hours=h)
            hit = one_day[one_day["datetime"] == target_dt].copy()

            if hit.empty:
                pred_val = np.nan
                actual_val = np.nan
            else:
                pred_val = hit["final_pred_net_load"].sum(min_count=1) if "final_pred_net_load" in hit.columns else np.nan
                actual_val = hit["actual_load"].sum(min_count=1) if "actual_load" in hit.columns else np.nan

            pct_val = calc_pct_error(pred_val, actual_val)

            pred_row[f"{h}:00"] = pred_val
            actual_row[f"{h}:00"] = actual_val
            error_row[f"{h}:00"] = pct_val

            if pd.notna(pred_val):
                pred_sum_day += pred_val
                pred_empty = False
            if pd.notna(actual_val):
                actual_sum_day += actual_val
                actual_empty = False

        pred_row["合计"] = np.nan if pred_empty else pred_sum_day
        actual_row["合计"] = np.nan if actual_empty else actual_sum_day
        error_row["合计"] = calc_pct_error(pred_sum_day, actual_sum_day) if (not pred_empty and not actual_empty) else np.nan

        rows.extend([pred_row, actual_row, error_row])

    final_cols = ["电量年月日", "类型"] + [f"{h}:00" for h in range(1, 25)] + ["合计"]
    return pd.DataFrame(rows)[final_cols]


# =========================================================
# 22. Excel格式化
# =========================================================
def format_validation_sheet(ws):
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align

    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 14
        elif col == 2:
            ws.column_dimensions[col_letter].width = 28
        elif col == 3:
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 12

    for row in range(2, max_row + 1):
        row_type = ws.cell(row=row, column=3).value
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).alignment = center_align

        if row_type == "预测":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = blue_fill
        elif row_type == "实际":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = green_fill
        elif row_type == "偏差":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = yellow_fill
            for col in range(4, max_col + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is None or val == "":
                    continue
                try:
                    abs_val = abs(float(val))
                except Exception:
                    continue
                cell.number_format = "0.00%"
                if abs_val >= 0.30:
                    cell.fill = red_fill
                elif abs_val >= 0.15:
                    cell.fill = orange_fill
                elif abs_val >= 0.05:
                    cell.fill = warn_fill


def format_total_validation_sheet(ws):
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align

    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 14
        elif col == 2:
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 12

    for row in range(2, max_row + 1):
        row_type = ws.cell(row=row, column=2).value
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).alignment = center_align

        if row_type == "预测":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = blue_fill
        elif row_type == "实际":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = green_fill
        elif row_type == "偏差":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = yellow_fill
            for col in range(3, max_col + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is None or val == "":
                    continue
                try:
                    abs_val = abs(float(val))
                except Exception:
                    continue
                cell.number_format = "0.00%"
                if abs_val >= 0.30:
                    cell.fill = red_fill
                elif abs_val >= 0.15:
                    cell.fill = orange_fill
                elif abs_val >= 0.05:
                    cell.fill = warn_fill


# =========================================================
# 23. 导出验证结果
# =========================================================
def export_validation_outputs(validation_df):
    log("\n导出验证宽表...")
    validation_df.to_csv(OUTPUT_VALIDATION / "validation_hourly_long_v2_6.csv", index=False, encoding="utf-8-sig")

    # 按用户导出宽表
    out_excel = OUTPUT_VALIDATION / "validation_hourly_wide_v2_6.xlsx"
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        for (uid, uname), g in validation_df.groupby(["用户编号", "用户名称"]):
            wide_df = build_validation_wide_for_one_user(g)
            sheet_name = str(uname)[:31] if str(uname) else str(uid)
            wide_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.book[sheet_name]
            format_validation_sheet(ws)
    log(f"已导出：{out_excel.name}")

    # 总量汇总宽表
    total_wide_df = build_validation_total_wide(validation_df)
    total_excel = OUTPUT_VALIDATION / "validation_total_hourly_wide_v2_6.xlsx"
    with pd.ExcelWriter(total_excel, engine="openpyxl") as writer:
        total_wide_df.to_excel(writer, sheet_name="总量汇总", index=False)
        ws = writer.book["总量汇总"]
        format_total_validation_sheet(ws)
    log(f"已导出：{total_excel.name}")


# =========================================================
# 24. 主流程
# =========================================================
def main():
    log("=== 开始 V2.6 验证 ===")
    log(f"PREDICT_START = {PREDICT_START_TS}")
    log(f"PREDICT_END   = {PREDICT_END_TS}")
    log(f"LOW_LOAD_THRESHOLD       = {LOW_LOAD_THRESHOLD}")
    log(f"LOW_LOAD_PROBA_THRESHOLD = {LOW_LOAD_PROBA_THRESHOLD}")

    # 数据加载
    user_master_df = load_user_master()
    actual_df = load_actual(user_master_df)
    pred_df = load_prediction_long()
    weather_df = load_hourly_weather()

    # 合并验证
    validation_df = build_validation_long(actual_df, pred_df, weather_df)

    # 验证分析
    export_metrics(validation_df)

    # V2 专项诊断
    export_total_load_diagnostics(validation_df)
    export_daytime_metrics(validation_df)
    export_pv_user_diagnostics(validation_df)
    export_low_load_classifier_report(validation_df)

    # 常规分析
    export_actual_value_bucket_metrics(validation_df)
    export_weather_match_metrics(validation_df)
    export_hourly_metrics(validation_df)

    # 导出 Excel
    export_validation_outputs(validation_df)

    log("\n=== V2.6 验证完成 ===")


if __name__ == "__main__":
    main()
