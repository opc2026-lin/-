# -*- coding: utf-8 -*-
import argparse
import glob
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill


BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parents[1]

INPUT_ROOT = PROJECT_DIR / "1-1负荷预测输入"
PREDICTION_ROOT = PROJECT_DIR / "1-2负荷预测输出"
EVAL_ROOT = PROJECT_DIR / "1-3负荷回测效验"

USER_MASTER_PATH = INPUT_ROOT / "用户主档案表.xlsx"
LOAD_DIR = INPUT_ROOT / "1.分时段历史用电信息"
WEATHER_DIR = INPUT_ROOT / "3.真实天气"
PREDICTION_DIR = PREDICTION_ROOT / "prediction"
MODEL_DIR = PREDICTION_ROOT / "model"
BACKTEST_DIR = EVAL_ROOT / "回测结果"

BACKTEST_DIR.mkdir(parents=True, exist_ok=True)

PV_EFFICIENCY = 0.75
PV_TEMP_COEFF = -0.004
MAPE_MIN_ACTUAL_KW = 1.0


def normalize_text(value):
    if pd.isna(value):
        return None
    return str(value).strip().replace("\u3000", "").replace(" ", "")


def convert_yes_no(value):
    text = normalize_text(value)
    return 1 if text in {"是", "有", "1", "true", "True", "Y", "y"} else 0


def compute_metrics(y_true, y_pred):
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    mae = np.mean(np.abs(y_true - y_pred))
    rmse = np.sqrt(np.mean((y_true - y_pred) ** 2))
    mape_mask = np.abs(y_true) >= MAPE_MIN_ACTUAL_KW
    if mape_mask.any():
        mape = np.mean(np.abs(y_true[mape_mask] - y_pred[mape_mask]) / np.abs(y_true[mape_mask])) * 100
        mape_count = int(mape_mask.sum())
    else:
        mape = np.nan
        mape_count = 0
    return (
        round(float(mae), 3),
        round(float(rmse), 3),
        round(float(mape), 3) if not np.isnan(mape) else np.nan,
        mape_count,
    )


def parse_args():
    parser = argparse.ArgumentParser(description="负荷预测真实回测 v4.3.8")
    parser.add_argument("--start-date", required=True, help="回测起始日期 YYYY-MM-DD")
    parser.add_argument("--days", type=int, default=1, help="连续回测天数，默认 1")
    return parser.parse_args()


def load_run_config():
    config_path = MODEL_DIR / "run_config_v3.csv"
    if not config_path.exists():
        raise FileNotFoundError(f"缺少运行配置文件: {config_path}")
    cfg = pd.read_csv(config_path, encoding="utf-8-sig").iloc[0]
    return float(cfg["LOW_LOAD_THRESHOLD"])


def load_user_master():
    df = pd.read_excel(USER_MASTER_PATH)
    df.columns = [str(col).strip() for col in df.columns]

    if "所在区" not in df.columns and {"所在市", "所在县"}.issubset(df.columns):
        df["所在区"] = df["所在市"].astype(str) + df["所在县"].astype(str)
    if "是否有光伏" not in df.columns:
        df["是否有光伏"] = "否"
    if "光伏容量(MW)" not in df.columns:
        df["光伏容量(MW)"] = 0.0

    df["用户编号"] = df["用户编号"].astype(str).str.strip()
    df["用户名称_norm"] = df["用户名称"].apply(normalize_text)
    df["所在市_norm"] = df["所在市"].apply(normalize_text)
    df["是否有光伏_flag"] = df["是否有光伏"].apply(convert_yes_no)
    df["光伏容量(MW)"] = pd.to_numeric(df["光伏容量(MW)"], errors="coerce").fillna(0.0)
    return df


def load_prediction_long():
    pred_path = PREDICTION_DIR / "prediction_result_v3_long.csv"
    if not pred_path.exists():
        raise FileNotFoundError(f"缺少预测长表文件: {pred_path}")
    df = pd.read_csv(pred_path, encoding="utf-8-sig")
    df["datetime"] = pd.to_datetime(df["datetime"], errors="coerce")
    df = df.dropna(subset=["datetime"]).copy()
    for col in ["pred_total_load", "pred_net_load", "pv_est", "proba_low"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df["用户编号"] = df["用户编号"].astype(str).str.strip()
    return df


def find_header_row(raw_df):
    for idx in range(min(len(raw_df), 6)):
        row_text = " ".join(str(x) for x in raw_df.iloc[idx].tolist() if pd.notna(x))
        if "电量年月日" in row_text or "电力用户名称" in row_text:
            return idx
    return None


def build_user_lookup(user_master_df):
    lookup = {}
    for _, row in user_master_df.iterrows():
        name_norm = row["用户名称_norm"]
        if not name_norm:
            continue
        lookup[name_norm] = row
        if "(" in name_norm:
            lookup[name_norm.split("(")[0]] = row
        if "（" in name_norm:
            lookup[name_norm.split("（")[0]] = row
    return lookup


def scan_actual_date_span():
    files = sorted(glob.glob(str(LOAD_DIR / "*.xlsx"))) + sorted(glob.glob(str(LOAD_DIR / "*.xls")))
    all_dates = []
    for fp in files:
        try:
            raw = pd.read_excel(fp, sheet_name=0, header=None)
        except Exception:
            continue
        header_row = find_header_row(raw)
        if header_row is None:
            continue
        df = pd.read_excel(fp, sheet_name=0, header=header_row)
        df.columns = [str(col).strip() for col in df.columns]
        if "电量年月日" not in df.columns:
            continue
        s = df["电量年月日"].astype(str).str.strip()
        s = s[s.str.fullmatch(r"\d{8}", na=False)]
        if s.empty:
            continue
        dates = pd.to_datetime(s, format="%Y%m%d", errors="coerce").dropna()
        all_dates.extend(dates.tolist())
    if not all_dates:
        return None, None
    return min(all_dates), max(all_dates)


def load_actual_load(user_master_df, predict_start, predict_end):
    files = sorted(glob.glob(str(LOAD_DIR / "*.xlsx"))) + sorted(glob.glob(str(LOAD_DIR / "*.xls")))
    if not files:
        raise FileNotFoundError(f"缺少真实负荷文件目录: {LOAD_DIR}")

    user_lookup = build_user_lookup(user_master_df)
    records = []
    valid_dates = {dt.normalize() for dt in pd.date_range(predict_start - pd.Timedelta(days=1), predict_end, freq="D")}

    for fp in files:
        try:
            raw = pd.read_excel(fp, sheet_name=0, header=None)
        except Exception:
            continue

        header_row = find_header_row(raw)
        if header_row is None:
            continue

        df = pd.read_excel(fp, sheet_name=0, header=header_row)
        df.columns = [str(col).strip() for col in df.columns]
        if "电量年月日" not in df.columns or "电力用户名称" not in df.columns:
            continue

        df["电量年月日"] = df["电量年月日"].astype(str).str.strip()
        df = df[df["电量年月日"].str.fullmatch(r"\d{8}", na=False)].copy()
        if df.empty:
            continue

        df["date"] = pd.to_datetime(df["电量年月日"], format="%Y%m%d", errors="coerce")
        df = df[df["date"].isin(valid_dates)].copy()
        if df.empty:
            continue

        hour_cols = sorted(
            [str(col).strip() for col in df.columns if re.fullmatch(r"(1?\d|2[0-4]):00", str(col).strip())],
            key=lambda x: int(x.split(":")[0]),
        )
        if len(hour_cols) < 20:
            continue

        df["电力用户名称_norm"] = df["电力用户名称"].apply(normalize_text)
        for user_name_norm, user_rows in df.groupby("电力用户名称_norm"):
            matched_user = user_lookup.get(user_name_norm)
            if matched_user is None and user_name_norm:
                for key, value in user_lookup.items():
                    if key and (key in user_name_norm or user_name_norm in key):
                        matched_user = value
                        break
            if matched_user is None:
                continue

            melted = user_rows.melt(
                id_vars=["date"],
                value_vars=hour_cols,
                var_name="hour_str",
                value_name="actual_load",
            )
            melted["actual_load"] = pd.to_numeric(melted["actual_load"], errors="coerce")

            def build_datetime(row):
                hour = int(str(row["hour_str"]).split(":")[0])
                base_date = row["date"]
                if pd.isna(base_date):
                    return pd.NaT
                if hour == 24:
                    return base_date + pd.Timedelta(days=1)
                return base_date + pd.Timedelta(hours=hour)

            melted["datetime"] = melted.apply(build_datetime, axis=1)
            melted = melted[(melted["datetime"] >= predict_start) & (melted["datetime"] < predict_end)].copy()
            if melted.empty:
                continue

            melted["用户编号"] = str(matched_user["用户编号"]).strip()
            melted["用户名称"] = matched_user["用户名称"]
            melted["所在市_norm"] = matched_user["所在市_norm"]
            melted["是否有光伏_flag"] = matched_user["是否有光伏_flag"]
            melted["光伏容量(MW)"] = matched_user["光伏容量(MW)"]
            records.append(
                melted[["用户编号", "用户名称", "所在市_norm", "datetime", "actual_load", "是否有光伏_flag", "光伏容量(MW)"]]
            )

    if not records:
        min_day, max_day = scan_actual_date_span()
        if min_day is None:
            raise ValueError("真实负荷目录里没有可识别日期数据。")
        raise ValueError(f"没有读到目标区间的真实负荷。真实数据当前覆盖: {min_day:%Y-%m-%d} 到 {max_day:%Y-%m-%d}")

    actual_df = pd.concat(records, ignore_index=True)
    actual_df = actual_df.dropna(subset=["datetime", "actual_load"]).copy()
    actual_df = (
        actual_df.groupby(
            ["用户编号", "用户名称", "所在市_norm", "datetime", "是否有光伏_flag", "光伏容量(MW)"],
            as_index=False,
        )["actual_load"].mean()
    )
    return actual_df.sort_values(["用户编号", "datetime"]).reset_index(drop=True)


def load_actual_weather():
    weather_files = sorted(glob.glob(str(WEATHER_DIR / "*.xlsx"))) + sorted(glob.glob(str(WEATHER_DIR / "*.xls")))
    city_sheet_map = {"宁德": "宁德", "莆田": "莆田", "福州": "福州", "泉州": "泉州"}
    records = []

    for fp in weather_files:
        try:
            xls = pd.ExcelFile(fp)
        except Exception:
            continue

        for sheet_name in xls.sheet_names:
            city_name = None
            for key, value in city_sheet_map.items():
                if key in str(sheet_name):
                    city_name = value
                    break
            if city_name is None:
                continue

            try:
                raw = pd.read_excel(fp, sheet_name=sheet_name, header=None)
            except Exception:
                continue

            row_idx = 0
            while row_idx < len(raw):
                header_text = str(raw.iloc[row_idx, 0]) if pd.notna(raw.iloc[row_idx, 0]) else ""
                date_match = re.search(r"(\d+)月(\d+)日", header_text)
                if not date_match:
                    row_idx += 1
                    continue

                month = int(date_match.group(1))
                day = int(date_match.group(2))
                date_base = pd.Timestamp(year=2026, month=month, day=day)

                for hour in range(24):
                    col = hour + 1
                    temperature = pd.to_numeric(raw.iloc[row_idx + 2, col], errors="coerce") if row_idx + 2 < len(raw) and col < raw.shape[1] else np.nan
                    radiation = pd.to_numeric(raw.iloc[row_idx + 3, col], errors="coerce") if row_idx + 3 < len(raw) and col < raw.shape[1] else np.nan
                    records.append(
                        {
                            "所在市_norm": city_name,
                            "datetime": date_base + pd.Timedelta(hours=hour),
                            "temperature": temperature,
                            "shortwave_radiation": radiation,
                        }
                    )

                row_idx += 8

    if not records:
        return pd.DataFrame(columns=["所在市_norm", "datetime", "temperature", "shortwave_radiation"])

    weather_df = pd.DataFrame(records)
    weather_df["datetime"] = pd.to_datetime(weather_df["datetime"]).dt.floor("h")
    weather_df = weather_df.drop_duplicates(subset=["所在市_norm", "datetime"], keep="last")
    return weather_df.sort_values(["所在市_norm", "datetime"]).reset_index(drop=True)


def estimate_actual_pv(actual_df, weather_df):
    df = actual_df.merge(weather_df, on=["所在市_norm", "datetime"], how="left")
    for col in ["temperature", "shortwave_radiation"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        city_median = df.groupby("所在市_norm")[col].transform("median")
        global_median = df[col].median()
        if pd.isna(global_median):
            global_median = 0.0
        df[col] = df[col].fillna(city_median).fillna(global_median)

    df["actual_pv_est"] = 0.0
    pv_mask = (df["是否有光伏_flag"] == 1) & (df["光伏容量(MW)"] > 0)
    df.loc[pv_mask, "actual_pv_est"] = (
        df.loc[pv_mask, "光伏容量(MW)"] * 1000 * PV_EFFICIENCY *
        (df.loc[pv_mask, "shortwave_radiation"] / 1000.0) *
        (1 + PV_TEMP_COEFF * (df.loc[pv_mask, "temperature"] - 25.0))
    )
    df["actual_pv_est"] = df["actual_pv_est"].clip(lower=0.0)
    df["actual_total_load"] = df["actual_load"] + df["actual_pv_est"]
    return df


def build_validation(actual_df, pred_df, low_load_threshold):
    keep_cols = [
        "用户编号", "用户名称", "datetime", "pv_capacity", "pv_est", "proba_low",
        "is_low_load_pred", "pred_total_load", "pred_net_load",
    ]
    pred_df = pred_df[[col for col in keep_cols if col in pred_df.columns]].copy()
    merged = actual_df.merge(pred_df, on=["用户编号", "用户名称", "datetime"], how="left")
    merged["hour"] = merged["datetime"].dt.hour
    for col in ["pred_net_load", "pred_total_load", "pv_est", "actual_load", "actual_total_load"]:
        merged[col] = pd.to_numeric(merged[col], errors="coerce")
    merged["pv_est"] = merged["pv_est"].fillna(0.0)
    merged["net_error"] = merged["pred_net_load"] - merged["actual_load"]
    merged["net_abs_error"] = merged["net_error"].abs()
    merged["net_rel_error"] = np.where(np.abs(merged["actual_load"]) >= 1e-9, merged["net_error"] / merged["actual_load"], np.nan)
    merged["total_error"] = merged["pred_total_load"] - merged["actual_total_load"]
    merged["total_abs_error"] = merged["total_error"].abs()
    merged["has_prediction"] = merged["pred_net_load"].notna().astype(int)
    merged["is_low_load_actual"] = (merged["actual_load"] < low_load_threshold).astype(int)
    return merged


def build_summary_sheet(validation_df, predict_start, predict_end):
    valid_net = validation_df.dropna(subset=["actual_load", "pred_net_load"]).copy()
    valid_total = validation_df.dropna(subset=["actual_total_load", "pred_total_load"]).copy()
    matched_count = int(validation_df["has_prediction"].sum())
    if matched_count == 0 or valid_net.empty or valid_total.empty:
        raise SystemExit(
            f"回测失败：{predict_start:%Y-%m-%d} 没有对齐到任何预测结果。"
            f"请先确认 1-2负荷预测输出\\prediction 下已生成该日期预测文件。"
        )

    net_mae, net_rmse, net_mape, net_mape_count = compute_metrics(valid_net["actual_load"], valid_net["pred_net_load"])
    total_mae, total_rmse, total_mape, total_mape_count = compute_metrics(valid_total["actual_total_load"], valid_total["pred_total_load"])

    summary = pd.DataFrame([{
        "起始日期": predict_start.strftime("%Y-%m-%d"),
        "结束日期": (predict_end - pd.Timedelta(days=1)).strftime("%Y-%m-%d"),
        "样本数": len(validation_df),
        "成功对齐预测样本": matched_count,
        "用户数": validation_df["用户编号"].nunique(),
        "净负荷_MAE_kW": net_mae,
        "净负荷_RMSE_kW": net_rmse,
        "净负荷_MAPE_pct": net_mape,
        "净负荷_MAPE样本数": net_mape_count,
        "总负荷_MAE_kW": total_mae,
        "总负荷_RMSE_kW": total_rmse,
        "总负荷_MAPE_pct": total_mape,
        "总负荷_MAPE样本数": total_mape_count,
        "真实净负荷_MWh": round(float(valid_net["actual_load"].sum() / 1000.0), 4),
        "预测净负荷_MWh": round(float(valid_net["pred_net_load"].sum() / 1000.0), 4),
        "真实总负荷_MWh": round(float(valid_total["actual_total_load"].sum() / 1000.0), 4),
        "预测总负荷_MWh": round(float(valid_total["pred_total_load"].sum() / 1000.0), 4),
    }])
    return summary, matched_count, net_mae, net_rmse, net_mape, net_mape_count, total_mae, total_rmse, total_mape, total_mape_count


def build_user_summary_sheet(validation_df):
    user_summary = (
        validation_df.dropna(subset=["actual_load", "pred_net_load"])
        .groupby(["用户编号", "用户名称"], as_index=False)
        .agg(
            真实净负荷_MWh=("actual_load", lambda s: round(float(s.sum() / 1000.0), 4)),
            预测净负荷_MWh=("pred_net_load", lambda s: round(float(s.sum() / 1000.0), 4)),
            净负荷_MAE_kW=("net_abs_error", lambda s: round(float(s.mean()), 3)),
            最大绝对误差_kW=("net_abs_error", lambda s: round(float(s.max()), 3)),
        )
        .sort_values("净负荷_MAE_kW", ascending=False)
    )
    return user_summary


def build_compare_sheet(validation_df, user_master_df):
    df = validation_df.dropna(subset=["actual_load", "pred_net_load"]).copy()
    if df.empty:
        return pd.DataFrame()

    user_meta = user_master_df[["用户编号", "是否有光伏_flag"]].drop_duplicates().copy()
    df = df.merge(user_meta, on="用户编号", how="left", suffixes=("", "_meta"))
    if "是否有光伏_flag_meta" in df.columns:
        df["是否有光伏_flag"] = df["是否有光伏_flag"].fillna(df["是否有光伏_flag_meta"])

    df["小时列"] = (df["hour"].replace({0: 24}).astype(int)).astype(str) + ":00"
    df["预测_MWh"] = (df["pred_net_load"] / 1000.0).round(4)
    df["实际_MWh"] = (df["actual_load"] / 1000.0).round(4)
    df["偏差"] = ((df["pred_net_load"] - df["actual_load"]) / df["actual_load"]).replace([np.inf, -np.inf], np.nan)

    hour_order = [f"{h}:00" for h in range(1, 25)]

    pivot_frames = []
    for type_name, value_col in [("预测", "预测_MWh"), ("实际", "实际_MWh"), ("偏差", "偏差")]:
        part = df.pivot_table(
            index=["用户编号", "用户名称", "是否有光伏_flag"],
            columns="小时列",
            values=value_col,
            aggfunc="mean",
        ).reset_index()
        for col in hour_order:
            if col not in part.columns:
                part[col] = np.nan
        part = part[["用户编号", "用户名称", "是否有光伏_flag"] + hour_order].copy()
        part["类型"] = type_name
        if type_name == "预测":
            part["日合计(MWh)"] = part[hour_order].sum(axis=1, skipna=True).round(4)
        elif type_name == "实际":
            part["日合计(MWh)"] = part[hour_order].sum(axis=1, skipna=True).round(4)
        else:
            pred_total = df.groupby(["用户编号", "用户名称"])["pred_net_load"].sum()
            actual_total = df.groupby(["用户编号", "用户名称"])["actual_load"].sum()
            total_rel = ((pred_total - actual_total) / actual_total.replace(0, np.nan)).replace([np.inf, -np.inf], np.nan)
            total_rel = total_rel.reset_index(name="日合计(MWh)")
            part = part.merge(total_rel, on=["用户编号", "用户名称"], how="left")
            part["日合计(MWh)"] = part["日合计(MWh)"]
        pivot_frames.append(part)

    compare_df = pd.concat(pivot_frames, ignore_index=True)
    compare_df["是否光伏"] = compare_df["是否有光伏_flag"].map({1: "是", 0: "否"}).fillna("否")
    compare_df = compare_df[["用户名称", "用户编号", "是否光伏", "类型"] + hour_order + ["日合计(MWh)"]]
    compare_df = compare_df.sort_values(["用户名称", "类型"], key=lambda s: s.map({"预测": 0, "实际": 1, "偏差": 2}).fillna(9) if s.name == "类型" else s).reset_index(drop=True)
    return compare_df


def apply_compare_sheet_format(workbook_path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    ws = wb["全用户对比"]

    pos_light = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
    pos_mid = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")
    pos_strong = PatternFill(fill_type="solid", start_color="F4B183", end_color="F4B183")
    pos_deep = PatternFill(fill_type="solid", start_color="E06666", end_color="E06666")

    neg_light = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
    neg_mid = PatternFill(fill_type="solid", start_color="BDD7EE", end_color="BDD7EE")
    neg_strong = PatternFill(fill_type="solid", start_color="9DC3E6", end_color="9DC3E6")
    neg_deep = PatternFill(fill_type="solid", start_color="5B9BD5", end_color="5B9BD5")

    headers = [cell.value for cell in ws[1]]
    hour_cols = [idx for idx, value in enumerate(headers, start=1) if isinstance(value, str) and re.fullmatch(r"\d{1,2}:00", value)]
    total_col = headers.index("日合计(MWh)") + 1
    type_col = headers.index("类型") + 1

    for row in range(2, ws.max_row + 1):
        row_type = ws.cell(row=row, column=type_col).value
        if row_type == "偏差":
            for col in hour_cols + [total_col]:
                cell = ws.cell(row=row, column=col)
                if cell.value is None or cell.value == "":
                    continue
                try:
                    raw_value = float(cell.value)
                    abs_value = abs(raw_value)
                except Exception:
                    continue

                cell.value = f"{raw_value * 100:.2f}%"
                cell.number_format = "@"

                if raw_value >= 0:
                    if abs_value < 0.10:
                        cell.fill = pos_light
                    elif abs_value < 0.20:
                        cell.fill = pos_mid
                    elif abs_value < 0.30:
                        cell.fill = pos_strong
                    else:
                        cell.fill = pos_deep
                else:
                    if abs_value < 0.10:
                        cell.fill = neg_light
                    elif abs_value < 0.20:
                        cell.fill = neg_mid
                    elif abs_value < 0.30:
                        cell.fill = neg_strong
                    else:
                        cell.fill = neg_deep
        else:
            for col in hour_cols + [total_col]:
                ws.cell(row=row, column=col).number_format = "0.0000"

    wb.save(workbook_path)


def format_date_range_name(predict_start, predict_end):
    end_day = predict_end - pd.Timedelta(days=1)
    if predict_start.normalize() == end_day.normalize():
        return f"{predict_start.month}月{predict_start.day}日_回测对比汇总.xlsx"
    return f"{predict_start.month}月{predict_start.day}日至{end_day.month}月{end_day.day}日_回测对比汇总.xlsx"


def save_reports(validation_df, user_master_df, predict_start, predict_end):
    summary, matched_count, net_mae, net_rmse, net_mape, net_mape_count, total_mae, total_rmse, total_mape, total_mape_count = build_summary_sheet(
        validation_df, predict_start, predict_end
    )
    user_summary = build_user_summary_sheet(validation_df)
    compare_df = build_compare_sheet(validation_df, user_master_df)

    workbook_path = BACKTEST_DIR / format_date_range_name(predict_start, predict_end)
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="回测汇总", index=False)
        compare_df.to_excel(writer, sheet_name="全用户对比", index=False)
        user_summary.to_excel(writer, sheet_name="用户误差汇总", index=False)
    apply_compare_sheet_format(workbook_path)

    print("=" * 60)
    print("真实回测结果")
    print("=" * 60)
    print(f"回测日期: {predict_start:%Y-%m-%d} 到 {(predict_end - pd.Timedelta(days=1)):%Y-%m-%d}")
    print(f"样本数: {len(validation_df)}")
    print(f"成功对齐预测样本: {matched_count}")
    print(f"净负荷 MAE: {net_mae:.3f} kW")
    print(f"净负荷 RMSE: {net_rmse:.3f} kW")
    print(f"净负荷 MAPE: {net_mape:.3f}% (仅统计真实负荷 >= {MAPE_MIN_ACTUAL_KW} kW 的 {net_mape_count} 个样本)")
    print(f"总负荷 MAE: {total_mae:.3f} kW")
    print(f"总负荷 RMSE: {total_rmse:.3f} kW")
    print(f"总负荷 MAPE: {total_mape:.3f}% (仅统计真实总负荷 >= {MAPE_MIN_ACTUAL_KW} kW 的 {total_mape_count} 个样本)")
    print(f"回测文件: {workbook_path}")


def main():
    args = parse_args()
    predict_start = pd.Timestamp(args.start_date).normalize()
    predict_end = predict_start + pd.Timedelta(days=max(int(args.days), 1))

    min_day, max_day = scan_actual_date_span()
    if max_day is None:
        raise SystemExit(f"真实负荷目录为空或无法识别: {LOAD_DIR}")
    if predict_end > (max_day + pd.Timedelta(days=1)):
        raise SystemExit(
            f"不能回测 {predict_start:%Y-%m-%d} 到 {(predict_end - pd.Timedelta(days=1)):%Y-%m-%d}。"
            f"真实负荷当前只到 {max_day:%Y-%m-%d}，真实数据目录: {LOAD_DIR}"
        )

    low_load_threshold = load_run_config()
    user_master_df = load_user_master()
    pred_df = load_prediction_long()
    actual_df = load_actual_load(user_master_df, predict_start, predict_end)
    weather_df = load_actual_weather()
    actual_df = estimate_actual_pv(actual_df, weather_df)
    validation_df = build_validation(actual_df, pred_df, low_load_threshold)
    save_reports(validation_df, user_master_df, predict_start, predict_end)


if __name__ == "__main__":
    main()
