# -*- coding: utf-8 -*-
"""
【V2.6 适配版】负荷预测验证脚本 - 适配新目录结构
P0-P4: 全系列诊断（关口误差/工厂真实需求/白天专项/光伏专项/低负荷混淆矩阵）
"""
import re, glob, warnings, sys, io
import numpy as np
import pandas as pd
from pathlib import Path
from sklearn.metrics import mean_absolute_error, mean_squared_error
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings("ignore")

# =========================================================
# 1. 路径配置 (适配新目录)
# =========================================================
BASE_DIR    = Path(__file__).resolve().parent
INPUT_DIR   = BASE_DIR / "1-1负荷预测输入"
OUTPUT_DIR  = BASE_DIR / "1-2负荷预测输出"

USER_MASTER_PATH = INPUT_DIR / "用户主档案表.xlsx"
LOAD_DIR         = INPUT_DIR / "1.分时段历史用电信息"
WEATHER_DIR      = INPUT_DIR / "3.真实天气"  # 历史天气数据

OUTPUT_MODEL      = OUTPUT_DIR / "model"
OUTPUT_PREDICTION = OUTPUT_DIR / "prediction"
OUTPUT_VALIDATION = OUTPUT_DIR / "validation"
OUTPUT_LOGS       = OUTPUT_DIR / "validation_logs"

for p in [OUTPUT_MODEL, OUTPUT_PREDICTION, OUTPUT_VALIDATION, OUTPUT_LOGS]:
    p.mkdir(parents=True, exist_ok=True)

# =========================================================
# 2. 日志
# =========================================================
LOG_FILE = OUTPUT_LOGS / "03_validate_v2_6_log.txt"
def log(msg):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(str(msg) + "\n")
with open(LOG_FILE, "w", encoding="utf-8") as f:
    f.write("=== V2.6 验证日志 ===\n")

# =========================================================
# 3. 读取运行配置
# =========================================================
CONFIG_PATH = OUTPUT_MODEL / "run_config_v2_6.csv"
if not CONFIG_PATH.exists():
    raise FileNotFoundError("未找到 run_config_v2_6.csv，请先运行 01_train_v2.py")

RUN_CONFIG = pd.read_csv(CONFIG_PATH, encoding="utf-8-sig").iloc[0]
PREDICT_START_TS         = pd.Timestamp(RUN_CONFIG["PREDICT_START"])
PREDICT_END_TS           = pd.Timestamp(RUN_CONFIG["PREDICT_END"])
LOW_LOAD_THRESHOLD       = float(RUN_CONFIG["LOW_LOAD_THRESHOLD"])
LOW_LOAD_PROBA_THRESHOLD = float(RUN_CONFIG["LOW_LOAD_PROBA_THRESHOLD"])

# =========================================================
# 4-5. 节假日 + 通用函数 (保持不变)
# =========================================================
HOLIDAY_MAP = {
    "2024-01-01": "元旦",
    "2024-02-10": "春节","2024-02-11": "春节","2024-02-12": "春节","2024-02-13": "春节",
    "2024-02-14": "春节","2024-02-15": "春节","2024-02-16": "春节","2024-02-17": "春节",
    "2024-04-04": "清明节","2024-04-05": "清明节","2024-04-06": "清明节",
    "2024-05-01": "劳动节","2024-05-02": "劳动节","2024-05-03": "劳动节","2024-05-04": "劳动节","2024-05-05": "劳动节",
    "2024-06-08": "端午节","2024-06-09": "端午节","2024-06-10": "端午节",
    "2024-09-15": "中秋节","2024-09-16": "中秋节","2024-09-17": "中秋节",
    "2024-10-01": "国庆节","2024-10-02": "国庆节","2024-10-03": "国庆节","2024-10-04": "国庆节",
    "2024-10-05": "国庆节","2024-10-06": "国庆节","2024-10-07": "国庆节",
    "2025-01-01": "元旦",
    "2025-01-28": "春节","2025-01-29": "春节","2025-01-30": "春节","2025-01-31": "春节",
    "2025-02-01": "春节","2025-02-02": "春节","2025-02-03": "春节","2025-02-04": "春节",
    "2025-04-04": "清明节","2025-04-05": "清明节","2025-04-06": "清明节",
    "2025-05-01": "劳动节","2025-05-02": "劳动节","2025-05-03": "劳动节","2025-05-04": "劳动节","2025-05-05": "劳动节",
    "2025-05-31": "端午节","2025-06-01": "端午节","2025-06-02": "端午节",
    "2025-10-01": "国庆节","2025-10-02": "国庆节","2025-10-03": "国庆节","2025-10-04": "国庆节",
    "2025-10-05": "国庆节","2025-10-06": "国庆节","2025-10-07": "国庆节","2025-10-08": "中秋节",
    "2026-01-01": "元旦",
    "2026-02-17": "春节","2026-02-18": "春节","2026-02-19": "春节","2026-02-20": "春节",
    "2026-02-21": "春节","2026-02-22": "春节","2026-02-23": "春节",
    "2026-04-04": "清明节","2026-04-05": "清明节","2026-04-06": "清明节",
    "2026-05-01": "劳动节","2026-05-02": "劳动节","2026-05-03": "劳动节",
    "2026-06-19": "端午节","2026-06-20": "端午节","2026-06-21": "端午节",
    "2026-09-25": "中秋节","2026-09-26": "中秋节","2026-09-27": "中秋节",
    "2026-10-01": "国庆节","2026-10-02": "国庆节","2026-10-03": "国庆节","2026-10-04": "国庆节",
    "2026-10-05": "国庆节","2026-10-06": "国庆节","2026-10-07": "国庆节",
}
ADJUST_WORKDAYS = set([
    "2024-02-04","2024-02-18","2024-04-07","2024-04-28","2024-05-11","2024-09-14","2024-09-29","2024-10-12",
    "2025-01-26","2025-02-08","2025-04-27","2025-09-28","2025-10-11",
    "2026-02-15","2026-02-28","2026-04-26","2026-05-09","2026-09-27","2026-10-10",
])

def normalize_text(x):
    if pd.isna(x): return None
    return str(x).strip().replace("\u3000","").replace(" ","")

def convert_yes_no(x):
    x = normalize_text(x)
    return 1 if x in ["是","有","1","true","True","Y","y"] else 0

def extract_city(s):
    if pd.isna(s): return None
    s = str(s).strip()
    for city in ["福州","泉州","莆田","宁德","厦门","漳州","龙岩","三明","南平"]:
        if city in s: return city
    return s

def calc_pct_error(pred, actual):
    if pd.isna(pred) or pd.isna(actual) or actual == 0: return np.nan
    return (pred - actual) / actual

def calc_metrics(df, actual_col="actual_load", pred_col="final_pred_net_load"):
    tmp = df.dropna(subset=[actual_col, pred_col]).copy()
    if tmp.empty: return {"mae":np.nan,"rmse":np.nan,"mape":np.nan,"sample_count":0}
    mae  = mean_absolute_error(tmp[actual_col], tmp[pred_col])
    rmse = np.sqrt(mean_squared_error(tmp[actual_col], tmp[pred_col]))
    ape  = np.where((tmp[actual_col].notna()) & (tmp[actual_col]!=0),
                    np.abs(tmp[pred_col]-tmp[actual_col])/tmp[actual_col], np.nan)
    return {"mae":mae,"rmse":rmse,"mape":np.nanmean(ape),"sample_count":len(tmp)}

def estimate_pv_generation(radiation, temp, capacity):
    if pd.isna(radiation) or pd.isna(temp) or pd.isna(capacity): return 0.0
    if radiation <= 0 or capacity <= 0: return 0.0
    temp_coeff = 1 + (-0.004)*(temp-25)
    return max(0.0, capacity*1000*0.75*(radiation/1000)*temp_coeff)

# =========================================================
# 6. 读取主档案 (适配版)
# =========================================================
def load_user_master():
    df = pd.read_excel(USER_MASTER_PATH)
    df.columns = [str(c).strip() for c in df.columns]
    if "所在区" not in df.columns: df["所在区"] = "未知"
    if "用户类型" not in df.columns: df["用户类型"] = df.get("用户类型标签","未知")
    if "光伏容量(MW)" not in df.columns: df["光伏容量(MW)"] = 0.0

    df["用户名称_norm"]   = df["用户名称"].apply(normalize_text)
    df["是否有光伏_flag"]  = df["是否有光伏"].apply(convert_yes_no)
    df["光伏容量(MW)"]     = pd.to_numeric(df["光伏容量(MW)"], errors="coerce").fillna(0)
    return df

# =========================================================
# 7. 读取气象 (适配新格式 - 纵向解析)
# =========================================================
def load_hourly_weather():
    log("读取小时气象文件（用于反推光伏）...")
    weather_files = sorted(glob.glob(str(WEATHER_DIR / "福建四地区历史天气数据_*.xlsx")))
    if not weather_files:
        log("[警告] 未找到历史天气文件，跳过硬气象诊断")
        return None

    SHEET_CITY_MAP = {"宁德":"宁德","莆田":"莆田","福州":"福州","泉州":"泉州"}
    all_rows = []
    for fp in weather_files:
        try: xls = pd.ExcelFile(fp)
        except: continue
        for sheet_name in xls.sheet_names:
            city = None
            for k,v in SHEET_CITY_MAP.items():
                if k in sheet_name: city=v; break
            if city is None: city = extract_city(sheet_name) or sheet_name
            try: raw = pd.read_excel(fp, sheet_name=sheet_name, header=None)
            except: continue
            if raw.empty: continue

            # Vertical format: each day = 8 rows (header + 变量 + temp + rad + cloud + humid + rain + blank)
            row_idx = 0
            nrows = len(raw)
            while row_idx < nrows:
                header_val = str(raw.iloc[row_idx,0]) if pd.notna(raw.iloc[row_idx,0]) else ""
                date_match = re.search(r"(\d+)月(\d+)日", header_val)
                if not date_match: row_idx += 1; continue
                month, day = int(date_match.group(1)), int(date_match.group(2))
                year = 2026
                d = pd.Timestamp(year=year, month=month, day=day)
                if row_idx + 3 >= nrows: row_idx += 1; continue

                def safe_num(r, c):
                    try: return pd.to_numeric(raw.iloc[r,c] if r<nrows else np.nan, errors="coerce")
                    except: return np.nan

                for h in range(24):
                    col = h + 1
                    if col >= raw.shape[1]: break
                    dt = d + pd.Timedelta(hours=h)
                    temp  = safe_num(row_idx+2, col)
                    rad   = safe_num(row_idx+3, col)
                    all_rows.append({"所在市_norm":city,"datetime":dt,
                                     "temperature":temp,"shortwave_radiation":rad})
                row_idx += 8

    if not all_rows:
        log("[警告] 未解析到有效气象数据")
        return None
    df = pd.DataFrame(all_rows).dropna(subset=["datetime"])
    df["datetime"] = pd.to_datetime(df["datetime"]).dt.floor("h")
    df = df.sort_values("datetime").drop_duplicates(subset=["所在市_norm","datetime"],keep="last").reset_index(drop=True)
    log(f"气象数据共 {len(df)} 条，城市 {df['所在市_norm'].nunique()} 个")
    return df

# =========================================================
# 8. 读取实际值 (适配新月度文件格式)
# =========================================================
def load_actual(user_master_df):
    log("读取实际预测区间数据...")
    files = sorted(glob.glob(str(LOAD_DIR / "*.xlsx")))
    if not files: raise FileNotFoundError("未找到负荷文件")

    # Name mapping
    user_name_map = {}
    for _, u in user_master_df.iterrows():
        n = normalize_text(u["用户名称"])
        user_name_map[n] = u
        user_name_map[n.replace("(套餐用户)","").strip()] = u

    all_list = []
    for fp in files:
        fname = Path(fp).name
        log(f"  读取: {fname}")
        try:
            raw = pd.read_excel(fp, sheet_name=0)
            header_idx = None
            for i in range(min(len(raw), 5)):
                row_str = " ".join([str(x) for x in raw.iloc[i].tolist() if pd.notna(x)])
                if "电量年月日" in row_str or "电力用户名称" in row_str:
                    header_idx = i; break
            if header_idx is not None and header_idx > 0:
                raw.columns = [str(x).strip() for x in raw.iloc[header_idx].tolist()]
                raw = raw.iloc[header_idx+1:].reset_index(drop=True)
            raw.columns = [str(c).strip() for c in raw.columns]
        except Exception as e:
            log(f"  [错误] {e}"); continue

        if "电量年月日" not in raw.columns: continue
        raw["电量年月日"] = raw["电量年月日"].astype(str).str.strip()
        raw = raw[raw["电量年月日"].str.fullmatch(r"\d{8}", na=False)].copy()
        if raw.empty: continue

        # Filter to prediction date range
        raw["date"] = pd.to_datetime(raw["电量年月日"], format="%Y%m%d", errors="coerce")
        pred_start_date = PREDICT_START_TS.normalize() - pd.Timedelta(days=1)  # include previous day for 24:00
        raw = raw[(raw["date"] >= pred_start_date) & (raw["date"] <= PREDICT_END_TS.normalize())].copy()
        if raw.empty: continue

        # Hour columns
        hour_cols = sorted([str(c).strip() for c in raw.columns if re.fullmatch(r"(1?\d|2[0-4]):00", str(c).strip())],
                          key=lambda x: int(x.split(":")[0]))
        if len(hour_cols) < 24: continue

        # Match users
        raw["电力用户名称_norm"] = raw.get("电力用户名称", pd.Series(dtype=str)).apply(normalize_text)
        for name_norm in raw["电力用户名称_norm"].unique():
            matched_u = None
            if name_norm in user_name_map:
                matched_u = user_name_map[name_norm]
            else:
                for k, v in user_name_map.items():
                    if k and name_norm and (k in name_norm or name_norm in k):
                        matched_u = v; break
            if matched_u is None: continue

            idx = raw[raw["电力用户名称_norm"]==name_norm].index
            # Melt hours
            melt = raw.loc[idx].melt(id_vars=["电量年月日","date"], value_vars=hour_cols,
                                     var_name="hour_str", value_name="actual_load")
            melt["actual_load"] = pd.to_numeric(melt["actual_load"], errors="coerce")

            def build_datetime(row):
                d = row["date"]; h = int(str(row["hour_str"]).split(":")[0])
                if pd.isna(d): return pd.NaT
                if h == 24: return d + pd.Timedelta(days=1)
                return d + pd.Timedelta(hours=h)
            melt["datetime"] = melt.apply(build_datetime, axis=1)

            # Filter to prediction window
            melt = melt[(melt["datetime"]>=PREDICT_START_TS) & (melt["datetime"]<PREDICT_END_TS)].copy()
            if melt.empty: continue

            melt["用户编号"]         = matched_u["用户编号"]
            melt["用户名称"]         = matched_u["用户名称"]
            melt["户号"]             = matched_u["用户编号"]
            melt["是否有光伏_flag"]   = matched_u["是否有光伏_flag"]
            melt["光伏容量(MW)"]      = matched_u["光伏容量(MW)"]
            all_list.append(melt[["用户编号","用户名称","户号","datetime","actual_load",
                                  "是否有光伏_flag","光伏容量(MW)"]])

    if not all_list: raise ValueError("未读取到预测区间实际数据")
    df = pd.concat(all_list, ignore_index=True).dropna(subset=["datetime"])
    df = df.sort_values(["用户编号","datetime"]).reset_index(drop=True)
    log(f"实际数据读取完成，共 {len(df)} 条，{df['用户编号'].nunique()} 个用户")
    return df

# =========================================================
# 9. 读取预测结果
# =========================================================
def load_prediction_long():
    pred_path = OUTPUT_PREDICTION / "predict_long_v2_6.csv"
    if not pred_path.exists():
        raise FileNotFoundError("未找到 predict_long_v2_6.csv，请先运行 02_predict_v2.py")
    pred_df = pd.read_csv(pred_path, encoding="utf-8-sig")
    pred_df["datetime"] = pd.to_datetime(pred_df["datetime"], errors="coerce")
    keep_cols = ["用户编号","用户名称","户号","datetime","final_pred_net_load","pred_total_load",
                 "pred_pv","is_low_load","proba_low","predict_status","是否有光伏_flag","weather_match_level"]
    keep_cols = [c for c in keep_cols if c in pred_df.columns]
    pred_df = pred_df[keep_cols].copy()
    if "predict_status" not in pred_df.columns: pred_df["predict_status"] = "已预测"
    return pred_df

# =========================================================
# 10. 合并验证（含光伏反推）
# =========================================================
def build_validation_long(actual_df, pred_df, weather_df):
    log("合并实际与预测...")
    merge_cols = ["用户编号","用户名称","户号","datetime"]
    df = actual_df.merge(pred_df, on=merge_cols, how="left", suffixes=("","_pred"))
    if "是否有光伏_flag_pred" in df.columns:
        df["是否有光伏_flag"] = df["是否有光伏_flag"].fillna(df["是否有光伏_flag_pred"])
        df = df.drop(columns=["是否有光伏_flag_pred"])

    # PV estimation for actual total load
    df["pv_est"] = 0.0
    mask_pv = df["是否有光伏_flag"] == 1
    if mask_pv.any():
        # Try using pred_pv from prediction
        if "pred_pv" in df.columns:
            df.loc[mask_pv, "pv_est"] = df.loc[mask_pv, "pred_pv"].fillna(0)
        elif weather_df is not None:
            # Match weather to actual_df for PV estimation
            for _, row in df[mask_pv].iterrows():
                city = extract_city(user_master_df[user_master_df["用户编号"]==row["用户编号"]]["所在市"].iloc[0])
                w = weather_df[(weather_df["所在市_norm"]==city) & (weather_df["datetime"]==row["datetime"])]
                if not w.empty:
                    df.at[row.name,"pv_est"] = estimate_pv_generation(
                        w["shortwave_radiation"].iloc[0], w["temperature"].iloc[0], row["光伏容量(MW)"])

    df["actual_total_load"] = df["actual_load"]
    if mask_pv.any():
        df.loc[mask_pv,"actual_total_load"] = df.loc[mask_pv,"actual_load"] + df.loc[mask_pv,"pv_est"]

    # Error calculations
    df["net_error"]     = df["final_pred_net_load"] - df["actual_load"]
    df["net_abs_error"] = df["net_error"].abs()
    df["total_error"]   = df["pred_total_load"] - df["actual_total_load"]
    df["total_abs_error"] = df["total_error"].abs()
    df["ape_net"]   = np.where((df["actual_load"].notna())&(df["actual_load"]!=0)&(df["final_pred_net_load"].notna()),
                               df["net_abs_error"]/df["actual_load"], np.nan)
    df["ape_total"] = np.where((df["actual_total_load"].notna())&(df["actual_total_load"]!=0)&(df["pred_total_load"].notna()),
                               df["total_abs_error"]/df["actual_total_load"], np.nan)
    df["status"]   = np.where(df["final_pred_net_load"].isna(), df["predict_status"].fillna("未预测"), "已预测")
    df["hour"]     = pd.to_datetime(df["datetime"]).dt.hour
    df["is_daytime_8_19"] = df["hour"].isin(list(range(8,20))).astype(int)
    df["is_low_load_actual"] = (df["actual_load"] < LOW_LOAD_THRESHOLD).astype(int)
    return df.sort_values(["用户编号","datetime"]).reset_index(drop=True)

# =========================================================
# 11-18. 专项诊断 (核心逻辑不变)
# =========================================================
def export_metrics(validation_df):
    total_metrics = calc_metrics(validation_df)
    pd.DataFrame([{"scope":"total","mae":total_metrics["mae"],"rmse":total_metrics["rmse"],
                   "mape":total_metrics["mape"],"sample_count":total_metrics["sample_count"]}]
                ).to_csv(OUTPUT_VALIDATION/"validation_metrics_total_v2_6.csv",index=False,encoding="utf-8-sig")

    by_user = [{"用户编号":uid,"用户名称":uname,**calc_metrics(g)}
               for (uid,uname),g in validation_df.groupby(["用户编号","用户名称"])]
    pd.DataFrame(by_user).to_csv(OUTPUT_VALIDATION/"validation_metrics_by_user_v2_6.csv",index=False,encoding="utf-8-sig")
    log("="*50)
    log(f"【全局最终关口电量】 MAE:{total_metrics['mae']:.2f}kW  RMSE:{total_metrics['rmse']:.2f}kW  MAPE:{total_metrics['mape']:.4%}  N:{total_metrics['sample_count']}")
    log("="*50)

def export_total_load_diagnostics(validation_df):
    log("\n[诊断1] 工厂真实用电需求拟合能力")
    valid = validation_df.dropna(subset=["pred_total_load","actual_total_load"])
    if valid.empty: log("  [跳过]"); return
    t_mae  = mean_absolute_error(valid["actual_total_load"],valid["pred_total_load"])
    t_rmse = np.sqrt(mean_squared_error(valid["actual_total_load"],valid["pred_total_load"]))
    net_v  = validation_df.dropna(subset=["final_pred_net_load","actual_load"])
    n_mae  = mean_absolute_error(net_v["actual_load"],net_v["final_pred_net_load"]) if not net_v.empty else np.nan
    log(f"  Total Load MAE:{t_mae:.2f}kW  Net Load MAE:{n_mae:.2f}kW")
    pd.DataFrame([{"total_load_mae":t_mae,"total_load_rmse":t_rmse,"net_load_mae":n_mae,
                   "pv_error_contribution":n_mae-t_mae,"sample_count":len(valid)}]
                ).to_csv(OUTPUT_VALIDATION/"validation_total_load_diagnostics_v2_6.csv",index=False,encoding="utf-8-sig")

def export_daytime_metrics(validation_df):
    log("\n[诊断2] 白天8:00-19:00分析")
    dd = validation_df[validation_df["is_daytime_8_19"]==1]
    m = calc_metrics(dd)
    pd.DataFrame([{"scope":"daytime","mae":m["mae"],"rmse":m["rmse"],"mape":m["mape"],"sample_count":m["sample_count"]}]
                ).to_csv(OUTPUT_VALIDATION/"validation_daytime_metrics_total_v2_6.csv",index=False,encoding="utf-8-sig")
    log(f"  白天 MAE:{m['mae']:.2f}kW  RMSE:{m['rmse']:.2f}kW  MAPE:{m['mape']:.4%}")
    by_user = [{"用户编号":uid,"用户名称":uname,**calc_metrics(g)} for (uid,uname),g in dd.groupby(["用户编号","用户名称"])]
    pd.DataFrame(by_user).to_csv(OUTPUT_VALIDATION/"validation_daytime_metrics_by_user_v2_6.csv",index=False,encoding="utf-8-sig")
    if "是否有光伏_flag" in dd.columns:
        by_pv = [{"是否有光伏_flag":p,"label":"光伏用户" if p==1 else "非光伏用户",**calc_metrics(g)} for p,g in dd.groupby("是否有光伏_flag")]
        pd.DataFrame(by_pv).to_csv(OUTPUT_VALIDATION/"validation_daytime_metrics_by_pv_v2_6.csv",index=False,encoding="utf-8-sig")
        for r in by_pv: log(f"  {r['label']}: MAE={r['mae']:.2f}kW RMSE={r['rmse']:.2f}kW")

def export_pv_user_diagnostics(validation_df):
    log("\n[诊断3] 光伏用户关口专项")
    mask = validation_df["是否有光伏_flag"]==1
    if not mask.any(): log("  [跳过]"); return
    m = calc_metrics(validation_df[mask]); md = calc_metrics(validation_df[mask & (validation_df["is_daytime_8_19"]==1)])
    log(f"  全局 MAE:{m['mae']:.2f}kW  白天 MAE:{md['mae']:.2f}kW")
    pd.DataFrame([{"scope":"pv_total",**m}]).to_csv(OUTPUT_VALIDATION/"validation_pv_user_total_v2_6.csv",index=False,encoding="utf-8-sig")
    pd.DataFrame([{"scope":"pv_daytime",**md}]).to_csv(OUTPUT_VALIDATION/"validation_pv_user_daytime_v2_6.csv",index=False,encoding="utf-8-sig")

def export_low_load_classifier_report(validation_df):
    log("\n[诊断4] 低负荷混淆矩阵")
    if "is_low_load" not in validation_df.columns: log("  [跳过]"); return
    ev = validation_df.dropna(subset=["is_low_load","is_low_load_actual"])
    if ev.empty: log("  [跳过]"); return
    ev["is_low_load"] = pd.to_numeric(ev["is_low_load"],errors="coerce").fillna(0).astype(int)
    tp=((ev["is_low_load"]==1)&(ev["is_low_load_actual"]==1)).sum()
    fp=((ev["is_low_load"]==1)&(ev["is_low_load_actual"]==0)).sum()
    tn=((ev["is_low_load"]==0)&(ev["is_low_load_actual"]==0)).sum()
    fn=((ev["is_low_load"]==0)&(ev["is_low_load_actual"]==1)).sum()
    log(f"  TP:{tp} FP:{fp} TN:{tn} FN:{fn}")
    prec=tp/(tp+fp) if tp+fp>0 else 0
    rec=tp/(tp+fn) if tp+fn>0 else 0
    f1=2*prec*rec/(prec+rec) if prec+rec>0 else 0
    log(f"  Precision:{prec:.4f} Recall:{rec:.4f} F1:{f1:.4f}")
    pd.DataFrame([{"total_count":len(ev),"pred_low_count":tp+fp,"actual_low_count":tp+fn,
                   "tp":tp,"fp":fp,"tn":tn,"fn":fn,"precision":prec,"recall":rec,"f1":f1,
                   "low_load_threshold":LOW_LOAD_THRESHOLD,"low_load_proba_threshold":LOW_LOAD_PROBA_THRESHOLD}]
                ).to_csv(OUTPUT_VALIDATION/"validation_low_load_classifier_report_v2_6.csv",index=False,encoding="utf-8-sig")

def export_actual_value_bucket_metrics(validation_df):
    log("\n按真实值分层分析...")
    def bucket(x):
        if pd.isna(x): return "missing"
        if x<20: return "<20"
        if x<50: return "20-50"
        if x<100: return "50-100"
        if x<200: return "100-200"
        return ">=200"
    df = validation_df.copy(); df["b"] = df["actual_load"].apply(bucket)
    ev = df[df["b"]!="missing"]
    order = ["<20","20-50","50-100","100-200",">=200"]
    rows = [{"actual_bucket":b,**calc_metrics(g)} for b,g in ev.groupby("b")]
    out = pd.DataFrame(rows)
    if not out.empty:
        out["actual_bucket"] = pd.Categorical(out["actual_bucket"],categories=order,ordered=True)
        out.sort_values("actual_bucket").to_csv(OUTPUT_VALIDATION/"validation_metrics_by_actual_bucket_v2_6.csv",index=False,encoding="utf-8-sig")
    dd = ev[ev["is_daytime_8_19"]==1]
    dr = [{"actual_bucket":b,**calc_metrics(g)} for b,g in dd.groupby("b")]
    out2 = pd.DataFrame(dr)
    if not out2.empty:
        out2["actual_bucket"] = pd.Categorical(out2["actual_bucket"],categories=order,ordered=True)
        out2.sort_values("actual_bucket").to_csv(OUTPUT_VALIDATION/"validation_daytime_metrics_by_actual_bucket_v2_6.csv",index=False,encoding="utf-8-sig")

def export_weather_match_metrics(validation_df):
    log("\n天气匹配层级分析...")
    if "weather_match_level" not in validation_df.columns: log("  [跳过]"); return
    df = validation_df.copy()
    df.groupby("weather_match_level",dropna=False).size().reset_index(name="count").to_csv(
        OUTPUT_VALIDATION/"validation_weather_match_level_count_v2_6.csv",index=False,encoding="utf-8-sig")

def export_hourly_metrics(validation_df):
    log("\n按小时分析...")
    df = validation_df.copy()
    rows = [{"hour":h,**calc_metrics(df[df["hour"]==h])} for h in range(24)]
    pd.DataFrame(rows).to_csv(OUTPUT_VALIDATION/"validation_metrics_by_hour_v2_6.csv",index=False,encoding="utf-8-sig")

# =========================================================
# 19-21. 宽表 + Excel格式化 (保持不变)
# =========================================================
def build_validation_wide_for_one_user(g):
    g = g.copy().sort_values("datetime")
    g["base_date"] = np.where(g["datetime"].dt.hour==0,
                              (g["datetime"]-pd.Timedelta(days=1)).dt.normalize(),
                              g["datetime"].dt.normalize())
    g["base_date"] = pd.to_datetime(g["base_date"])
    day_list = sorted(g["base_date"].dropna().unique())
    rows = []
    for d in day_list:
        day = pd.to_datetime(d); one_day = g[g["base_date"]==day]
        acc = str(one_day["户号"].dropna().iloc[0]) if one_day["户号"].notna().any() else ""
        pr={"电量年月日":day.strftime("%Y%m%d"),"户号":acc,"类型":"预测"}
        ar={"电量年月日":day.strftime("%Y%m%d"),"户号":acc,"类型":"实际"}
        er={"电量年月日":day.strftime("%Y%m%d"),"户号":acc,"类型":"偏差"}
        ps,as_ = 0.0,0.0; pe,ae=True,True
        for h in range(1,25):
            target_dt = day+pd.Timedelta(days=1) if h==24 else day+pd.Timedelta(hours=h)
            hit = one_day[one_day["datetime"]==target_dt]
            pv = hit["final_pred_net_load"].iloc[0] if not hit.empty and "final_pred_net_load" in hit.columns and pd.notna(hit["final_pred_net_load"].iloc[0]) else np.nan
            av = hit["actual_load"].iloc[0] if not hit.empty and "actual_load" in hit.columns and pd.notna(hit["actual_load"].iloc[0]) else np.nan
            pr[f"{h}:00"]=pv; ar[f"{h}:00"]=av; er[f"{h}:00"]=calc_pct_error(pv,av)
            if pd.notna(pv): ps+=pv; pe=False
            if pd.notna(av): as_+=av; ae=False
        pr["合计"]=np.nan if pe else ps; ar["合计"]=np.nan if ae else as_
        er["合计"]=calc_pct_error(ps,as_) if (not pe and not ae) else np.nan
        rows.extend([pr,ar,er])
    return pd.DataFrame(rows)[["电量年月日","户号","类型"]+[f"{h}:00" for h in range(1,25)]+["合计"]]

def build_validation_total_wide(validation_df):
    df = validation_df.copy().sort_values("datetime")
    df["base_date"] = np.where(df["datetime"].dt.hour==0,
                               (df["datetime"]-pd.Timedelta(days=1)).dt.normalize(),
                               df["datetime"].dt.normalize())
    df["base_date"] = pd.to_datetime(df["base_date"])
    rows = []
    for d in sorted(df["base_date"].dropna().unique()):
        day = pd.to_datetime(d); one_day = df[df["base_date"]==day]
        pr={"电量年月日":day.strftime("%Y%m%d"),"类型":"预测"}
        ar={"电量年月日":day.strftime("%Y%m%d"),"类型":"实际"}
        er={"电量年月日":day.strftime("%Y%m%d"),"类型":"偏差"}
        ps,as_=0.0,0.0; pe,ae=True,True
        for h in range(1,25):
            target_dt = day+pd.Timedelta(days=1) if h==24 else day+pd.Timedelta(hours=h)
            hit = one_day[one_day["datetime"]==target_dt]
            pv = hit["final_pred_net_load"].sum(min_count=1) if not hit.empty and "final_pred_net_load" in hit.columns else np.nan
            av = hit["actual_load"].sum(min_count=1) if not hit.empty and "actual_load" in hit.columns else np.nan
            pr[f"{h}:00"]=pv; ar[f"{h}:00"]=av; er[f"{h}:00"]=calc_pct_error(pv,av)
            if pd.notna(pv): ps+=pv; pe=False
            if pd.notna(av): as_+=av; ae=False
        pr["合计"]=np.nan if pe else ps; ar["合计"]=np.nan if ae else as_
        er["合计"]=calc_pct_error(ps,as_) if (not pe and not ae) else np.nan
        rows.extend([pr,ar,er])
    return pd.DataFrame(rows)[["电量年月日","类型"]+[f"{h}:00" for h in range(1,25)]+["合计"]]

def format_validation_sheet(ws):
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col); c.fill = header_fill; c.font = bold_font; c.alignment = center_align
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        cl = get_column_letter(col)
        ws.column_dimensions[cl].width = {1:14,2:28,3:10}.get(col,12)
    for row in range(2, ws.max_row + 1):
        rt = ws.cell(row=row, column=3).value
        for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).alignment = center_align
        if rt == "预测":
            for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).fill = blue_fill
        elif rt == "实际":
            for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).fill = green_fill
        elif rt == "偏差":
            for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).fill = yellow_fill
            for col in range(4, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                if c.value is None: continue
                try: abs_val = abs(float(c.value))
                except: continue
                c.number_format = "0.00%"
                if abs_val >= 0.30: c.fill = red_fill
                elif abs_val >= 0.15: c.fill = orange_fill
                elif abs_val >= 0.05: c.fill = warn_fill

def format_total_validation_sheet(ws):
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col); c.fill = header_fill; c.font = bold_font; c.alignment = center_align
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        cl = get_column_letter(col)
        ws.column_dimensions[cl].width = {1:14,2:10}.get(col,12)
    for row in range(2, ws.max_row + 1):
        rt = ws.cell(row=row, column=2).value
        for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).alignment = center_align
        if rt == "预测":
            for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).fill = blue_fill
        elif rt == "实际":
            for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).fill = green_fill
        elif rt == "偏差":
            for col in range(1, ws.max_column + 1): ws.cell(row=row, column=col).fill = yellow_fill
            for col in range(3, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                if c.value is None: continue
                try: abs_val = abs(float(c.value))
                except: continue
                c.number_format = "0.00%"
                if abs_val >= 0.30: c.fill = red_fill
                elif abs_val >= 0.15: c.fill = orange_fill
                elif abs_val >= 0.05: c.fill = warn_fill

def export_validation_outputs(validation_df):
    log("\n导出验证宽表...")
    validation_df.to_csv(OUTPUT_VALIDATION / "validation_hourly_long_v2_6.csv", index=False, encoding="utf-8-sig")
    out_excel = OUTPUT_VALIDATION / "validation_hourly_wide_v2_6.xlsx"
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        for (uid, uname), g in validation_df.groupby(["用户编号", "用户名称"]):
            wide_df = build_validation_wide_for_one_user(g)
            sn = str(uname)[:31] if str(uname) else str(uid)
            wide_df.to_excel(writer, sheet_name=sn, index=False)
            format_validation_sheet(writer.book[sn])
    log(f"已导出: {out_excel.name}")
    total_wide = build_validation_total_wide(validation_df)
    total_xl = OUTPUT_VALIDATION / "validation_total_hourly_wide_v2_6.xlsx"
    with pd.ExcelWriter(total_xl, engine="openpyxl") as writer:
        total_wide.to_excel(writer, sheet_name="总量汇总", index=False)
        format_total_validation_sheet(writer.book["总量汇总"])
    log(f"已导出: {total_xl.name}")

# =========================================================
# 22. 主流程
# =========================================================
def main():
    log("=== 开始 V2.6 验证 ===")
    log(f"PREDICT_START = {PREDICT_START_TS}")
    log(f"PREDICT_END   = {PREDICT_END_TS}")
    log(f"LOW_LOAD_THRESHOLD       = {LOW_LOAD_THRESHOLD}")
    log(f"LOW_LOAD_PROBA_THRESHOLD = {LOW_LOAD_PROBA_THRESHOLD}")

    user_master_df = load_user_master()
    actual_df = load_actual(user_master_df)
    pred_df = load_prediction_long()
    weather_df = load_hourly_weather()

    validation_df = build_validation_long_fixed4(actual_df, pred_df, weather_df, user_master_df)

    # 验证分析
    export_metrics(validation_df)
    export_total_load_diagnostics(validation_df)
    export_daytime_metrics(validation_df)
    export_pv_user_diagnostics(validation_df)
    export_low_load_classifier_report(validation_df)
    export_actual_value_bucket_metrics(validation_df)
    export_weather_match_metrics(validation_df)
    export_hourly_metrics(validation_df)

    # 导出 Excel
    export_validation_outputs(validation_df)

    log("\n=== V2.6 验证完成 ===")

'''
def build_validation_long_fixed(actual_df, pred_df, weather_df, user_master_df):
    """用真实天气重算 actual_total_load，避免把预测 PV 混进验证口径。"""
    log("合并实际与预测（fixed）...")
    merge_cols = ["鐢ㄦ埛缂栧彿","鐢ㄦ埛鍚嶇О","鎴峰彿","datetime"]
    df = actual_df.merge(pred_df, on=merge_cols, how="left", suffixes=("","_pred"))
    if "鏄惁鏈夊厜浼廮flag_pred" in df.columns:
        df["鏄惁鏈夊厜浼廮flag"] = df["鏄惁鏈夊厜浼廮flag"].fillna(df["鏄惁鏈夊厜浼廮flag_pred"])
        df = df.drop(columns=["鏄惁鏈夊厜浼廮flag_pred"])

    user_city_map = {}
    for _, u in user_master_df.iterrows():
        uid = u.get("鐢ㄦ埛缂栧彿")
        if pd.notna(uid):
            user_city_map[uid] = extract_city(u.get("鎵€鍦ㄥ競"))

    df["pv_est"] = 0.0
    mask_pv = df["鏄惁鏈夊厜浼廮flag"] == 1
    if mask_pv.any() and weather_df is not None and not weather_df.empty:
        w = weather_df.copy()
        w["datetime"] = pd.to_datetime(w["datetime"]).dt.floor("h")
        for idx, row in df[mask_pv].iterrows():
            city = user_city_map.get(row["鐢ㄦ埛缂栧彿"])
            if not city:
                continue
            hit = w[(w["鎵€鍦ㄥ競_norm"] == city) & (w["datetime"] == row["datetime"])]
            if hit.empty:
                continue
            df.at[idx, "pv_est"] = estimate_pv_generation(
                hit["shortwave_radiation"].iloc[0],
                hit["temperature"].iloc[0],
                row["鍏변紡瀹归噺(MW)"],
            )

    df["actual_total_load"] = df["actual_load"]
    if mask_pv.any():
        df.loc[mask_pv, "actual_total_load"] = df.loc[mask_pv, "actual_load"] + df.loc[mask_pv, "pv_est"]

    df["net_error"] = df["final_pred_net_load"] - df["actual_load"]
    df["net_abs_error"] = df["net_error"].abs()
    df["total_error"] = df["pred_total_load"] - df["actual_total_load"]
    df["total_abs_error"] = df["total_error"].abs()
    df["ape_net"] = np.where(
        (df["actual_load"].notna()) & (df["actual_load"] != 0) & (df["final_pred_net_load"].notna()),
        df["net_abs_error"] / df["actual_load"],
        np.nan,
    )
    df["ape_total"] = np.where(
        (df["actual_total_load"].notna()) & (df["actual_total_load"] != 0) & (df["pred_total_load"].notna()),
        df["total_abs_error"] / df["actual_total_load"],
        np.nan,
    )
    df["status"] = np.where(df["final_pred_net_load"].isna(), df["predict_status"].fillna("鏈娴?), "宸查娴?")
    df["status"] = np.where(df["final_pred_net_load"].isna(), df["predict_status"].fillna("unpredicted"), "predicted")
    df["hour"] = pd.to_datetime(df["datetime"]).dt.hour
    df["is_daytime_8_19"] = df["hour"].isin(list(range(8, 20))).astype(int)
    df["is_low_load_actual"] = (df["actual_total_load"] < LOW_LOAD_THRESHOLD).astype(int)
    return df.sort_values(["鐢ㄦ埛缂栧彿","datetime"]).reset_index(drop=True)
'''
'''
def build_validation_long_fixed2(actual_df, pred_df, weather_df, user_master_df):
    """用真实天气重算 actual_total_load，避免把预测 PV 混进验证口径。"""
    log("合并实际与预测（fixed2）...")

    def pick_col(df, keys):
        for c in df.columns:
            sc = str(c)
            if any(k in sc for k in keys):
                return c
        raise KeyError(f"未找到匹配列: {keys}")

    uid_col = pick_col(actual_df, ["缂栧彿"])
    name_col = pick_col(actual_df, ["鍚嶇О"])
    account_col = pick_col(actual_df, ["鎴峰彿"])

    merge_cols = [uid_col, name_col, account_col, "datetime"]
    df = actual_df.merge(pred_df, on=merge_cols, how="left", suffixes=("", "_pred"))

    flag_col = pick_col(df, ["鏈夊厜浼", "光伏", "pv"])
    if f"{flag_col}_pred" in df.columns:
        df[flag_col] = df[flag_col].fillna(df[f"{flag_col}_pred"])
        df = df.drop(columns=[f"{flag_col}_pred"])

    city_col = pick_col(user_master_df, ["鍦ㄥ競"])
    user_city_map = {}
    for _, u in user_master_df.iterrows():
        uid = u[uid_col] if uid_col in u.index else u.get(uid_col)
        if pd.notna(uid):
            user_city_map[uid] = extract_city(u[city_col])

    weather_city_col = pick_col(weather_df, ["鍦ㄥ競_norm", "city"])

    cap_col = pick_col(df, ["MW"])
    df["pv_est"] = 0.0
    mask_pv = df[flag_col] == 1
    if mask_pv.any() and weather_df is not None and not weather_df.empty:
        w = weather_df.copy()
        w["datetime"] = pd.to_datetime(w["datetime"]).dt.floor("h")
        for idx, row in df[mask_pv].iterrows():
            city = user_city_map.get(row[uid_col])
            if not city:
                continue
            hit = w[(w[weather_city_col] == city) & (w["datetime"] == row["datetime"])]
            if hit.empty:
                continue
            df.at[idx, "pv_est"] = estimate_pv_generation(
                hit["shortwave_radiation"].iloc[0],
                hit["temperature"].iloc[0],
                row[cap_col],
            )

    df["actual_total_load"] = df["actual_load"]
    if mask_pv.any():
        df.loc[mask_pv, "actual_total_load"] = df.loc[mask_pv, "actual_load"] + df.loc[mask_pv, "pv_est"]

    df["net_error"] = df["final_pred_net_load"] - df["actual_load"]
    df["net_abs_error"] = df["net_error"].abs()
    df["total_error"] = df["pred_total_load"] - df["actual_total_load"]
    df["total_abs_error"] = df["total_error"].abs()
    df["ape_net"] = np.where(
        (df["actual_load"].notna()) & (df["actual_load"] != 0) & (df["final_pred_net_load"].notna()),
        df["net_abs_error"] / df["actual_load"],
        np.nan,
    )
    df["ape_total"] = np.where(
        (df["actual_total_load"].notna()) & (df["actual_total_load"] != 0) & (df["pred_total_load"].notna()),
        df["total_abs_error"] / df["actual_total_load"],
        np.nan,
    )
    df["status"] = np.where(df["final_pred_net_load"].isna(), df["predict_status"].fillna("鏈娴?), "宸查娴?")
    df["hour"] = pd.to_datetime(df["datetime"]).dt.hour
    df["is_daytime_8_19"] = df["hour"].isin(list(range(8, 20))).astype(int)
    df["is_low_load_actual"] = (df["actual_total_load"] < LOW_LOAD_THRESHOLD).astype(int)
    return df.sort_values([uid_col, "datetime"]).reset_index(drop=True)

'''
'''
def build_validation_long_fixed3(actual_df, pred_df, weather_df, user_master_df):
    """Validate against net load, and rebuild actual total load from actual weather instead of predicted PV."""
    merge_id_cols = [c for c in list(actual_df.columns[:3]) if c in pred_df.columns]
    merge_cols = merge_id_cols + ["datetime"]
    df = actual_df.merge(pred_df, on=merge_cols, how="left", suffixes=("", "_pred"))

    flag_candidates = [c for c in actual_df.columns if c not in merge_cols + ["actual_load"]]
    flag_col = flag_candidates[0] if flag_candidates else None
    cap_col = flag_candidates[1] if len(flag_candidates) > 1 else None
    if flag_col and f"{flag_col}_pred" in df.columns:
        df[flag_col] = df[flag_col].fillna(df[f"{flag_col}_pred"])
        df = df.drop(columns=[f"{flag_col}_pred"])

    user_master_uid_col = user_master_df.columns[0]
    user_master_city_col = user_master_df.columns[2]
    weather_city_col = weather_df.columns[0] if weather_df is not None and not weather_df.empty else None

    city_by_uid = {}
    for _, row in user_master_df.iterrows():
        uid = row[user_master_uid_col]
        if pd.notna(uid):
            city_by_uid[uid] = extract_city(row[user_master_city_col])

    df["pv_est"] = 0.0
    if flag_col and cap_col and weather_df is not None and not weather_df.empty:
        weather = weather_df.copy()
        weather["datetime"] = pd.to_datetime(weather["datetime"]).dt.floor("h")
        pv_mask = df[flag_col] == 1
        for idx, row in df[pv_mask].iterrows():
            uid = row[merge_id_cols[0]]
            city = city_by_uid.get(uid)
            if not city:
                continue
            match = weather[(weather[weather_city_col] == city) & (weather["datetime"] == row["datetime"])]
            if match.empty:
                continue
            df.at[idx, "pv_est"] = estimate_pv_generation(
                match["shortwave_radiation"].iloc[0],
                match["temperature"].iloc[0],
                row[cap_col],
            )

    df["actual_total_load"] = df["actual_load"]
    if flag_col:
        pv_mask = df[flag_col] == 1
        df.loc[pv_mask, "actual_total_load"] = df.loc[pv_mask, "actual_load"] + df.loc[pv_mask, "pv_est"]

    df["net_error"] = df["final_pred_net_load"] - df["actual_load"]
    df["net_abs_error"] = df["net_error"].abs()
    df["total_error"] = df["pred_total_load"] - df["actual_total_load"]
    df["total_abs_error"] = df["total_error"].abs()
    df["ape_net"] = np.where(
        (df["actual_load"].notna()) & (df["actual_load"] != 0) & (df["final_pred_net_load"].notna()),
        df["net_abs_error"] / df["actual_load"],
        np.nan,
    )
    df["ape_total"] = np.where(
        (df["actual_total_load"].notna()) & (df["actual_total_load"] != 0) & (df["pred_total_load"].notna()),
        df["total_abs_error"] / df["actual_total_load"],
        np.nan,
    )
    df["status"] = np.where(df["final_pred_net_load"].isna(), df["predict_status"].fillna("未预测"), "已预测")
    df["hour"] = pd.to_datetime(df["datetime"]).dt.hour
    df["is_daytime_8_19"] = df["hour"].isin(list(range(8, 20))).astype(int)
    df["is_low_load_actual"] = (df["actual_total_load"] < LOW_LOAD_THRESHOLD).astype(int)
    return df.sort_values([merge_id_cols[0], "datetime"]).reset_index(drop=True)

'''
def build_validation_long_fixed4(actual_df, pred_df, weather_df, user_master_df):
    """Use actual weather to rebuild total load validation with schema-agnostic column detection."""
    id_cols = [c for c in list(actual_df.columns[:3]) if c in pred_df.columns]
    merge_cols = id_cols + ["datetime"]
    df = actual_df.merge(pred_df, on=merge_cols, how="left", suffixes=("", "_pred"))

    actual_only_cols = [c for c in actual_df.columns if c not in merge_cols + ["actual_load"]]
    flag_col = actual_only_cols[0] if len(actual_only_cols) >= 1 else None
    cap_col = actual_only_cols[1] if len(actual_only_cols) >= 2 else None

    if flag_col and f"{flag_col}_pred" in df.columns:
        df[flag_col] = df[flag_col].fillna(df[f"{flag_col}_pred"])
        df = df.drop(columns=[f"{flag_col}_pred"])

    uid_col = user_master_df.columns[0]
    city_col = user_master_df.columns[2]
    city_by_uid = {}
    for _, row in user_master_df.iterrows():
        uid = row[uid_col]
        if pd.notna(uid):
            city_by_uid[uid] = extract_city(row[city_col])

    df["pv_est"] = 0.0
    if weather_df is not None and not weather_df.empty and flag_col and cap_col:
        weather = weather_df.copy()
        weather["datetime"] = pd.to_datetime(weather["datetime"]).dt.floor("h")
        weather_city_col = weather.columns[0]
        pv_mask = df[flag_col] == 1
        for idx, row in df[pv_mask].iterrows():
            uid = row[id_cols[0]]
            city = city_by_uid.get(uid)
            if not city:
                continue
            hit = weather[(weather[weather_city_col] == city) & (weather["datetime"] == row["datetime"])]
            if hit.empty:
                continue
            df.at[idx, "pv_est"] = estimate_pv_generation(
                hit["shortwave_radiation"].iloc[0],
                hit["temperature"].iloc[0],
                row[cap_col],
            )

    df["actual_total_load"] = df["actual_load"]
    if flag_col:
        pv_mask = df[flag_col] == 1
        df.loc[pv_mask, "actual_total_load"] = df.loc[pv_mask, "actual_load"] + df.loc[pv_mask, "pv_est"]

    df["net_error"] = df["final_pred_net_load"] - df["actual_load"]
    df["net_abs_error"] = df["net_error"].abs()
    df["total_error"] = df["pred_total_load"] - df["actual_total_load"]
    df["total_abs_error"] = df["total_error"].abs()
    df["ape_net"] = np.where(
        (df["actual_load"].notna()) & (df["actual_load"] != 0) & (df["final_pred_net_load"].notna()),
        df["net_abs_error"] / df["actual_load"],
        np.nan,
    )
    df["ape_total"] = np.where(
        (df["actual_total_load"].notna()) & (df["actual_total_load"] != 0) & (df["pred_total_load"].notna()),
        df["total_abs_error"] / df["actual_total_load"],
        np.nan,
    )
    df["status"] = np.where(df["final_pred_net_load"].isna(), df["predict_status"].fillna("unpredicted"), "predicted")
    df["hour"] = pd.to_datetime(df["datetime"]).dt.hour
    df["is_daytime_8_19"] = df["hour"].isin(list(range(8, 20))).astype(int)
    df["is_low_load_actual"] = (df["actual_total_load"] < LOW_LOAD_THRESHOLD).astype(int)
    return df.sort_values([id_cols[0], "datetime"]).reset_index(drop=True)

if __name__ == "__main__":
    main()
