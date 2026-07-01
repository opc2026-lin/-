# -*- coding: utf-8 -*-

import re
import warnings
import numpy as np
import pandas as pd

from pathlib import Path
from sklearn.metrics import mean_absolute_error, mean_squared_error
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# =========================================================
# 1. 参数配置
# 与训练/预测保持一致
# =========================================================
TARGET_USER_NAME = "福建俊杰新材料科技股份有限公司"

PREDICT_START = "2026-05-01 00:00:00"   # 左闭
PREDICT_END = "2026-06-01 00:00:00"     # 右开


# =========================================================
# 2. 路径配置
# =========================================================
BASE_DIR = Path(__file__).resolve().parent

USER_MASTER_PATH = BASE_DIR / "input" / "user_master" / "01_用户主档案表.csv"
LOAD_DIR = BASE_DIR / "input" / "load"

OUTPUT_MODEL = BASE_DIR / "output" / "model"
OUTPUT_PREDICTION = BASE_DIR / "output" / "prediction"
OUTPUT_VALIDATION = BASE_DIR / "output" / "validation"
OUTPUT_LOGS = BASE_DIR / "output" / "logs"

for p in [OUTPUT_MODEL, OUTPUT_PREDICTION, OUTPUT_VALIDATION, OUTPUT_LOGS]:
    p.mkdir(parents=True, exist_ok=True)


# =========================================================
# 3. 日志
# =========================================================
SAFE_USER_NAME = TARGET_USER_NAME.replace("/", "_").replace("\\", "_")
LOG_FILE = OUTPUT_LOGS / f"03_validate_single_user_v6_{SAFE_USER_NAME}.log"


def log(msg):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(str(msg) + "\n")


with open(LOG_FILE, "w", encoding="utf-8") as f:
    f.write("=== 单用户V6验证日志 ===\n")


# =========================================================
# 4. 时间范围
# =========================================================
PREDICT_START_TS = pd.Timestamp(PREDICT_START)
PREDICT_END_TS = pd.Timestamp(PREDICT_END)


# =========================================================
# 5. 通用函数
# =========================================================
def normalize_text(x):
    if pd.isna(x):
        return None
    x = str(x).strip()
    x = x.replace("　", "").replace(" ", "")
    return x


def safe_read_table(path):
    path = str(path)
    if path.lower().endswith(".csv"):
        try:
            return pd.read_csv(path, encoding="utf-8-sig")
        except Exception:
            try:
                return pd.read_csv(path, encoding="gbk")
            except Exception:
                return pd.read_csv(path)
    else:
        return pd.read_excel(path)


def normalize_load_sheet(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    header_row_idx = None
    for i in range(min(len(df), 10)):
        row_values = [str(x).strip() for x in df.iloc[i].tolist()]
        joined = "".join(row_values)
        if ("电量年月日" in joined) and ("户号" in joined):
            header_row_idx = i
            break

    if header_row_idx is not None:
        new_header = [str(x).strip() for x in df.iloc[header_row_idx].tolist()]
        df = df.iloc[header_row_idx + 1:].copy()
        df.columns = new_header
    else:
        if "电量年月日" not in df.columns or "户号" not in df.columns:
            raise ValueError("未识别到负荷表头")

    df.columns = [str(c).strip() for c in df.columns]
    return df


def calc_pct_error(pred, actual):
    if pd.isna(pred) or pd.isna(actual) or actual == 0:
        return np.nan
    return (pred - actual) / actual


def calc_metrics(df, actual_col="actual_load", pred_col="predicted_load"):
    tmp = df.dropna(subset=[actual_col, pred_col]).copy()
    if tmp.empty:
        return {"mae": np.nan, "rmse": np.nan, "mape": np.nan, "sample_count": 0}

    mae = mean_absolute_error(tmp[actual_col], tmp[pred_col])
    rmse = np.sqrt(mean_squared_error(tmp[actual_col], tmp[pred_col]))

    ape = np.where(
        (tmp[actual_col].notna()) & (tmp[actual_col] != 0),
        np.abs(tmp[pred_col] - tmp[actual_col]) / tmp[actual_col],
        np.nan
    )
    mape = np.nanmean(ape)

    return {
        "mae": mae,
        "rmse": rmse,
        "mape": mape,
        "sample_count": len(tmp)
    }


# =========================================================
# 6. 读取主档案
# =========================================================
def load_user_master():
    df = safe_read_table(USER_MASTER_PATH)
    df.columns = [str(c).strip() for c in df.columns]
    df["用户名称_norm"] = df["用户名称"].apply(normalize_text)
    return df


# =========================================================
# 7. 读取单用户配置
# =========================================================
def load_single_user_run_config():
    config_path = OUTPUT_MODEL / f"single_user_run_config_v6_{SAFE_USER_NAME}.csv"
    if not config_path.exists():
        raise FileNotFoundError(f"未找到单用户V6配置文件：{config_path.name}")
    return pd.read_csv(config_path, encoding="utf-8-sig").iloc[0]


# =========================================================
# 8. 读取单用户实际值
# =========================================================
def parse_actual_sheet(file_path, user_info):
    target_sheet = f"{str(PREDICT_START_TS.year)[2:]}.{PREDICT_START_TS.month}"

    raw = pd.read_excel(file_path, sheet_name=target_sheet, header=0)
    df = normalize_load_sheet(raw)

    required = ["电量年月日", "户号"]
    for c in required:
        if c not in df.columns:
            raise ValueError(f"{Path(file_path).name}-{target_sheet} 缺少字段: {c}")

    hour_cols = [f"{h}:00" for h in range(1, 25)]
    for c in hour_cols:
        if c not in df.columns:
            raise ValueError(f"{Path(file_path).name}-{target_sheet} 缺少小时列: {c}")

    df["电量年月日"] = df["电量年月日"].astype(str).str.strip()
    df = df[df["电量年月日"].str.fullmatch(r"\d{8}", na=False)].copy()
    if df.empty:
        return None

    long_df = df.melt(
        id_vars=["电量年月日", "户号"],
        value_vars=hour_cols,
        var_name="hour_str",
        value_name="actual_load"
    )

    long_df["date"] = pd.to_datetime(long_df["电量年月日"], format="%Y%m%d", errors="coerce")
    long_df["actual_load"] = pd.to_numeric(long_df["actual_load"], errors="coerce")
    long_df["户号"] = long_df["户号"].astype(str).str.strip()

    def build_datetime(row):
        d = row["date"]
        h = int(str(row["hour_str"]).split(":")[0])
        if pd.isna(d):
            return pd.NaT
        if h == 24:
            return d + pd.Timedelta(days=1)
        return d + pd.Timedelta(hours=h)

    long_df["datetime"] = long_df.apply(build_datetime, axis=1)
    long_df["用户编号"] = user_info["用户编号"]
    long_df["用户名称"] = user_info["用户名称"]

    long_df = long_df[
        (long_df["datetime"] >= PREDICT_START_TS) &
        (long_df["datetime"] < PREDICT_END_TS)
    ].copy()

    return long_df[["用户编号", "用户名称", "户号", "datetime", "actual_load"]].dropna(subset=["datetime"])


def load_single_user_actual(user_master_df):
    target_norm = normalize_text(TARGET_USER_NAME)
    matched = user_master_df[user_master_df["用户名称_norm"] == target_norm]

    if matched.empty:
        raise ValueError(f"主档案表中未找到目标用户：{TARGET_USER_NAME}")

    user_info = matched.iloc[0]

    files = list(Path(LOAD_DIR).glob("*.xlsx")) + list(Path(LOAD_DIR).glob("*.xls"))
    target_file = None
    for fp in files:
        if normalize_text(fp.stem) == target_norm:
            target_file = fp
            break

    if target_file is None:
        raise FileNotFoundError(f"未找到该用户负荷文件：{TARGET_USER_NAME}")

    return parse_actual_sheet(target_file, user_info)


# =========================================================
# 9. 读取单用户预测结果
# =========================================================
def load_single_user_prediction_long():
    pred_path = OUTPUT_PREDICTION / f"single_user_predict_long_v6_{SAFE_USER_NAME}.csv"
    if not pred_path.exists():
        raise FileNotFoundError(f"未找到单用户V6预测长表：{pred_path.name}")

    pred_df = pd.read_csv(pred_path, encoding="utf-8-sig")
    pred_df["datetime"] = pd.to_datetime(pred_df["datetime"], errors="coerce")

    keep_cols = [
        "用户编号", "用户名称", "户号", "datetime",
        "predicted_load", "predict_status",
        "pred_daytime_total_load", "pred_day_type",
        "weather_match_level"
    ]
    keep_cols = [c for c in keep_cols if c in pred_df.columns]
    pred_df = pred_df[keep_cols].copy()

    if "predict_status" not in pred_df.columns:
        pred_df["predict_status"] = "已预测"

    return pred_df


# =========================================================
# 10. 合并验证
# =========================================================
def build_validation_long(actual_df, pred_df):
    log("合并实际与预测...")

    merge_cols = ["用户编号", "用户名称", "户号", "datetime"]
    df = actual_df.merge(pred_df, on=merge_cols, how="left")

    df["error"] = df["predicted_load"] - df["actual_load"]
    df["abs_error"] = df["error"].abs()
    df["pct_error"] = df.apply(lambda x: calc_pct_error(x["predicted_load"], x["actual_load"]), axis=1)

    df["ape"] = np.where(
        (df["actual_load"].notna()) & (df["actual_load"] != 0) & (df["predicted_load"].notna()),
        df["abs_error"] / df["actual_load"],
        np.nan
    )

    df["status"] = np.where(
        df["predicted_load"].isna(),
        df["predict_status"].fillna("未预测"),
        "已预测"
    )

    df["hour"] = pd.to_datetime(df["datetime"]).dt.hour
    df["is_daytime_8_19"] = df["hour"].isin(list(range(8, 20))).astype(int)
    df["date_day"] = pd.to_datetime(df["datetime"]).dt.normalize()

    return df.sort_values("datetime").reset_index(drop=True)


# =========================================================
# 11. 导出总体指标
# =========================================================
def export_metrics(validation_df):
    total_metrics = calc_metrics(validation_df)

    pd.DataFrame([{
        "scope": "single_user_total_v6",
        "mae": total_metrics["mae"],
        "rmse": total_metrics["rmse"],
        "mape": total_metrics["mape"],
        "sample_count": total_metrics["sample_count"]
    }]).to_csv(
        OUTPUT_VALIDATION / f"single_user_validation_metrics_total_v6_{SAFE_USER_NAME}.csv",
        index=False,
        encoding="utf-8-sig"
    )

    log(f"单用户V6总体验证 MAE: {total_metrics['mae']:.4f}")
    log(f"单用户V6总体验证 RMSE: {total_metrics['rmse']:.4f}")
    log(f"单用户V6总体验证 MAPE: {total_metrics['mape']:.4%}")


# =========================================================
# 12. 白天专项指标
# =========================================================
def export_daytime_metrics(validation_df):
    log("导出单用户V6白天专项验证指标（8:00~19:00）...")

    day_df = validation_df[validation_df["is_daytime_8_19"] == 1].copy()
    total_metrics = calc_metrics(day_df)

    pd.DataFrame([{
        "scope": "single_user_daytime_8_19_v6",
        "mae": total_metrics["mae"],
        "rmse": total_metrics["rmse"],
        "mape": total_metrics["mape"],
        "sample_count": total_metrics["sample_count"]
    }]).to_csv(
        OUTPUT_VALIDATION / f"single_user_validation_daytime_metrics_total_v6_{SAFE_USER_NAME}.csv",
        index=False,
        encoding="utf-8-sig"
    )

    log(f"单用户V6白天 MAE: {total_metrics['mae']:.4f}")
    log(f"单用户V6白天 RMSE: {total_metrics['rmse']:.4f}")
    log(f"单用户V6白天 MAPE: {total_metrics['mape']:.4%}")


# =========================================================
# 13. 按真实值分层
# =========================================================
def export_actual_value_bucket_metrics(validation_df):
    log("导出单用户V6按真实值分层误差分析...")

    df = validation_df.copy()

    def get_bucket(x):
        if pd.isna(x):
            return "missing"
        elif x < 20:
            return "<20"
        elif x < 50:
            return "20-50"
        elif x < 100:
            return "50-100"
        elif x < 200:
            return "100-200"
        else:
            return ">=200"

    df["actual_bucket"] = df["actual_load"].apply(get_bucket)
    df_eval = df[df["actual_bucket"] != "missing"].copy()
    bucket_order = ["<20", "20-50", "50-100", "100-200", ">=200"]

    rows_all = []
    for bucket, g in df_eval.groupby("actual_bucket"):
        m = calc_metrics(g)
        rows_all.append({
            "actual_bucket": bucket,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    out_all = pd.DataFrame(rows_all)
    if not out_all.empty:
        out_all["actual_bucket"] = pd.Categorical(out_all["actual_bucket"], categories=bucket_order, ordered=True)
        out_all = out_all.sort_values("actual_bucket")

    out_all.to_csv(
        OUTPUT_VALIDATION / f"single_user_validation_metrics_by_actual_bucket_v6_{SAFE_USER_NAME}.csv",
        index=False,
        encoding="utf-8-sig"
    )

    day_df = df_eval[df_eval["is_daytime_8_19"] == 1].copy()

    rows_day = []
    for bucket, g in day_df.groupby("actual_bucket"):
        m = calc_metrics(g)
        rows_day.append({
            "actual_bucket": bucket,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    out_day = pd.DataFrame(rows_day)
    if not out_day.empty:
        out_day["actual_bucket"] = pd.Categorical(out_day["actual_bucket"], categories=bucket_order, ordered=True)
        out_day = out_day.sort_values("actual_bucket")

    out_day.to_csv(
        OUTPUT_VALIDATION / f"single_user_validation_daytime_metrics_by_actual_bucket_v6_{SAFE_USER_NAME}.csv",
        index=False,
        encoding="utf-8-sig"
    )

    log("单用户V6按真实值分层误差分析导出完成")


# =========================================================
# 14. 日级总量与日类型分析
# =========================================================
def export_day_load_and_type_metrics(validation_df):
    log("导出单用户V6日级总量 / 日类型分析...")

    # 实际日级白天总量
    actual_day = (
        validation_df[validation_df["is_daytime_8_19"] == 1]
        .groupby("date_day", as_index=False)["actual_load"]
        .sum()
        .rename(columns={"actual_load": "actual_daytime_total_load"})
    )

    pred_day = (
        validation_df[["date_day", "pred_daytime_total_load", "pred_day_type"]]
        .drop_duplicates(subset=["date_day"])
        .copy()
    )

    merged = actual_day.merge(pred_day, on="date_day", how="left")

    # 实际日类型（按预测值同样阈值逻辑）
    q30 = merged["actual_daytime_total_load"].quantile(0.3)
    q70 = merged["actual_daytime_total_load"].quantile(0.7)

    def map_day_type(x):
        if pd.isna(x):
            return "mid_day"
        elif x <= q30:
            return "low_day"
        elif x <= q70:
            return "mid_day"
        else:
            return "high_day"

    merged["actual_day_type"] = merged["actual_daytime_total_load"].apply(map_day_type)

    merged.to_csv(
        OUTPUT_VALIDATION / f"single_user_day_compare_v6_{SAFE_USER_NAME}.csv",
        index=False,
        encoding="utf-8-sig"
    )

    # 日级白天总量误差
    day_eval = merged.dropna(subset=["actual_daytime_total_load", "pred_daytime_total_load"]).copy()
    if not day_eval.empty:
        mae = mean_absolute_error(day_eval["actual_daytime_total_load"], day_eval["pred_daytime_total_load"])
        rmse = np.sqrt(mean_squared_error(day_eval["actual_daytime_total_load"], day_eval["pred_daytime_total_load"]))
        ape = np.where(
            day_eval["actual_daytime_total_load"] != 0,
            np.abs(day_eval["pred_daytime_total_load"] - day_eval["actual_daytime_total_load"]) / day_eval["actual_daytime_total_load"],
            np.nan
        )
        mape = np.nanmean(ape)

        pd.DataFrame([{
            "mae": mae,
            "rmse": rmse,
            "mape": mape,
            "sample_count": len(day_eval)
        }]).to_csv(
            OUTPUT_VALIDATION / f"single_user_daytime_total_load_metrics_v6_{SAFE_USER_NAME}.csv",
            index=False,
            encoding="utf-8-sig"
        )

        log(f"单用户V6日级白天总量 MAE: {mae:.4f}")
        log(f"单用户V6日级白天总量 RMSE: {rmse:.4f}")
        log(f"单用户V6日级白天总量 MAPE: {mape:.4%}")

    # 日类型映射准确率
    valid_type = merged.dropna(subset=["actual_day_type", "pred_day_type"]).copy()
    if not valid_type.empty:
        acc = (valid_type["actual_day_type"] == valid_type["pred_day_type"]).mean()

        pd.DataFrame([{
            "day_type_accuracy": acc,
            "sample_count": len(valid_type)
        }]).to_csv(
            OUTPUT_VALIDATION / f"single_user_day_type_accuracy_v6_{SAFE_USER_NAME}.csv",
            index=False,
            encoding="utf-8-sig"
        )

        log(f"单用户V6日类型映射准确率: {acc:.4f}")


# =========================================================
# 15. 天气匹配层级分析
# =========================================================
def export_weather_match_metrics(validation_df):
    log("导出单用户V6天气匹配层级分析...")

    if "weather_match_level" not in validation_df.columns:
        log("[警告] 预测结果中不存在 weather_match_level，跳过天气匹配层级分析")
        return

    df = validation_df.copy()

    count_df = (
        df.groupby("weather_match_level", dropna=False)
        .size()
        .reset_index(name="sample_count")
    )
    count_df.to_csv(
        OUTPUT_VALIDATION / f"single_user_weather_match_level_count_v6_{SAFE_USER_NAME}.csv",
        index=False,
        encoding="utf-8-sig"
    )

    metric_rows = []
    for level, g in df.groupby("weather_match_level", dropna=False):
        m = calc_metrics(g)
        metric_rows.append({
            "weather_match_level": level,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    pd.DataFrame(metric_rows).to_csv(
        OUTPUT_VALIDATION / f"single_user_weather_match_level_metrics_v6_{SAFE_USER_NAME}.csv",
        index=False,
        encoding="utf-8-sig"
    )

    log("单用户V6天气匹配层级分析导出完成")


# =========================================================
# 16. 构建宽表（每天3行）
# =========================================================
def build_validation_wide(validation_df):
    g = validation_df.copy().sort_values("datetime")

    g["base_date"] = np.where(
        g["datetime"].dt.hour == 0,
        (g["datetime"] - pd.Timedelta(days=1)).dt.normalize(),
        g["datetime"].dt.normalize()
    )
    g["base_date"] = pd.to_datetime(g["base_date"])

    day_list = sorted(g["base_date"].dropna().unique())
    rows = []

    for d in day_list:
        day = pd.to_datetime(d)
        one_day = g[g["base_date"] == day].copy()

        account_value = one_day["户号"].dropna().astype(str).iloc[0] if one_day["户号"].notna().any() else ""

        pred_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "预测"}
        actual_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "实际"}
        error_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "偏差"}

        pred_sum = 0.0
        actual_sum = 0.0
        pred_empty = True
        actual_empty = True

        for h in range(1, 25):
            target_dt = day + pd.Timedelta(days=1) if h == 24 else day + pd.Timedelta(hours=h)
            hit = one_day[one_day["datetime"] == target_dt]

            if hit.empty:
                pred_val = np.nan
                actual_val = np.nan
            else:
                pred_val = hit["predicted_load"].iloc[0] if "predicted_load" in hit.columns else np.nan
                actual_val = hit["actual_load"].iloc[0] if "actual_load" in hit.columns else np.nan

            pct_val = calc_pct_error(pred_val, actual_val)

            pred_row[f"{h}:00"] = pred_val
            actual_row[f"{h}:00"] = actual_val
            error_row[f"{h}:00"] = pct_val

            if pd.notna(pred_val):
                pred_sum += pred_val
                pred_empty = False
            if pd.notna(actual_val):
                actual_sum += actual_val
                actual_empty = False

        pred_row["合计"] = np.nan if pred_empty else pred_sum
        actual_row["合计"] = np.nan if actual_empty else actual_sum
        error_row["合计"] = calc_pct_error(pred_sum, actual_sum) if (not pred_empty and not actual_empty) else np.nan

        rows.extend([pred_row, actual_row, error_row])

    final_cols = ["电量年月日", "户号", "类型"] + [f"{h}:00" for h in range(1, 25)] + ["合计"]
    return pd.DataFrame(rows)[final_cols]


# =========================================================
# 17. Excel格式化
# =========================================================
def format_validation_sheet(ws):
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align

    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 14
        elif col == 2:
            ws.column_dimensions[col_letter].width = 28
        elif col == 3:
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 12

    for row in range(2, max_row + 1):
        row_type = ws.cell(row=row, column=3).value

        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).alignment = center_align

        if row_type == "预测":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = blue_fill

        elif row_type == "实际":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = green_fill

        elif row_type == "偏差":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

            for col in range(4, max_col + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value

                if val is None or val == "":
                    continue

                try:
                    abs_val = abs(float(val))
                except Exception:
                    continue

                cell.number_format = "0.00%"

                if abs_val >= 0.30:
                    cell.fill = red_fill
                elif abs_val >= 0.15:
                    cell.fill = orange_fill
                elif abs_val >= 0.05:
                    cell.fill = warn_fill


# =========================================================
# 18. 导出宽表
# =========================================================
def export_validation_outputs(validation_df):
    log("导出单用户V6验证结果...")
    validation_df.to_csv(
        OUTPUT_VALIDATION / f"single_user_validation_hourly_long_v6_{SAFE_USER_NAME}.csv",
        index=False,
        encoding="utf-8-sig"
    )

    wide_df = build_validation_wide(validation_df)
    out_excel = OUTPUT_VALIDATION / f"single_user_validation_hourly_wide_v6_{SAFE_USER_NAME}.xlsx"

    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        sheet_name = f"{str(PREDICT_START_TS.year)[2:]}.{PREDICT_START_TS.month}"
        wide_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        ws = writer.book[sheet_name[:31]]
        format_validation_sheet(ws)

    log(f"已导出：{out_excel.name}")


# =========================================================
# 19. 主流程
# =========================================================
def main():
    log("=== 开始单用户V6验证 ===")
    log(f"目标用户 = {TARGET_USER_NAME}")
    log(f"PREDICT_START = {PREDICT_START_TS}")
    log(f"PREDICT_END   = {PREDICT_END_TS}")

    user_master_df = load_user_master()
    actual_df = load_single_user_actual(user_master_df)
    pred_df = load_single_user_prediction_long()

    validation_df = build_validation_long(actual_df, pred_df)

    export_metrics(validation_df)
    export_daytime_metrics(validation_df)
    export_actual_value_bucket_metrics(validation_df)
    export_day_load_and_type_metrics(validation_df)
    export_weather_match_metrics(validation_df)
    export_validation_outputs(validation_df)

    log("=== 单用户V6验证完成 ===")


if __name__ == "__main__":
    main()