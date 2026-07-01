# -*- coding: utf-8 -*-

import re
import glob
import warnings
import numpy as np
import pandas as pd

from pathlib import Path
from sklearn.metrics import mean_absolute_error, mean_squared_error
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# =========================================================
# 1. 路径配置
# 输出文件名固定，方便你手动备份不同版本结果
# =========================================================
BASE_DIR = Path(__file__).resolve().parent

USER_MASTER_PATH = BASE_DIR / "input" / "user_master" / "01_用户主档案表.csv"
LOAD_DIR = BASE_DIR / "input" / "load"

OUTPUT_MODEL = BASE_DIR / "output" / "model"
OUTPUT_PREDICTION = BASE_DIR / "output" / "prediction"
OUTPUT_VALIDATION = BASE_DIR / "output" / "validation"
OUTPUT_LOGS = BASE_DIR / "output" / "logs"

for p in [OUTPUT_MODEL, OUTPUT_PREDICTION, OUTPUT_VALIDATION, OUTPUT_LOGS]:
    p.mkdir(parents=True, exist_ok=True)


# =========================================================
# 2. 日志
# =========================================================
LOG_FILE = OUTPUT_LOGS / "03_validate_log.txt"


def log(msg):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(str(msg) + "\n")


with open(LOG_FILE, "w", encoding="utf-8") as f:
    f.write("=== 验证日志 ===\n")


# =========================================================
# 3. 读取运行配置
# =========================================================
CONFIG_PATH = OUTPUT_MODEL / "run_config_v2_5_weatherfix.csv"
if not CONFIG_PATH.exists():
    raise FileNotFoundError("未找到 run_config_v2_5_weatherfix.csv，请先运行 01_train_v2_5_weatherfix.py")

RUN_CONFIG = pd.read_csv(CONFIG_PATH, encoding="utf-8-sig").iloc[0]
PREDICT_START_TS = pd.Timestamp(RUN_CONFIG["PREDICT_START"])
PREDICT_END_TS = pd.Timestamp(RUN_CONFIG["PREDICT_END"])
LOW_LOAD_THRESHOLD = float(RUN_CONFIG["LOW_LOAD_THRESHOLD"])
LOW_LOAD_PROBA_THRESHOLD = float(RUN_CONFIG["LOW_LOAD_PROBA_THRESHOLD"])


# =========================================================
# 4. 通用函数
# =========================================================
def normalize_text(x):
    if pd.isna(x):
        return None
    x = str(x).strip()
    x = x.replace("　", "").replace(" ", "")
    return x


def safe_read_table(path):
    path = str(path)
    if path.lower().endswith(".csv"):
        try:
            return pd.read_csv(path, encoding="utf-8-sig")
        except Exception:
            try:
                return pd.read_csv(path, encoding="gbk")
            except Exception:
                return pd.read_csv(path)
    else:
        return pd.read_excel(path)


def normalize_load_sheet(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    header_row_idx = None
    for i in range(min(len(df), 10)):
        row_values = [str(x).strip() for x in df.iloc[i].tolist()]
        joined = "".join(row_values)
        if ("电量年月日" in joined) and ("户号" in joined):
            header_row_idx = i
            break

    if header_row_idx is not None:
        new_header = [str(x).strip() for x in df.iloc[header_row_idx].tolist()]
        df = df.iloc[header_row_idx + 1:].copy()
        df.columns = new_header
    else:
        if "电量年月日" not in df.columns or "户号" not in df.columns:
            raise ValueError("未识别到负荷表头")

    df.columns = [str(c).strip() for c in df.columns]
    return df


def calc_pct_error(pred, actual):
    if pd.isna(pred) or pd.isna(actual) or actual == 0:
        return np.nan
    return (pred - actual) / actual


def calc_metrics(df, actual_col="actual_load", pred_col="predicted_load"):
    tmp = df.dropna(subset=[actual_col, pred_col]).copy()
    if tmp.empty:
        return {"mae": np.nan, "rmse": np.nan, "mape": np.nan, "sample_count": 0}

    mae = mean_absolute_error(tmp[actual_col], tmp[pred_col])
    rmse = np.sqrt(mean_squared_error(tmp[actual_col], tmp[pred_col]))

    ape = np.where(
        (tmp[actual_col].notna()) & (tmp[actual_col] != 0),
        np.abs(tmp[pred_col] - tmp[actual_col]) / tmp[actual_col],
        np.nan
    )
    mape = np.nanmean(ape)

    return {
        "mae": mae,
        "rmse": rmse,
        "mape": mape,
        "sample_count": len(tmp)
    }


# =========================================================
# 5. 读取主档案
# =========================================================
def load_user_master():
    df = safe_read_table(USER_MASTER_PATH)
    df.columns = [str(c).strip() for c in df.columns]
    df["用户名称_norm"] = df["用户名称"].apply(normalize_text)
    df["是否有光伏_flag"] = df["是否有光伏"].astype(str).str.strip().isin(
        ["是", "有", "1", "true", "True", "Y", "y"]
    ).astype(int)
    return df


# =========================================================
# 6. 读取实际值（按配置预测区间）
# =========================================================
def parse_actual_sheet(file_path, user_info):
    target_sheet = f"{str(PREDICT_START_TS.year)[2:]}.{PREDICT_START_TS.month}"

    raw = pd.read_excel(file_path, sheet_name=target_sheet, header=0)
    df = normalize_load_sheet(raw)

    required = ["电量年月日", "户号"]
    for c in required:
        if c not in df.columns:
            raise ValueError(f"{Path(file_path).name}-{target_sheet} 缺少字段: {c}")

    hour_cols = [f"{h}:00" for h in range(1, 25)]
    for c in hour_cols:
        if c not in df.columns:
            raise ValueError(f"{Path(file_path).name}-{target_sheet} 缺少小时列: {c}")

    df["电量年月日"] = df["电量年月日"].astype(str).str.strip()
    df = df[df["电量年月日"].str.fullmatch(r"\d{8}", na=False)].copy()
    if df.empty:
        return None

    long_df = df.melt(
        id_vars=["电量年月日", "户号"],
        value_vars=hour_cols,
        var_name="hour_str",
        value_name="actual_load"
    )

    long_df["date"] = pd.to_datetime(long_df["电量年月日"], format="%Y%m%d", errors="coerce")
    long_df["actual_load"] = pd.to_numeric(long_df["actual_load"], errors="coerce")
    long_df["户号"] = long_df["户号"].astype(str).str.strip()

    def build_datetime(row):
        d = row["date"]
        h = int(str(row["hour_str"]).split(":")[0])
        if pd.isna(d):
            return pd.NaT
        if h == 24:
            return d + pd.Timedelta(days=1)
        return d + pd.Timedelta(hours=h)

    long_df["datetime"] = long_df.apply(build_datetime, axis=1)
    long_df["用户编号"] = user_info["用户编号"]
    long_df["用户名称"] = user_info["用户名称"]
    long_df["是否有光伏_flag"] = user_info["是否有光伏_flag"]

    long_df = long_df[
        (long_df["datetime"] >= PREDICT_START_TS) &
        (long_df["datetime"] < PREDICT_END_TS)
    ].copy()

    return long_df[["用户编号", "用户名称", "户号", "datetime", "actual_load", "是否有光伏_flag"]].dropna(subset=["datetime"])


def load_actual(user_master_df):
    log("读取实际预测区间数据...")
    files = glob.glob(str(LOAD_DIR / "*.xlsx")) + glob.glob(str(LOAD_DIR / "*.xls"))

    all_list = []
    target_sheet = f"{str(PREDICT_START_TS.year)[2:]}.{PREDICT_START_TS.month}"

    for fp in files:
        file_name = Path(fp).stem
        user_name_norm = normalize_text(file_name)

        matched = user_master_df[user_master_df["用户名称_norm"] == user_name_norm]
        if matched.empty:
            continue

        user_info = matched.iloc[0]

        try:
            xls = pd.ExcelFile(fp)
        except Exception as e:
            log(f"[错误] 打开失败：{Path(fp).name} -> {e}")
            continue

        if target_sheet not in xls.sheet_names:
            log(f"[警告] 缺少 {target_sheet} sheet：{Path(fp).name}")
            continue

        try:
            one = parse_actual_sheet(fp, user_info)
            if one is not None and not one.empty:
                all_list.append(one)
                log(f"已读取实际：{Path(fp).name} - {target_sheet}")
        except Exception as e:
            log(f"[错误] 读取实际失败：{Path(fp).name} -> {e}")

    if not all_list:
        raise ValueError("未读取到任何预测区间实际数据")

    df = pd.concat(all_list, ignore_index=True)
    df = df.sort_values(["用户编号", "datetime"]).reset_index(drop=True)
    return df


# =========================================================
# 7. 读取预测结果
# =========================================================
def load_prediction_long():
    pred_path = OUTPUT_PREDICTION / "predict_long_v2_5_weatherfix.csv"
    if not pred_path.exists():
        raise FileNotFoundError("未找到 predict_long_v2_5_weatherfix.csv，请先运行 02_predict_v2_5_weatherfix.py")

    pred_df = pd.read_csv(pred_path, encoding="utf-8-sig")
    pred_df["datetime"] = pd.to_datetime(pred_df["datetime"], errors="coerce")

    keep_cols = [
        "用户编号", "用户名称", "户号", "datetime",
        "predicted_load", "predict_status",
        "是否有光伏_flag", "low_load_pred_flag", "low_load_pred_proba",
        "weather_match_level"
    ]
    keep_cols = [c for c in keep_cols if c in pred_df.columns]
    pred_df = pred_df[keep_cols].copy()

    if "predict_status" not in pred_df.columns:
        pred_df["predict_status"] = "已预测"

    return pred_df


# =========================================================
# 8. 合并验证
# =========================================================
def build_validation_long(actual_df, pred_df):
    log("合并实际与预测...")

    merge_cols = ["用户编号", "用户名称", "户号", "datetime"]
    df = actual_df.merge(pred_df, on=merge_cols, how="left", suffixes=("", "_pred"))

    if "是否有光伏_flag_pred" in df.columns:
        df["是否有光伏_flag"] = df["是否有光伏_flag"].fillna(df["是否有光伏_flag_pred"])
        df = df.drop(columns=["是否有光伏_flag_pred"])

    df["error"] = df["predicted_load"] - df["actual_load"]
    df["abs_error"] = df["error"].abs()
    df["pct_error"] = df.apply(lambda x: calc_pct_error(x["predicted_load"], x["actual_load"]), axis=1)

    df["ape"] = np.where(
        (df["actual_load"].notna()) & (df["actual_load"] != 0) & (df["predicted_load"].notna()),
        df["abs_error"] / df["actual_load"],
        np.nan
    )

    df["status"] = np.where(
        df["predicted_load"].isna(),
        df["predict_status"].fillna("未预测"),
        "已预测"
    )

    df["hour"] = pd.to_datetime(df["datetime"]).dt.hour
    df["is_daytime_8_19"] = df["hour"].isin(list(range(8, 20))).astype(int)
    df["is_low_load_actual"] = (df["actual_load"] < LOW_LOAD_THRESHOLD).astype(int)

    return df.sort_values(["用户编号", "datetime"]).reset_index(drop=True)


# =========================================================
# 9. 全时段指标
# =========================================================
def export_metrics(validation_df):
    total_metrics = calc_metrics(validation_df)

    pd.DataFrame([{
        "scope": "total",
        "mae": total_metrics["mae"],
        "rmse": total_metrics["rmse"],
        "mape": total_metrics["mape"],
        "sample_count": total_metrics["sample_count"]
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_metrics_total.csv",
        index=False, encoding="utf-8-sig"
    )

    by_user_rows = []
    for (uid, uname), g in validation_df.groupby(["用户编号", "用户名称"]):
        m = calc_metrics(g)
        by_user_rows.append({
            "用户编号": uid,
            "用户名称": uname,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    pd.DataFrame(by_user_rows).to_csv(
        OUTPUT_VALIDATION / "validation_metrics_by_user.csv",
        index=False, encoding="utf-8-sig"
    )

    log(f"总体验证 MAE: {total_metrics['mae']:.4f}")
    log(f"总体验证 RMSE: {total_metrics['rmse']:.4f}")
    log(f"总体验证 MAPE: {total_metrics['mape']:.4%}")


# =========================================================
# 10. 白天专项指标
# =========================================================
def export_daytime_metrics(validation_df):
    log("导出白天专项验证指标（8:00~19:00）...")

    day_df = validation_df[validation_df["is_daytime_8_19"] == 1].copy()

    total_metrics = calc_metrics(day_df)
    pd.DataFrame([{
        "scope": "daytime_8_19_total",
        "mae": total_metrics["mae"],
        "rmse": total_metrics["rmse"],
        "mape": total_metrics["mape"],
        "sample_count": total_metrics["sample_count"]
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_daytime_metrics_total.csv",
        index=False, encoding="utf-8-sig"
    )

    by_user_rows = []
    for (uid, uname), g in day_df.groupby(["用户编号", "用户名称"]):
        m = calc_metrics(g)
        by_user_rows.append({
            "用户编号": uid,
            "用户名称": uname,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })
    pd.DataFrame(by_user_rows).to_csv(
        OUTPUT_VALIDATION / "validation_daytime_metrics_by_user.csv",
        index=False, encoding="utf-8-sig"
    )

    if "是否有光伏_flag" in day_df.columns:
        by_pv_rows = []
        for pv_flag, g in day_df.groupby("是否有光伏_flag"):
            m = calc_metrics(g)
            by_pv_rows.append({
                "是否有光伏_flag": pv_flag,
                "mae": m["mae"],
                "rmse": m["rmse"],
                "mape": m["mape"],
                "sample_count": m["sample_count"]
            })
        pd.DataFrame(by_pv_rows).to_csv(
            OUTPUT_VALIDATION / "validation_daytime_metrics_by_pv.csv",
            index=False, encoding="utf-8-sig"
        )

    log(f"白天8:00~19:00 总体 MAE: {total_metrics['mae']:.4f}")
    log(f"白天8:00~19:00 总体 RMSE: {total_metrics['rmse']:.4f}")
    log(f"白天8:00~19:00 总体 MAPE: {total_metrics['mape']:.4%}")


# =========================================================
# 11. 按真实值分层
# =========================================================
def export_actual_value_bucket_metrics(validation_df):
    log("导出按真实值分层误差分析...")

    df = validation_df.copy()

    def get_bucket(x):
        if pd.isna(x):
            return "missing"
        elif x < 20:
            return "<20"
        elif x < 50:
            return "20-50"
        elif x < 100:
            return "50-100"
        elif x < 200:
            return "100-200"
        else:
            return ">=200"

    df["actual_bucket"] = df["actual_load"].apply(get_bucket)
    df_eval = df[df["actual_bucket"] != "missing"].copy()
    bucket_order = ["<20", "20-50", "50-100", "100-200", ">=200"]

    rows_all = []
    for bucket, g in df_eval.groupby("actual_bucket"):
        m = calc_metrics(g)
        rows_all.append({
            "actual_bucket": bucket,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    out_all = pd.DataFrame(rows_all)
    if not out_all.empty:
        out_all["actual_bucket"] = pd.Categorical(out_all["actual_bucket"], categories=bucket_order, ordered=True)
        out_all = out_all.sort_values("actual_bucket")

    out_all.to_csv(
        OUTPUT_VALIDATION / "validation_metrics_by_actual_bucket.csv",
        index=False, encoding="utf-8-sig"
    )

    day_df = df_eval[df_eval["is_daytime_8_19"] == 1].copy()

    rows_day = []
    for bucket, g in day_df.groupby("actual_bucket"):
        m = calc_metrics(g)
        rows_day.append({
            "actual_bucket": bucket,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    out_day = pd.DataFrame(rows_day)
    if not out_day.empty:
        out_day["actual_bucket"] = pd.Categorical(out_day["actual_bucket"], categories=bucket_order, ordered=True)
        out_day = out_day.sort_values("actual_bucket")

    out_day.to_csv(
        OUTPUT_VALIDATION / "validation_daytime_metrics_by_actual_bucket.csv",
        index=False, encoding="utf-8-sig"
    )

    log("按真实值分层误差分析导出完成")


# =========================================================
# 12. 低负荷识别分析
# =========================================================
def export_low_load_usage_metrics(validation_df):
    log("导出低负荷状态识别使用情况分析...")

    df = validation_df.copy()

    if "low_load_pred_flag" not in df.columns:
        log("[警告] 预测结果中不存在 low_load_pred_flag，跳过低负荷分类分析")
        return

    eval_df = df.dropna(subset=["low_load_pred_flag"]).copy()
    if eval_df.empty:
        log("[警告] low_load_pred_flag 全为空，跳过低负荷分类分析")
        return

    eval_df["low_load_pred_flag"] = pd.to_numeric(eval_df["low_load_pred_flag"], errors="coerce")
    eval_df = eval_df.dropna(subset=["low_load_pred_flag"]).copy()
    eval_df["low_load_pred_flag"] = eval_df["low_load_pred_flag"].astype(int)

    total_count = len(eval_df)
    pred_low_count = (eval_df["low_load_pred_flag"] == 1).sum()
    actual_low_count = (eval_df["is_low_load_actual"] == 1).sum()

    tp = ((eval_df["low_load_pred_flag"] == 1) & (eval_df["is_low_load_actual"] == 1)).sum()
    fp = ((eval_df["low_load_pred_flag"] == 1) & (eval_df["is_low_load_actual"] == 0)).sum()
    tn = ((eval_df["low_load_pred_flag"] == 0) & (eval_df["is_low_load_actual"] == 0)).sum()
    fn = ((eval_df["low_load_pred_flag"] == 0) & (eval_df["is_low_load_actual"] == 1)).sum()

    pd.DataFrame([{
        "total_count": total_count,
        "pred_low_count": pred_low_count,
        "actual_low_count": actual_low_count,
        "tp": tp,
        "fp": fp,
        "tn": tn,
        "fn": fn,
        "pred_low_ratio": pred_low_count / total_count if total_count > 0 else np.nan,
        "actual_low_ratio": actual_low_count / total_count if total_count > 0 else np.nan,
        "low_load_threshold": LOW_LOAD_THRESHOLD,
        "low_load_proba_threshold": LOW_LOAD_PROBA_THRESHOLD
    }]).to_csv(
        OUTPUT_VALIDATION / "validation_low_load_usage_summary.csv",
        index=False,
        encoding="utf-8-sig"
    )

    day_df = eval_df[eval_df["is_daytime_8_19"] == 1].copy()
    if not day_df.empty:
        total_count = len(day_df)
        pred_low_count = (day_df["low_load_pred_flag"] == 1).sum()
        actual_low_count = (day_df["is_low_load_actual"] == 1).sum()

        tp = ((day_df["low_load_pred_flag"] == 1) & (day_df["is_low_load_actual"] == 1)).sum()
        fp = ((day_df["low_load_pred_flag"] == 1) & (day_df["is_low_load_actual"] == 0)).sum()
        tn = ((day_df["low_load_pred_flag"] == 0) & (day_df["is_low_load_actual"] == 0)).sum()
        fn = ((day_df["low_load_pred_flag"] == 0) & (day_df["is_low_load_actual"] == 1)).sum()

        pd.DataFrame([{
            "total_count": total_count,
            "pred_low_count": pred_low_count,
            "actual_low_count": actual_low_count,
            "tp": tp,
            "fp": fp,
            "tn": tn,
            "fn": fn,
            "pred_low_ratio": pred_low_count / total_count if total_count > 0 else np.nan,
            "actual_low_ratio": actual_low_count / total_count if total_count > 0 else np.nan,
            "low_load_threshold": LOW_LOAD_THRESHOLD,
            "low_load_proba_threshold": LOW_LOAD_PROBA_THRESHOLD
        }]).to_csv(
            OUTPUT_VALIDATION / "validation_daytime_low_load_usage_summary.csv",
            index=False,
            encoding="utf-8-sig"
        )

    log("低负荷状态识别使用情况分析导出完成")


# =========================================================
# 13. 天气匹配层级分析（新增）
# =========================================================
def export_weather_match_metrics(validation_df):
    log("导出天气匹配层级分析...")

    if "weather_match_level" not in validation_df.columns:
        log("[警告] 预测结果中不存在 weather_match_level，跳过天气匹配层级分析")
        return

    df = validation_df.copy()

    count_df = (
        df.groupby("weather_match_level", dropna=False)
        .size()
        .reset_index(name="sample_count")
    )
    count_df.to_csv(
        OUTPUT_VALIDATION / "validation_weather_match_level_count.csv",
        index=False,
        encoding="utf-8-sig"
    )

    metric_rows = []
    for level, g in df.groupby("weather_match_level", dropna=False):
        m = calc_metrics(g)
        metric_rows.append({
            "weather_match_level": level,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    pd.DataFrame(metric_rows).to_csv(
        OUTPUT_VALIDATION / "validation_weather_match_level_metrics.csv",
        index=False,
        encoding="utf-8-sig"
    )

    day_df = df[df["is_daytime_8_19"] == 1].copy()
    day_rows = []
    for level, g in day_df.groupby("weather_match_level", dropna=False):
        m = calc_metrics(g)
        day_rows.append({
            "weather_match_level": level,
            "mae": m["mae"],
            "rmse": m["rmse"],
            "mape": m["mape"],
            "sample_count": m["sample_count"]
        })

    pd.DataFrame(day_rows).to_csv(
        OUTPUT_VALIDATION / "validation_daytime_weather_match_level_metrics.csv",
        index=False,
        encoding="utf-8-sig"
    )

    log("天气匹配层级分析导出完成")


# =========================================================
# 14. 每个用户宽表：每天3行
# =========================================================
def build_validation_wide_for_one_user(g):
    g = g.copy().sort_values("datetime")

    g["base_date"] = np.where(
        g["datetime"].dt.hour == 0,
        (g["datetime"] - pd.Timedelta(days=1)).dt.normalize(),
        g["datetime"].dt.normalize()
    )
    g["base_date"] = pd.to_datetime(g["base_date"])

    day_list = sorted(g["base_date"].dropna().unique())
    rows = []

    for d in day_list:
        day = pd.to_datetime(d)
        one_day = g[g["base_date"] == day].copy()

        account_value = one_day["户号"].dropna().astype(str).iloc[0] if one_day["户号"].notna().any() else ""

        pred_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "预测"}
        actual_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "实际"}
        error_row = {"电量年月日": day.strftime("%Y%m%d"), "户号": account_value, "类型": "偏差"}

        pred_sum = 0.0
        actual_sum = 0.0
        pred_empty = True
        actual_empty = True

        for h in range(1, 25):
            target_dt = day + pd.Timedelta(days=1) if h == 24 else day + pd.Timedelta(hours=h)
            hit = one_day[one_day["datetime"] == target_dt]

            if hit.empty:
                pred_val = np.nan
                actual_val = np.nan
            else:
                pred_val = hit["predicted_load"].iloc[0] if "predicted_load" in hit.columns else np.nan
                actual_val = hit["actual_load"].iloc[0] if "actual_load" in hit.columns else np.nan

            pct_val = calc_pct_error(pred_val, actual_val)

            pred_row[f"{h}:00"] = pred_val
            actual_row[f"{h}:00"] = actual_val
            error_row[f"{h}:00"] = pct_val

            if pd.notna(pred_val):
                pred_sum += pred_val
                pred_empty = False
            if pd.notna(actual_val):
                actual_sum += actual_val
                actual_empty = False

        pred_row["合计"] = np.nan if pred_empty else pred_sum
        actual_row["合计"] = np.nan if actual_empty else actual_sum
        error_row["合计"] = calc_pct_error(pred_sum, actual_sum) if (not pred_empty and not actual_empty) else np.nan

        rows.extend([pred_row, actual_row, error_row])

    final_cols = ["电量年月日", "户号", "类型"] + [f"{h}:00" for h in range(1, 25)] + ["合计"]
    return pd.DataFrame(rows)[final_cols]


# =========================================================
# 15. 全体汇总宽表：每天3行
# =========================================================
def build_validation_total_wide(validation_df):
    df = validation_df.copy().sort_values("datetime")

    df["base_date"] = np.where(
        df["datetime"].dt.hour == 0,
        (df["datetime"] - pd.Timedelta(days=1)).dt.normalize(),
        df["datetime"].dt.normalize()
    )
    df["base_date"] = pd.to_datetime(df["base_date"])

    day_list = sorted(df["base_date"].dropna().unique())
    rows = []

    for d in day_list:
        day = pd.to_datetime(d)
        one_day = df[df["base_date"] == day].copy()

        pred_row = {"电量年月日": day.strftime("%Y%m%d"), "类型": "预测"}
        actual_row = {"电量年月日": day.strftime("%Y%m%d"), "类型": "实际"}
        error_row = {"电量年月日": day.strftime("%Y%m%d"), "类型": "偏差"}

        pred_sum_day = 0.0
        actual_sum_day = 0.0
        pred_empty = True
        actual_empty = True

        for h in range(1, 25):
            target_dt = day + pd.Timedelta(days=1) if h == 24 else day + pd.Timedelta(hours=h)
            hit = one_day[one_day["datetime"] == target_dt].copy()

            if hit.empty:
                pred_val = np.nan
                actual_val = np.nan
            else:
                pred_val = hit["predicted_load"].sum(min_count=1)
                actual_val = hit["actual_load"].sum(min_count=1)

            pct_val = calc_pct_error(pred_val, actual_val)

            pred_row[f"{h}:00"] = pred_val
            actual_row[f"{h}:00"] = actual_val
            error_row[f"{h}:00"] = pct_val

            if pd.notna(pred_val):
                pred_sum_day += pred_val
                pred_empty = False
            if pd.notna(actual_val):
                actual_sum_day += actual_val
                actual_empty = False

        pred_row["合计"] = np.nan if pred_empty else pred_sum_day
        actual_row["合计"] = np.nan if actual_empty else actual_sum_day
        error_row["合计"] = calc_pct_error(pred_sum_day, actual_sum_day) if (not pred_empty and not actual_empty) else np.nan

        rows.extend([pred_row, actual_row, error_row])

    final_cols = ["电量年月日", "类型"] + [f"{h}:00" for h in range(1, 25)] + ["合计"]
    return pd.DataFrame(rows)[final_cols]


# =========================================================
# 16. Excel格式化：用户表
# =========================================================
def format_validation_sheet(ws):
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align

    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 14
        elif col == 2:
            ws.column_dimensions[col_letter].width = 28
        elif col == 3:
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 12

    for row in range(2, max_row + 1):
        row_type = ws.cell(row=row, column=3).value

        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).alignment = center_align

        if row_type == "预测":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = blue_fill

        elif row_type == "实际":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = green_fill

        elif row_type == "偏差":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

            for col in range(4, max_col + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value

                if val is None or val == "":
                    continue

                try:
                    abs_val = abs(float(val))
                except Exception:
                    continue

                cell.number_format = "0.00%"

                if abs_val >= 0.30:
                    cell.fill = red_fill
                elif abs_val >= 0.15:
                    cell.fill = orange_fill
                elif abs_val >= 0.05:
                    cell.fill = warn_fill


# =========================================================
# 17. Excel格式化：总量表
# =========================================================
def format_total_validation_sheet(ws):
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align

    ws.freeze_panes = "A2"

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 14
        elif col == 2:
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 12

    for row in range(2, max_row + 1):
        row_type = ws.cell(row=row, column=2).value

        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).alignment = center_align

        if row_type == "预测":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = blue_fill

        elif row_type == "实际":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = green_fill

        elif row_type == "偏差":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

            for col in range(3, max_col + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value

                if val is None or val == "":
                    continue

                try:
                    abs_val = abs(float(val))
                except Exception:
                    continue

                cell.number_format = "0.00%"

                if abs_val >= 0.30:
                    cell.fill = red_fill
                elif abs_val >= 0.15:
                    cell.fill = orange_fill
                elif abs_val >= 0.05:
                    cell.fill = warn_fill


# =========================================================
# 18. 导出验证结果
# =========================================================
def export_validation_outputs(validation_df):
    log("导出验证结果...")
    validation_df.to_csv(OUTPUT_VALIDATION / "validation_hourly_long.csv", index=False, encoding="utf-8-sig")

    out_excel = OUTPUT_VALIDATION / "validation_hourly_wide.xlsx"
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        for (uid, uname), g in validation_df.groupby(["用户编号", "用户名称"]):
            wide_df = build_validation_wide_for_one_user(g)
            sheet_name = str(uname)[:31] if str(uname) else str(uid)
            wide_df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.book[sheet_name]
            format_validation_sheet(ws)

    log(f"已导出：{out_excel.name}")

    total_wide_df = build_validation_total_wide(validation_df)
    total_excel = OUTPUT_VALIDATION / "validation_total_hourly_wide.xlsx"

    with pd.ExcelWriter(total_excel, engine="openpyxl") as writer:
        total_wide_df.to_excel(writer, sheet_name="总量汇总", index=False)
        ws = writer.book["总量汇总"]
        format_total_validation_sheet(ws)

    log(f"已导出：{total_excel.name}")


# =========================================================
# 19. 主流程
# =========================================================
def main():
    log("=== 开始V2.5_WeatherFix验证 ===")
    log(f"PREDICT_START = {PREDICT_START_TS}")
    log(f"PREDICT_END   = {PREDICT_END_TS}")
    log(f"LOW_LOAD_THRESHOLD = {LOW_LOAD_THRESHOLD}")
    log(f"LOW_LOAD_PROBA_THRESHOLD = {LOW_LOAD_PROBA_THRESHOLD}")

    user_master_df = load_user_master()
    actual_df = load_actual(user_master_df)
    pred_df = load_prediction_long()

    validation_df = build_validation_long(actual_df, pred_df)

    export_metrics(validation_df)
    export_daytime_metrics(validation_df)
    export_actual_value_bucket_metrics(validation_df)
    export_low_load_usage_metrics(validation_df)
    export_weather_match_metrics(validation_df)
    export_validation_outputs(validation_df)

    log("=== V2.5_WeatherFix验证完成 ===")


if __name__ == "__main__":
    main()