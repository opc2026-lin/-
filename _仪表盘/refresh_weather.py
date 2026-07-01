"""从 fujian_pv_daily.xlsx 提取最新天气数据，写入 weather_today.json"""
import openpyxl, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
VAULT = os.path.dirname(SCRIPT_DIR)
PV_FILE = os.path.join(VAULT, "1-1负荷预测输入", "2.预测天气", "fujian_pv_daily.xlsx")
OUTPUT = os.path.join(SCRIPT_DIR, "weather_today.json")

COLORS = {"宁德": "#5b89ff", "莆田": "#f472b6", "福州": "#fbbf24", "泉州": "#34d399"}
CAPS = {"宁德": "2000kW", "莆田": "900kW", "福州": "400kW", "泉州": "400kW"}

def cloud_to_weather(pct):
    if pct < 30: return "晴", "☀"
    elif pct < 60: return "多云", "⛅"
    elif pct < 90: return "阴", "☁"
    else: return "雨", "🌧"

from datetime import date, timedelta

wb = openpyxl.load_workbook(PV_FILE, data_only=True)

# 匹配今天日期的 sheet，找不到则用 D+2 日
today_str = f"{date.today().month}月{date.today().day}日"
d2_str = f"{(date.today()+timedelta(days=2)).month}月{(date.today()+timedelta(days=2)).day}日"
sheet_name = None
for s in wb.sheetnames:
    clean = s.strip("'\"")
    if clean == today_str:
        sheet_name = s; break
if not sheet_name:
    for s in wb.sheetnames:
        if s.strip("'\"") == d2_str:
            sheet_name = s; break
if not sheet_name:
    sheet_name = wb.sheetnames[0]  # fallback

ws = wb[sheet_name]
date_str = sheet_name.strip("'\"")

regions = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
    if row[0] and "公司" in str(row[0]):
        parts = str(row[0]).split("|")
        city = parts[1].strip()
        cloud_pct = int(parts[3].strip().replace("云", "").replace("%", ""))
        weather, icon = cloud_to_weather(cloud_pct)

        temp_row = list(ws.iter_rows(min_row=i+3, max_row=i+3, values_only=True))[0]
        temps = [float(v) if v and v != "-" else None for v in temp_row[1:25]]
        peak_temp = max((t for t in temps if t is not None), default=0)

        rad_row = list(ws.iter_rows(min_row=i+4, max_row=i+4, values_only=True))[0]
        rads = [float(v) if v and v != "-" else None for v in rad_row[1:25]]
        peak_rad = max((r for r in rads if r is not None), default=0)

        pv_row = list(ws.iter_rows(min_row=i+5, max_row=i+5, values_only=True))[0]
        pvs = [float(v) if v and v != "-" else None for v in pv_row[1:25]]
        total_pv = sum(p for p in pvs if p is not None)

        regions.append({
            "city": city, "temp": round(peak_temp, 1),
            "weather": weather, "icon": icon,
            "w": f"{icon} {weather}", "cloud": cloud_pct,
            "rad": round(peak_rad, 0), "pv": round(total_pv, 3),
            "dot": COLORS.get(city, "#5b89ff"),
            "cap": CAPS.get(city, "")
        })

data = {"date": date_str, "regions": regions}
with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print(f"✅ {date_str} 天气数据已写入 weather_today.json")
for r in regions:
    print(f"  {r['city']}: {r['temp']}° {r['icon']} {r['weather']}")
