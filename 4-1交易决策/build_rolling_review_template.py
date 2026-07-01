from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT_DIR = Path(r"D:\亿云能源科技售电交易数据库\4-1交易决策")
OUTPUT_PATH = OUTPUT_DIR / "滚动撮合复盘标准模板.xlsx"


THIN = Side(style="thin", color="BFBFBF")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill("solid", fgColor="E2F0D9")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")


def style_cell(cell, *, bold=False, fill=None, center=False, wrap=False):
    cell.font = Font(bold=bold)
    if fill:
        cell.fill = fill
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap,
    )


def build_cover(ws):
    ws.title = "填写说明"
    ws["A1"] = "滚动撮合复盘标准模板"
    ws["A2"] = "用途：复盘中长期底仓之上的连续滚动撮合申报，不复盘最终用户总负荷曲线。"
    ws["A4"] = "使用顺序"
    ws["A5"] = "1. 先填基础信息"
    ws["A6"] = "2. 再填24时段逐小时主表"
    ws["A7"] = "3. 每小时只选一个结果分类"
    ws["A8"] = "4. 最后写当日汇总和规则沉淀"
    ws["A10"] = "关键公式"
    ws["A11"] = "净敞口 = 预测用电 - 中长期持仓"
    ws["A12"] = "报量比例 = 申报量 / 净敞口"
    ws["A13"] = "价差 = 申报价 - 市场成交参考价"
    ws["A14"] = "修复比例 = 成交量 / 净敞口"
    ws["A16"] = "结果分类枚举"
    options = [
        "方向对，量少了",
        "方向对，价低了",
        "方向对，执行可以",
        "方向错，不该买",
        "方向错，应该更积极",
        "试探未成，可接受",
        "成交了，但买多了",
        "成交了，但不是关键时段",
    ]
    for idx, text in enumerate(options, start=17):
        ws[f"A{idx}"] = text

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws["A1"].font = Font(bold=True, size=14)
    for row in [4, 10, 16]:
        style_cell(ws[f"A{row}"], bold=True, fill=SECTION_FILL)
    for row in range(1, 25):
        style_cell(ws[f"A{row}"], wrap=True)


def build_base_info(ws):
    ws.title = "基础信息"
    rows = [
        ("目标日", ""),
        ("复盘日期", ""),
        ("天气判断", ""),
        ("是否高温日", ""),
        ("是否工作日", ""),
        ("中长期底仓口径", ""),
        ("滚动撮合场次", ""),
        ("复盘人", ""),
        ("备注", ""),
    ]
    ws["A1"] = "基础信息"
    style_cell(ws["A1"], bold=True, fill=SECTION_FILL, center=True)
    for i, (label, value) in enumerate(rows, start=3):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        style_cell(ws[f"A{i}"], bold=True, fill=HEADER_FILL)
        style_cell(ws[f"B{i}"])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42


def build_hourly_review(ws):
    ws.title = "24时段主表"
    headers = [
        "时段",
        "预测用电",
        "中长期持仓",
        "净敞口",
        "时段分层",
        "动作类型",
        "申报量",
        "报量比例",
        "申报价",
        "市场成交参考价",
        "价差",
        "成交量",
        "修复比例",
        "没成交性质",
        "结果分类",
        "备注",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        style_cell(cell, bold=True, fill=HEADER_FILL, center=True, wrap=True)

    for hour in range(1, 25):
        row = hour + 1
        ws.cell(row=row, column=1, value=hour)
        style_cell(ws.cell(row=row, column=1), center=True)
        for col in range(2, 17):
            style_cell(ws.cell(row=row, column=col))

        ws.cell(row=row, column=4, value=f"=IF(OR(B{row}=\"\",C{row}=\"\"),\"\",B{row}-C{row})")
        ws.cell(row=row, column=8, value=f"=IF(OR(D{row}=\"\",D{row}=0,G{row}=\"\"),\"\",G{row}/D{row})")
        ws.cell(row=row, column=11, value=f"=IF(OR(I{row}=\"\",J{row}=\"\"),\"\",I{row}-J{row})")
        ws.cell(row=row, column=13, value=f"=IF(OR(D{row}=\"\",D{row}=0,L{row}=\"\"),\"\",L{row}/D{row})")

    widths = {
        "A": 8,
        "B": 12,
        "C": 14,
        "D": 12,
        "E": 12,
        "F": 16,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 16,
        "K": 10,
        "L": 12,
        "M": 12,
        "N": 14,
        "O": 20,
        "P": 24,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"


def build_rules(ws):
    ws.title = "判定规则"
    rows = [
        ("时段分层", "低价值时段：本来不缺或缺口小；一般时段：有缺口但非核心；关键时段：缺口大且属于保供/高价时段"),
        ("报量比例", "0~30% 试探型；30%~70% 中性型；70%~100% 进攻型；>100% 激进型，必须解释原因"),
        ("价差判定", "价差>=0 说明价格达到成交要求；-1~0 只差临门一脚；-3~-1 价格偏低；<-3 明显不是成交价"),
        ("修复比例", "<30% 修复弱；30%~70% 修复一般；>70% 有效修复；>100% 可能过补"),
        ("没成交性质", "只填：可接受 / 不可接受"),
        ("动作类型", "试探补仓 / 正常补仓 / 抢量补仓 / 防高价少报 / 判断不买"),
    ]
    ws["A1"] = "固定判定规则"
    style_cell(ws["A1"], bold=True, fill=SECTION_FILL, center=True)
    for i, (label, desc) in enumerate(rows, start=3):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = desc
        style_cell(ws[f"A{i}"], bold=True, fill=HEADER_FILL)
        style_cell(ws[f"B{i}"], wrap=True)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 100


def build_daily_summary(ws):
    ws.title = "当日汇总"
    prompts = [
        "今天最关键的错因是什么？",
        "今天最值得保留的动作是什么？",
        "今天最不该重复的动作是什么？",
        "如果同样的盘面再来一次，哪些时段会改？",
        "明天开始固定执行的规则是什么？",
    ]
    ws["A1"] = "当日汇总"
    style_cell(ws["A1"], bold=True, fill=SECTION_FILL, center=True)
    row = 3
    for prompt in prompts:
        ws[f"A{row}"] = prompt
        style_cell(ws[f"A{row}"], bold=True, fill=HEADER_FILL, wrap=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=6)
        for r in range(row, row + 3):
            for c in range(2, 7):
                style_cell(ws.cell(row=r, column=c))
        row += 4
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 22 if col == "A" else 18


def build_rule_settlement(ws):
    ws.title = "规则沉淀"
    ws["A1"] = "规则沉淀"
    style_cell(ws["A1"], bold=True, fill=SECTION_FILL, center=True)
    headers = ["规则编号", "适用条件", "执行动作", "是否纳入长期规则"]
    for col, header in enumerate(headers, start=1):
        style_cell(ws.cell(row=3, column=col, value=header), bold=True, fill=HEADER_FILL, center=True)

    for i in range(4, 14):
        ws.cell(row=i, column=1, value=f"规则{i-3}")
        for col in range(1, 5):
            style_cell(ws.cell(row=i, column=col), wrap=(col in [2, 3]))

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 18


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_cover(wb.active)
    build_base_info(wb.create_sheet())
    build_hourly_review(wb.create_sheet())
    build_rules(wb.create_sheet())
    build_daily_summary(wb.create_sheet())
    build_rule_settlement(wb.create_sheet())
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
