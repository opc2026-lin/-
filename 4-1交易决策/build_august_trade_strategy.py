from pathlib import Path
import math
import re

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(r"D:\亿云能源科技售电交易数据库")
OUTPUT_DIR = BASE_DIR / "4-1交易决策"

ANNUAL_BILATERAL = BASE_DIR / "_交易运营" / "交易记录" / "年度双边" / "2026年年度双边协商交易交易结果.xlsx"
ANNUAL_LISTING_FILES = [
    BASE_DIR / "_交易运营" / "交易记录" / "年度挂牌交易" / "2026年年度挂牌交易-轮次2交易结果.xlsx",
    BASE_DIR / "_交易运营" / "交易记录" / "年度挂牌交易" / "2026年年度挂牌交易-轮次3交易结果.xlsx",
    BASE_DIR / "_交易运营" / "交易记录" / "年度挂牌交易" / "2026年年度挂牌交易-轮次4交易结果.xlsx",
]
MONTHLY_MARKET_PRICE = BASE_DIR / "_交易运营" / "交易记录" / "挂牌与竞价申报数据" / "1-7月月竞市场成交价格汇总.xlsx"
JULY_HOLDING = BASE_DIR / "_交易运营" / "日滚搓申报" / "日持仓量" / "2026年7月日持仓量明细.xlsx"
JUNE_DAILY_DIR = BASE_DIR / "_交易运营" / "交易记录" / "挂牌与竞价申报数据" / "6月交易"

AUGUST_DAYS = 31
RECOMMENDED_MONTHLY_LISTING_MWH_PER_H = 2.000


def extract_hour(text: str) -> int | None:
    nums = re.findall(r"\d+", str(text))
    if len(nums) >= 2:
        return int(nums[1]) + 1
    if len(nums) == 1:
        return int(nums[0])
    return None


def extract_month(period_text: str) -> int | None:
    match = re.search(r"(\d{4})-(\d{2})-\d{2}", str(period_text))
    if match:
        return int(match.group(2))
    return None


def load_annual_flat_base() -> pd.DataFrame:
    rows = []

    bilateral = pd.read_excel(ANNUAL_BILATERAL, sheet_name=0)
    rows.extend(_parse_trade_result(bilateral, "annual_bilateral"))

    for file_path in ANNUAL_LISTING_FILES:
        df = pd.read_excel(file_path, sheet_name=0)
        rows.extend(_parse_trade_result(df, "annual_listing"))

    flat = pd.DataFrame(rows)
    flat = flat[flat["month"] == 8].copy()
    summary = flat.groupby(["source", "hour"], as_index=False)["mwh_month"].sum()
    summary["mwh_per_h"] = summary["mwh_month"] / AUGUST_DAYS
    pivot = summary.pivot(index="hour", columns="source", values="mwh_per_h").fillna(0.0).reset_index()
    if "annual_bilateral" not in pivot.columns:
        pivot["annual_bilateral"] = 0.0
    if "annual_listing" not in pivot.columns:
        pivot["annual_listing"] = 0.0
    pivot["annual_total_flat"] = pivot["annual_bilateral"] + pivot["annual_listing"]
    return pivot.sort_values("hour").reset_index(drop=True)


def _parse_trade_result(df: pd.DataFrame, source_name: str) -> list[dict]:
    period_col = df.columns[3]
    hour_col = df.columns[4]
    qty_col = df.columns[5]
    rows = []
    for _, row in df.iterrows():
        rows.append(
            {
                "source": source_name,
                "month": extract_month(row[period_col]),
                "hour": extract_hour(row[hour_col]),
                "mwh_month": float(pd.to_numeric(row[qty_col], errors="coerce") or 0.0),
            }
        )
    return rows


def load_proxy_forecast_shape() -> pd.DataFrame:
    raw = pd.read_excel(JULY_HOLDING, sheet_name=0, header=None)
    labels = raw.iloc[3:9, 0].astype(str).tolist()
    values = raw.iloc[3:9, 1:25].copy()
    values.index = labels
    values.columns = list(range(1, 25))

    target_label = labels[5]
    forecast = pd.to_numeric(values.loc[target_label], errors="coerce")
    return pd.DataFrame({"hour": range(1, 25), "proxy_load_mwh_per_h": forecast.values})


def load_monthly_market_prices() -> pd.DataFrame:
    xls = pd.ExcelFile(MONTHLY_MARKET_PRICE)
    rows = []
    for sheet in xls.sheet_names:
        month_match = re.search(r"(\d+)", sheet)
        month = int(month_match.group(1)) if month_match else None
        df = pd.read_excel(MONTHLY_MARKET_PRICE, sheet_name=sheet)
        if df.empty:
            continue
        values = df.iloc[0, 1:25]
        for hour, value in enumerate(values, start=1):
            rows.append({"month": month, "hour": hour, "market_price": float(pd.to_numeric(value, errors="coerce"))})
    out = pd.DataFrame(rows)
    july = out[out["month"] == 7].copy().sort_values("hour").reset_index(drop=True)
    if july.empty:
        raise FileNotFoundError("未找到 7 月月竞市场价。")
    return july


def load_june_daily_price_stats() -> pd.DataFrame:
    rows = []
    for file_path in sorted(JUNE_DAILY_DIR.glob("6-*.xlsx")):
        df = pd.read_excel(file_path, sheet_name=0)
        for _, row in df.iterrows():
            hour = extract_hour(row.iloc[5])
            if hour is None:
                continue
            deal_qty = pd.to_numeric(row.iloc[10], errors="coerce")
            deal_price = pd.to_numeric(row.iloc[11], errors="coerce")
            if pd.isna(deal_qty) or pd.isna(deal_price):
                continue
            rows.append({"hour": hour, "deal_qty": float(deal_qty), "deal_price": float(deal_price)})

    daily = pd.DataFrame(rows)
    grouped = daily.groupby("hour")["deal_price"]
    summary = grouped.agg(["mean", "median", "min", "max"]).reset_index()
    summary["p75"] = grouped.quantile(0.75).values
    summary["p90"] = grouped.quantile(0.90).values
    return summary.sort_values("hour").reset_index(drop=True)


def round_price(value: float) -> float:
    return math.ceil(value * 10) / 10


def build_strategy() -> tuple[pd.DataFrame, dict]:
    annual = load_annual_flat_base()
    forecast = load_proxy_forecast_shape()
    july_market = load_monthly_market_prices()
    june_daily = load_june_daily_price_stats()

    df = annual.merge(forecast, on="hour", how="left")
    df = df.merge(july_market[["hour", "market_price"]].rename(columns={"market_price": "july_monthly_market_price"}), on="hour", how="left")
    df = df.merge(june_daily[["hour", "mean", "median", "p75", "p90"]], on="hour", how="left")

    df["recommended_monthly_listing_mwh_per_h"] = RECOMMENDED_MONTHLY_LISTING_MWH_PER_H
    df["total_flat_after_listing"] = df["annual_total_flat"] + df["recommended_monthly_listing_mwh_per_h"]
    df["recommended_monthly_comp_mwh_per_h"] = (df["proxy_load_mwh_per_h"] - df["total_flat_after_listing"]).clip(lower=0)

    def price_rule(row: pd.Series) -> float | None:
        qty = float(row["recommended_monthly_comp_mwh_per_h"])
        if qty <= 0:
            return None
        anchor = max(float(row["july_monthly_market_price"]), float(row["p75"]))
        if qty >= 2.0:
            margin = 2.0
        elif qty >= 1.0:
            margin = 1.0
        elif qty >= 0.3:
            margin = 0.5
        else:
            margin = 0.3
        return round_price(anchor + margin)

    df["recommended_monthly_comp_price"] = df.apply(price_rule, axis=1)
    df["monthly_comp_monthly_mwh"] = df["recommended_monthly_comp_mwh_per_h"] * AUGUST_DAYS

    assumptions = {
        "annual_flat_avg_mwh_per_h": round(float(df["annual_total_flat"].mean()), 3),
        "monthly_listing_target_mwh_per_h": RECOMMENDED_MONTHLY_LISTING_MWH_PER_H,
        "monthly_listing_total_mwh": round(RECOMMENDED_MONTHLY_LISTING_MWH_PER_H * 24 * AUGUST_DAYS, 3),
        "flat_after_listing_avg_mwh_per_h": round(float(df["total_flat_after_listing"].mean()), 3),
        "proxy_load_avg_mwh_per_h": round(float(df["proxy_load_mwh_per_h"].mean()), 3),
        "recommended_monthly_comp_avg_mwh_per_h": round(float(df["recommended_monthly_comp_mwh_per_h"].mean()), 3),
        "recommended_monthly_comp_total_mwh": round(float(df["monthly_comp_monthly_mwh"].sum()), 3),
    }
    return df, assumptions


def save_outputs(strategy_df: pd.DataFrame, assumptions: dict) -> tuple[Path, Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OUTPUT_DIR / "2026年8月挂牌与月竞申报策略.xlsx"
    md_path = OUTPUT_DIR / "2026年8月挂牌与月竞申报策略说明.md"

    listing_df = pd.DataFrame(
        [
            {
                "项目": "年度双边+年度挂牌底仓",
                "数值": assumptions["annual_flat_avg_mwh_per_h"],
                "单位": "MWh/h",
                "说明": "按 2026 年 8 月年度双边与年度挂牌结果折算",
            },
            {
                "项目": "建议月挂牌申报量",
                "数值": assumptions["monthly_listing_target_mwh_per_h"],
                "单位": "MWh/h",
                "说明": "建议按平仓直线申报",
            },
            {
                "项目": "建议月挂牌整月总量",
                "数值": assumptions["monthly_listing_total_mwh"],
                "单位": "MWh",
                "说明": "2.000 × 24 × 31",
            },
            {
                "项目": "挂牌后总平仓",
                "数值": assumptions["flat_after_listing_avg_mwh_per_h"],
                "单位": "MWh/h",
                "说明": "年度底仓 + 月挂牌",
            },
            {
                "项目": "代理负荷均值",
                "数值": assumptions["proxy_load_avg_mwh_per_h"],
                "单位": "MWh/h",
                "说明": "用 7 月 D+2 形状代替 8 月曲线",
            },
            {
                "项目": "建议月竞整月总量",
                "数值": assumptions["recommended_monthly_comp_total_mwh"],
                "单位": "MWh",
                "说明": "仅补 24 时段形状缺口",
            },
        ]
    )

    output_df = strategy_df[
        [
            "hour",
            "annual_bilateral",
            "annual_listing",
            "annual_total_flat",
            "proxy_load_mwh_per_h",
            "recommended_monthly_listing_mwh_per_h",
            "total_flat_after_listing",
            "recommended_monthly_comp_mwh_per_h",
            "monthly_comp_monthly_mwh",
            "july_monthly_market_price",
            "median",
            "p75",
            "p90",
            "recommended_monthly_comp_price",
        ]
    ].copy()
    output_df.columns = [
        "时段",
        "年度双边(MWh/h)",
        "年度挂牌(MWh/h)",
        "年度底仓合计(MWh/h)",
        "代理负荷(MWh/h)",
        "建议月挂牌(MWh/h)",
        "挂牌后总平仓(MWh/h)",
        "建议月竞量(MWh/h)",
        "建议月竞整月量(MWh)",
        "7月月竞市场价",
        "6月日成交中位价",
        "6月日成交75分位价",
        "6月日成交90分位价",
        "建议月竞报价",
    ]

    history_df = strategy_df[
        [
            "hour",
            "july_monthly_market_price",
            "mean",
            "median",
            "p75",
            "p90",
        ]
    ].copy()
    history_df.columns = [
        "时段",
        "7月月竞市场价",
        "6月日成交均价",
        "6月日成交中位价",
        "6月日成交75分位价",
        "6月日成交90分位价",
    ]

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        listing_df.to_excel(writer, sheet_name="总览", index=False)
        output_df.to_excel(writer, sheet_name="24时段申报表", index=False)
        history_df.to_excel(writer, sheet_name="历史成交价参考", index=False)

    wb = openpyxl.load_workbook(xlsx_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 24)

    ws = wb["24时段申报表"]
    for row in range(2, ws.max_row + 1):
        comp_qty = ws[f"H{row}"].value
        comp_price = ws[f"N{row}"].value
        if comp_qty and comp_qty > 0:
            ws[f"H{row}"].fill = PatternFill("solid", fgColor="FFF2CC")
            ws[f"N{row}"].fill = PatternFill("solid", fgColor="F4CCCC")
        if comp_qty and comp_qty >= 2:
            for col in ["H", "I", "N"]:
                ws[f"{col}{row}"].font = Font(bold=True, color="9C0006")
        if comp_price is None:
            ws[f"N{row}"].value = ""

    wb.save(xlsx_path)

    md_text = f"""# 2026年8月挂牌与月竞申报策略

## 核心结论

- 建议月挂牌申报：`{assumptions["monthly_listing_target_mwh_per_h"]:.3f} MWh/h`
- 建议月挂牌整月总量：`{assumptions["monthly_listing_total_mwh"]:.3f} MWh`
- 年度底仓均值：`{assumptions["annual_flat_avg_mwh_per_h"]:.3f} MWh/h`
- 挂牌后总平仓均值：`{assumptions["flat_after_listing_avg_mwh_per_h"]:.3f} MWh/h`
- 用 7 月 D+2 形状代理的负荷均值：`{assumptions["proxy_load_avg_mwh_per_h"]:.3f} MWh/h`
- 建议月竞整月补仓量：`{assumptions["recommended_monthly_comp_total_mwh"]:.3f} MWh`

## 口径

- 8 月负荷曲线：暂用 `7月D+2` 形状代替
- 月竞报价锚点：取 `7月月竞市场价` 与 `6月日成交75分位价` 的较高值，再按缺口大小加价
- 月竞只补形状，不承担主体底仓

## 文件

- 主表：`{xlsx_path}`
"""
    md_path.write_text(md_text, encoding="utf-8")
    return xlsx_path, md_path


def main():
    strategy_df, assumptions = build_strategy()
    xlsx_path, md_path = save_outputs(strategy_df, assumptions)
    print(f"OK: {xlsx_path}")
    print(f"OK: {md_path}")


if __name__ == "__main__":
    main()
